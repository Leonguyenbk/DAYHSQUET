import argparse
import os
import queue
import subprocess
import sys
import threading
import time
import tkinter as tk
from datetime import datetime
from tkinter import filedialog, messagebox, ttk
from tkinter.scrolledtext import ScrolledText

from openpyxl import Workbook, load_workbook

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager


URL = "https://dla.mplis.gov.vn/dc/CungCapThongTinHoSoTiepNhan/v2"
SAVE_INTERVAL = 10


def normalize_path(path: str) -> str:
    return os.path.normpath(os.path.abspath(path))


def safe_console_print(message):
    text = str(message)
    stream = sys.stdout

    try:
        stream.write(text + "\n")
        stream.flush()
        return
    except UnicodeEncodeError:
        pass

    encoding = getattr(stream, "encoding", None) or "utf-8"
    data = (text + "\n").encode(encoding, errors="replace")
    stream.buffer.write(data)
    stream.flush()


def get_login_fields(wait):
    username_box = wait.until(
        EC.presence_of_element_located(
            (
                By.CSS_SELECTOR,
                "input[autocomplete='username'], input[name='username']",
            )
        )
    )
    password_box = wait.until(
        EC.presence_of_element_located(
            (
                By.CSS_SELECTOR,
                "input[autocomplete='current-password'], input[name='password']",
            )
        )
    )
    return username_box, password_box


def wait_ajax(driver, timeout=20):
    end_time = time.time() + timeout
    while time.time() < end_time:
        try:
            is_done = driver.execute_script(
                """
                if (typeof jQuery !== 'undefined') {
                    return jQuery.active === 0;
                }
                return true;
                """
            )
            if is_done:
                return True
        except Exception:
            return True
        time.sleep(0.2)
    return False


def sanitize_filename(name: str) -> str:
    invalid = r'<>:"/\\|?*'
    for ch in invalid:
        name = name.replace(ch, "_")
    return name.strip()


def wait_download_complete(folder: str, before_files: set, timeout=120):
    end_time = time.time() + timeout
    while time.time() < end_time:
        current_files = set(os.listdir(folder))
        new_files = current_files - before_files

        cr_files = [f for f in current_files if f.endswith(".crdownload") or f.endswith(".tmp")]
        if cr_files:
            time.sleep(0.5)
            continue

        completed_files = [
            f for f in new_files
            if not f.endswith(".crdownload") and not f.endswith(".tmp")
        ]
        if completed_files:
            full_paths = [os.path.join(folder, f) for f in completed_files]
            return max(full_paths, key=os.path.getctime)

        time.sleep(0.5)
    return None


def enable_download_behavior(driver, download_dir: str):
    try:
        driver.execute_cdp_cmd(
            "Page.setDownloadBehavior",
            {
                "behavior": "allow",
                "downloadPath": download_dir,
            },
        )
    except Exception:
        pass


def trigger_pdf_download(driver, wait, download_dir: str, before_files: set, timeout=120):
    btn_download = wait.until(EC.presence_of_element_located((By.ID, "download")))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn_download)
    time.sleep(0.5)

    try:
        wait.until(EC.element_to_be_clickable((By.ID, "download")))
        driver.execute_script("arguments[0].click();", btn_download)
    except Exception:
        driver.execute_script("arguments[0].click();", btn_download)

    driver.switch_to.default_content()
    downloaded_file = wait_download_complete(download_dir, before_files, timeout=8)
    if downloaded_file:
        return downloaded_file

    iframe = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#mdlPreviewFileLocal iframe")))
    driver.switch_to.frame(iframe)
    driver.execute_script(
        """
        if (window.PDFViewerApplication && typeof window.PDFViewerApplication.download === 'function') {
            window.PDFViewerApplication.download();
        } else {
            const btn = document.getElementById('download');
            if (btn) btn.click();
        }
        """
    )
    driver.switch_to.default_content()
    return wait_download_complete(download_dir, before_files, timeout=timeout)


def rename_downloaded_file(downloaded_file: str, mahoso: str, target_folder: str, ten_file: str = ""):
    ext = os.path.splitext(downloaded_file)[1] or ".pdf"
    file_label = sanitize_filename(ten_file) if ten_file else ""
    mahoso_label = sanitize_filename(mahoso)
    base = f"{file_label}_{mahoso_label}" if file_label else mahoso_label
    new_path = os.path.join(target_folder, base + ext)

    if os.path.exists(new_path):
        i = 1
        while True:
            candidate = os.path.join(target_folder, f"{base}_{i}{ext}")
            if not os.path.exists(candidate):
                new_path = candidate
                break
            i += 1

    os.rename(downloaded_file, new_path)
    return new_path


def read_mahoso_from_excel(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    result = []
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=1).value
        if value is not None:
            ma = str(value).strip()
            if ma:
                result.append(ma)
    return result


def write_result_excel(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "KetQua"

    ws["A1"] = "STT"
    ws["B1"] = "Mã hồ sơ"
    ws["C1"] = "Kết quả"

    for i, (stt, mahoso, ketqua) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=stt)
        ws.cell(row=i, column=2, value=mahoso)
        ws.cell(row=i, column=3, value=ketqua)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 20
    wb.save(path)


def cho_file_hoac_notify(driver, timeout=10):
    start = time.time()
    while time.time() - start < timeout:
        previews = driver.find_elements(By.ID, "mdlPreviewFileLocal")
        if previews:
            try:
                cls = previews[0].get_attribute("class") or ""
                if previews[0].is_displayed() and "in" in cls.split():
                    return "file"
            except Exception:
                pass

        notifies = driver.find_elements(By.CSS_SELECTOR, "body > div.notifyjs-corner > div")
        if notifies:
            try:
                if notifies[0].is_displayed():
                    return "notify"
            except Exception:
                pass

        time.sleep(0.3)
    return "timeout"


def dong_modal_preview(driver, wait):
    driver.switch_to.default_content()
    preview = wait.until(EC.visibility_of_element_located((By.ID, "mdlPreviewFileLocal")))
    btn_close_preview = preview.find_element(By.CSS_SELECTOR, "div.modal-header > button.close")
    driver.execute_script("arguments[0].click();", btn_close_preview)
    wait.until(
        lambda d: "in" not in (d.find_element(By.ID, "mdlPreviewFileLocal").get_attribute("class") or "").split()
    )


def dong_modal_ds_cho_bo_sung(driver, wait):
    modal_ds = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "[id^='mdlDSChoBoSung'].in")))
    btn_dong_ds = modal_ds.find_element(By.CSS_SELECTOR, ".flex-panel-footer button[data-dismiss='modal']")
    driver.execute_script("arguments[0].click();", btn_dong_ds)
    wait.until(lambda d: len(d.find_elements(By.CSS_SELECTOR, "[id^='mdlDSChoBoSung'].in")) == 0)


def parse_row_datetime(row):
    cells = row.find_elements(By.TAG_NAME, "td")
    for index in (6, 3):
        if len(cells) <= index:
            continue
        text = cells[index].text.strip()
        if not text:
            continue
        for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
    return None


def get_latest_file_button_row(modal):
    rows = modal.find_elements(By.CSS_SELECTOR, "#tbDSChoBoSung tbody tr")
    best_row = None
    best_date = None

    for row in rows:
        buttons = row.find_elements(By.CSS_SELECTOR, "#btnXemFileDungCho")
        if not buttons:
            continue

        row_date = parse_row_datetime(row)
        if best_row is None or (row_date is not None and (best_date is None or row_date > best_date)):
            best_row = row
            best_date = row_date

    return best_row


def process_one_mahoso(driver, wait, mahoso, download_dir):
    input_box = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#frmSearchHoSoTiepNhan input")))
    input_box.click()
    input_box.send_keys(Keys.CONTROL, "a")
    input_box.send_keys(Keys.DELETE)
    input_box.send_keys(mahoso)
    input_box.send_keys(Keys.ENTER)

    selector_ma = "#lstHoSoTiepNhan > ul > li.list-group-item.title > span.value > span"
    wait.until(lambda d: d.find_element(By.CSS_SELECTOR, selector_ma).text.strip() == mahoso)
    wait_ajax(driver, 10)

    try:
        trang_thai_element = driver.find_element(By.CSS_SELECTOR, "#lstHoSoTiepNhan span.value[name='trangThai']")
        if "Đang chờ bổ sung" not in trang_thai_element.text:
            return f"Bỏ qua (Trạng thái: {trang_thai_element.text})"
    except Exception:
        pass

    dong_ket_qua = wait.until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "#lstHoSoTiepNhan > ul > li.list-group-item.title"))
    )
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", dong_ket_qua)
    time.sleep(0.5)
    driver.execute_script("arguments[0].click();", dong_ket_qua)

    wait.until(
        lambda d: d.find_element(By.ID, "wpThongTinChiTiet").get_attribute("class").strip() == "thongtinchitiet-wrapper"
    )
    wait_ajax(driver, 10)

    btn_cho_bo_sung = wait.until(EC.element_to_be_clickable((By.ID, "btnDSChoBoSung")))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn_cho_bo_sung)
    time.sleep(0.5)
    driver.execute_script("arguments[0].click();", btn_cho_bo_sung)

    wait.until(
        lambda d: "in" in (d.find_element(By.CSS_SELECTOR, "[id^='mdlDSChoBoSung']").get_attribute("class") or "").split()
    )
    wait.until(EC.visibility_of_element_located((By.ID, "tbDSChoBoSung")))
    wait_ajax(driver, 10)

    modal = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "[id^='mdlDSChoBoSung'].in")))
    selected_row = get_latest_file_button_row(modal)
    if selected_row is None:
        dong_modal_ds_cho_bo_sung(driver, wait)
        return "Không có file"

    try:
        ten_file = selected_row.find_element(
            By.CSS_SELECTOR,
            "input[name='tenFileNhanBoSung'], input[name='tenFile']",
        ).get_attribute("value").strip()
    except Exception:
        ten_file = ""

    btn_xem_file = selected_row.find_element(By.CSS_SELECTOR, "#btnXemFileDungCho")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn_xem_file)
    time.sleep(0.5)
    driver.execute_script("arguments[0].click();", btn_xem_file)

    ket_qua = cho_file_hoac_notify(driver, timeout=12)

    if ket_qua == "notify":
        dong_modal_ds_cho_bo_sung(driver, wait)
        return "Không có file"

    if ket_qua == "file":
        before_files = set(os.listdir(download_dir))
        wait.until(lambda d: "in" in (d.find_element(By.ID, "mdlPreviewFileLocal").get_attribute("class") or ""))

        iframe = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#mdlPreviewFileLocal iframe")))
        driver.switch_to.frame(iframe)

        wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
        wait.until(EC.visibility_of_element_located((By.ID, "mainContainer")))
        wait.until(EC.visibility_of_element_located((By.ID, "toolbarViewerRight")))
        wait.until(lambda d: len(d.find_elements(By.CSS_SELECTOR, ".page canvas")) > 0)

        downloaded_file = trigger_pdf_download(driver, wait, download_dir, before_files, timeout=120)

        time.sleep(1)
        dong_modal_preview(driver, wait)
        time.sleep(0.5)
        dong_modal_ds_cho_bo_sung(driver, wait)

        if downloaded_file:
            rename_downloaded_file(downloaded_file, mahoso, download_dir, ten_file=ten_file)
            return "Có file"
        return "Không có file"

    try:
        driver.switch_to.default_content()
    except Exception:
        pass
    try:
        dong_modal_preview(driver, wait)
    except Exception:
        pass
    try:
        dong_modal_ds_cho_bo_sung(driver, wait)
    except Exception:
        pass
    return "Không có file"


def run_batch(username, password, input_excel, download_dir, output_excel, log_fn=print):
    if not username or not password:
        raise Exception("Thiếu username hoặc password.")
    if not os.path.isfile(input_excel):
        raise Exception(f"Không tìm thấy file Excel đầu vào: {input_excel}")

    download_dir = normalize_path(download_dir)
    os.makedirs(download_dir, exist_ok=True)
    ma_hoso_list = read_mahoso_from_excel(input_excel)
    if not ma_hoso_list:
        raise Exception("Không có mã hồ sơ nào trong cột A từ dòng 2 trở đi.")

    options = Options()
    options.add_argument("--start-maximized")
    prefs = {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "plugins.always_open_pdf_externally": True,
    }
    options.add_experimental_option("prefs", prefs)

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    enable_download_behavior(driver, download_dir)
    wait = WebDriverWait(driver, 30)

    results = []
    try:
        log_fn("Đăng nhập MPLIS...")
        driver.get(URL)
        username_box, password_box = get_login_fields(wait)
        username_box.send_keys(username)
        password_box.send_keys(password)
        password_box.send_keys(Keys.ENTER)
        wait.until(EC.presence_of_element_located((By.ID, "frmSearchHoSoTiepNhan")))
        wait_ajax(driver, 20)
        log_fn("Đăng nhập xong.")

        total = len(ma_hoso_list)
        for idx, mahoso in enumerate(ma_hoso_list, start=1):
            log_fn(f"[{idx}/{total}] Xử lý hồ sơ: {mahoso}")
            try:
                result = process_one_mahoso(driver, wait, mahoso, download_dir)
            except Exception as e:
                log_fn(f"Lỗi hồ sơ {mahoso}: {e}")
                result = "Không có file"
                try:
                    driver.switch_to.default_content()
                except Exception:
                    pass

            results.append((idx, mahoso, result if result == "Có file" else "Không có file"))
            if idx % SAVE_INTERVAL == 0:
                write_result_excel(output_excel, results)
                log_fn(f"Đã lưu tạm Excel sau {idx} hồ sơ: {output_excel}")

        write_result_excel(output_excel, results)
        log_fn(f"Hoàn thành. Đã ghi kết quả: {output_excel}")
    finally:
        driver.quit()


class LauncherApp:
    def __init__(self, root):
        self.root = root
        self.root.title("MPLIS Launcher (GUI -> CLI)")
        self.root.geometry("860x620")

        self.proc = None
        self.log_queue = queue.Queue()

        self.username_var = tk.StringVar()
        self.password_var = tk.StringVar()
        self.input_var = tk.StringVar()
        self.folder_var = tk.StringVar()
        self.output_var = tk.StringVar()

        self.build_ui()
        self.root.after(100, self.poll_log)

    def build_ui(self):
        main = ttk.Frame(self.root, padding=12)
        main.pack(fill="both", expand=True)

        ttk.Label(main, text="Username:").grid(row=0, column=0, sticky="w", pady=6)
        ttk.Entry(main, textvariable=self.username_var, width=50).grid(row=0, column=1, sticky="ew", padx=6, pady=6)

        ttk.Label(main, text="Password:").grid(row=1, column=0, sticky="w", pady=6)
        ttk.Entry(main, textvariable=self.password_var, width=50, show="*").grid(row=1, column=1, sticky="ew", padx=6, pady=6)

        ttk.Label(main, text="Excel input:").grid(row=2, column=0, sticky="w", pady=6)
        ttk.Entry(main, textvariable=self.input_var, width=70).grid(row=2, column=1, sticky="ew", padx=6, pady=6)
        ttk.Button(main, text="Chọn...", command=self.pick_input).grid(row=2, column=2, pady=6)

        ttk.Label(main, text="Folder tải:").grid(row=3, column=0, sticky="w", pady=6)
        ttk.Entry(main, textvariable=self.folder_var, width=70).grid(row=3, column=1, sticky="ew", padx=6, pady=6)
        ttk.Button(main, text="Chọn...", command=self.pick_folder).grid(row=3, column=2, pady=6)

        ttk.Label(main, text="Excel kết quả:").grid(row=4, column=0, sticky="w", pady=6)
        ttk.Entry(main, textvariable=self.output_var, width=70).grid(row=4, column=1, sticky="ew", padx=6, pady=6)
        ttk.Button(main, text="Chọn...", command=self.pick_output).grid(row=4, column=2, pady=6)

        row = ttk.Frame(main)
        row.grid(row=5, column=1, sticky="ew", padx=6, pady=10)
        self.start_btn = ttk.Button(row, text="Bắt đầu", command=self.start)
        self.start_btn.pack(side="left")
        self.status_label = ttk.Label(row, text="Sẵn sàng")
        self.status_label.pack(side="right")

        self.log_box = ScrolledText(main, height=24)
        self.log_box.grid(row=6, column=0, columnspan=3, sticky="nsew", pady=(8, 0))

        main.columnconfigure(1, weight=1)
        main.rowconfigure(6, weight=1)

    def pick_input(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xlsm")])
        if path:
            self.input_var.set(path)

    def pick_folder(self):
        path = filedialog.askdirectory()
        if path:
            self.folder_var.set(path)

    def pick_output(self):
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.output_var.set(path)

    def append_log(self, text):
        self.log_box.insert("end", text + "\n")
        self.log_box.see("end")

    def start(self):
        if self.proc is not None and self.proc.poll() is None:
            return

        username = self.username_var.get().strip()
        password = self.password_var.get().strip()
        input_excel = self.input_var.get().strip()
        download_dir = self.folder_var.get().strip()
        output_excel = self.output_var.get().strip()

        if not username or not password or not input_excel or not download_dir or not output_excel:
            messagebox.showerror("Lỗi", "Vui lòng nhập/chọn đầy đủ thông tin.")
            return

        self.log_box.delete("1.0", "end")
        self.status_label.config(text="Đang chạy...")
        self.start_btn.config(state="disabled")

        cmd = [
            sys.executable,
            os.path.abspath(__file__),
            "--run",
            "--username", username,
            "--password", password,
            "--input", input_excel,
            "--download", download_dir,
            "--output", output_excel,
        ]

        self.proc = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
        )

        def reader():
            assert self.proc and self.proc.stdout
            for line in self.proc.stdout:
                self.log_queue.put(line.rstrip("\n"))
            rc = self.proc.wait()
            self.log_queue.put(f"__EXIT__:{rc}")

        threading.Thread(target=reader, daemon=True).start()

    def poll_log(self):
        try:
            while True:
                item = self.log_queue.get_nowait()
                if item.startswith("__EXIT__:"):
                    code = int(item.split(":", 1)[1])
                    self.start_btn.config(state="normal")
                    self.status_label.config(text="Hoàn thành" if code == 0 else "Có lỗi")
                    if code == 0:
                        messagebox.showinfo("Xong", "Đã xử lý xong toàn bộ mã hồ sơ.")
                    else:
                        messagebox.showerror("Lỗi", "Job kết thúc với lỗi. Xem log để biết chi tiết.")
                else:
                    self.append_log(item)
        except queue.Empty:
            pass
        self.root.after(100, self.poll_log)


def cli_main(args):
    if hasattr(sys.stdout, "reconfigure"):
        try:
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass

    def log(msg):
        safe_console_print(msg)

    run_batch(args.username, args.password, args.input, args.download, args.output, log_fn=log)


def gui_main():
    root = tk.Tk()
    app = LauncherApp(root)
    root.mainloop()


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--run", action="store_true", help="Run batch in CLI mode")
    parser.add_argument("--username")
    parser.add_argument("--password")
    parser.add_argument("--input")
    parser.add_argument("--download")
    parser.add_argument("--output")
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    if args.run:
        cli_main(args)
    else:
        gui_main()
