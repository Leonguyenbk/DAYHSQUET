import os
import sys
import json
import re
import queue
import threading
import traceback
import time
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime, timezone
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from openpyxl import load_workbook, Workbook

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys


# =========================
# API CONFIG
# =========================

REFERER_URL = "https://dla.mplis.gov.vn/dc/DonDangKy/KeKhaiDangKyV2"

API_ADVANCED_SEARCH      = "https://dla.mplis.gov.vn/dc/DangKyAjax/AdvancedSearchTinhHinhDangKy"
URL_GET_THONG_TIN_DANGKY = "https://dla.mplis.gov.vn/dc/DangKyAjax/GetThongTinDangKyByTinhHinhDangKyIds"
URL_DELETE_DON_DANGKY    = "https://dla.mplis.gov.vn/dc/DangKyAjax/DeleteDonDangKyByTinhHinhDangKyId"

# True  = chỉ kiểm tra, KHÔNG xóa thật (in ra cái nào sẽ bị xóa)
# False = XÓA THẬT
# Mặc định True cho an toàn vì xóa là không hồi lại được.
DRY_RUN = True

# Số luồng xử lý song song
MAX_WORKERS = 3
API_SEARCH_TIMEOUT = 120
API_DELETE_TIMEOUT = 180
API_SEARCH_RETRIES = 3
API_RETRY_BACKOFF_SECONDS = 5

# Ghi Excel sau mỗi N dòng kết quả
WRITE_EVERY_N = 10


# =========================
# LOG REDIRECT (thread-safe)
# =========================

class ThreadSafeLogger:
    """Logger thread-safe, dùng chung cho nhiều worker."""
    def __init__(self, log_queue):
        self.log_queue = log_queue
        self._lock = threading.Lock()

    def log(self, msg):
        if msg and str(msg).strip():
            with self._lock:
                self.log_queue.put(str(msg))

    def write(self, text):
        self.log(text)

    def flush(self):
        pass


# =========================
# HELPER
# =========================

def safe_int(value, default=0):
    try:
        if value is None or value == "":
            return default
        return int(value)
    except Exception:
        return default


def now_iso_z():
    return datetime.now(timezone.utc).isoformat(timespec="milliseconds").replace("+00:00", "Z")


def find_key_recursive(obj, target_key):
    """
    Tìm đệ quy giá trị của key 'target_key' đầu tiên trong dict/list.
    Trả về giá trị (có thể là [], {}...) hoặc None nếu không thấy.
    """
    if isinstance(obj, dict):
        if target_key in obj:
            return obj[target_key]
        for v in obj.values():
            found = find_key_recursive(v, target_key)
            if found is not None:
                return found
    elif isinstance(obj, list):
        for x in obj:
            found = find_key_recursive(x, target_key)
            if found is not None:
                return found
    return None


# =========================
# EXCEL
# =========================

def doc_excel(path_excel):
    """Đọc Excel, chỉ cần 2 cột: soto | sothua."""
    wb = load_workbook(path_excel, data_only=True)
    ws = wb.active

    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value:
            headers[str(value).strip().lower()] = col

    required_cols = ["soto", "sothua"]
    missing = [c for c in required_cols if c not in headers]
    if missing:
        raise RuntimeError(f"Thiếu cột trong Excel: {', '.join(missing)} (cần: soto | sothua)")

    data = []
    for row in range(2, ws.max_row + 1):
        soto   = ws.cell(row=row, column=headers["soto"]).value
        sothua = ws.cell(row=row, column=headers["sothua"]).value

        if not soto and not sothua:
            continue

        data.append({
            "row":    row,
            "soto":   str(soto).strip()   if soto   is not None else "",
            "sothua": str(sothua).strip() if sothua is not None else "",
        })

    return data


def ghi_excel_output(rows, output_path, lock):
    """Ghi kết quả ra Excel, dùng lock để tránh ghi đồng thời."""
    with lock:
        wb = Workbook()
        ws = wb.active
        ws.title = "KetQua"

        headers = [
            "STT", "Dòng Excel", "Số tờ", "Số thửa",
            "tinhHinhDangKyId", "Chủ sử dụng",
            "Số GCN", "soPhatHanh", "Hành động", "Trạng thái", "Ghi chú"
        ]
        ws.append(headers)

        sorted_rows = sorted(
            rows,
            key=lambda x: (x.get("row_excel") or 0, str(x.get("tinhHinhDangKyId") or ""))
        )

        for i, r in enumerate(sorted_rows, start=1):
            ws.append([
                i,
                r.get("row_excel"),
                r.get("soto"),
                r.get("sothua"),
                r.get("tinhHinhDangKyId"),
                r.get("chu_su_dung"),
                r.get("so_gcn"),
                r.get("so_phat_hanh"),
                r.get("hanh_dong"),
                r.get("status"),
                r.get("note"),
            ])

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

        wb.save(output_path)


# =========================
# SELENIUM LOGIN + TOKEN
# =========================

def lay_token_tu_trang(driver):
    js = """
    return (
        document.querySelector('input[name="__RequestVerificationToken"]')?.value ||
        document.querySelector('input[name="__requestverificationtoken"]')?.value ||
        document.querySelector('meta[name="__RequestVerificationToken"]')?.content ||
        document.querySelector('meta[name="__requestverificationtoken"]')?.content ||
        document.querySelector('meta[name="RequestVerificationToken"]')?.content ||
        ''
    );
    """
    return driver.execute_script(js)


def tao_session_tu_selenium(driver):
    session = requests.Session()

    user_agent = driver.execute_script("return navigator.userAgent;")
    token = lay_token_tu_trang(driver)

    if not token:
        raise RuntimeError("Không lấy được __requestverificationtoken.")

    session.headers.update({
        "User-Agent": user_agent,
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "X-Requested-With": "XMLHttpRequest",
        "Origin": "https://dla.mplis.gov.vn",
        "Referer": REFERER_URL,
        "__requestverificationtoken": token,
        "__RequestVerificationToken": token,
        "RequestVerificationToken": token,
    })

    for c in driver.get_cookies():
        session.cookies.set(
            name=c["name"],
            value=c["value"],
            domain=c.get("domain"),
            path=c.get("path", "/")
        )

    return session


def login_mplis(username, password):
    options = Options()
    options.add_argument("--start-maximized")
    options.add_argument("--window-position=100,100")
    options.add_argument("--window-size=1400,900")

    driver = webdriver.Chrome(options=options)
    driver.get(REFERER_URL)

    try:
        time.sleep(2)

        inputs = driver.find_elements(By.CSS_SELECTOR, "input")
        user_box = None
        pass_box = None

        for inp in inputs:
            typ         = (inp.get_attribute("type")        or "").lower()
            name        = (inp.get_attribute("name")        or "").lower()
            placeholder = (inp.get_attribute("placeholder") or "").lower()
            input_id    = (inp.get_attribute("id")          or "").lower()
            text_all    = f"{name} {placeholder} {input_id}"

            if not user_box and typ in ["text", "email"] and any(
                k in text_all for k in ["user", "username", "login", "account", "ten"]
            ):
                user_box = inp

            if not pass_box and typ == "password":
                pass_box = inp

        if not user_box:
            for inp in inputs:
                typ = (inp.get_attribute("type") or "").lower()
                if typ in ["text", "email"]:
                    user_box = inp
                    break

        if user_box and pass_box:
            user_box.clear()
            user_box.send_keys(username)
            pass_box.clear()
            pass_box.send_keys(password)
            pass_box.send_keys(Keys.ENTER)
        else:
            pass  # người dùng tự đăng nhập tay

    except Exception:
        pass

    messagebox.showinfo(
        "Đăng nhập / Authenticator",
        "Hoàn tất đăng nhập MPLIS và xác thực Authenticator trên Chrome.\n"
        "Sau khi vào được màn hình kê khai đăng ký thì bấm OK để tiếp tục."
    )

    return driver


# =========================
# HTTP RETRY
# =========================

def post_retry(session, url, *, retries, timeout, **kwargs):
    last_error = None
    for attempt in range(1, retries + 1):
        try:
            return session.post(url, timeout=timeout, **kwargs), None
        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
            last_error = e
            if attempt < retries:
                time.sleep(API_RETRY_BACKOFF_SECONDS * attempt)
        except requests.exceptions.RequestException as e:
            return None, f"Loi request API: {e}"
    return None, f"Timeout/loi mang sau {retries} lan thu: {last_error}"


# =========================
# API 1: ADVANCED SEARCH TÌNH HÌNH ĐĂNG KÝ -> lấy danh sách tinhHinhDangKyId
# =========================

def tao_payload_advanced_search(xa_id, so_to, so_thua):
    """
    Payload tìm tình hình đăng ký theo tờ/thửa.
    Lấy đúng từ F12 (Form Data) của AdvancedSearchTinhHinhDangKy — dùng prefix model[...].
      - model[soHieuToBanDo] = số tờ
      - model[soThuTuThua]   = số thửa
      - model[xaId]          = mã xã
    """
    return {
        "draw": "4",

        "columns[0][data]": "",                 "columns[0][name]": "",
        "columns[0][searchable]": "true",       "columns[0][orderable]": "false",
        "columns[0][search][value]": "",        "columns[0][search][regex]": "false",

        "columns[1][data]": "tinhHinhDangKyId", "columns[1][name]": "tinhHinhDangKyId",
        "columns[1][searchable]": "true",       "columns[1][orderable]": "true",
        "columns[1][search][value]": "",        "columns[1][search][regex]": "false",

        "columns[2][data]": "maDon",            "columns[2][name]": "maDon",
        "columns[2][searchable]": "true",       "columns[2][orderable]": "true",
        "columns[2][search][value]": "",        "columns[2][search][regex]": "false",

        "columns[3][data]": "soThuTu",          "columns[3][name]": "soThuTu",
        "columns[3][searchable]": "true",       "columns[3][orderable]": "true",
        "columns[3][search][value]": "",        "columns[3][search][regex]": "false",

        "columns[4][data]": "DaiDienKhaiTrinh", "columns[4][name]": "DaiDienKhaiTrinh",
        "columns[4][searchable]": "true",       "columns[4][orderable]": "false",
        "columns[4][search][value]": "",        "columns[4][search][regex]": "false",

        "columns[5][data]": "ngayTiepNhan",     "columns[5][name]": "ngayTiepNhan",
        "columns[5][searchable]": "true",       "columns[5][orderable]": "true",
        "columns[5][search][value]": "",        "columns[5][search][regex]": "false",

        "columns[6][data]": "thoiDiemDangKy",   "columns[6][name]": "thoiDiemDangKy",
        "columns[6][searchable]": "true",       "columns[6][orderable]": "true",
        "columns[6][search][value]": "",        "columns[6][search][regex]": "false",

        "order[0][column]": "5",
        "order[0][dir]": "desc",

        # length lớn để lấy hết bản ghi của 1 tờ/thửa
        "start": "0", "length": "100",
        "search[value]": "", "search[regex]": "false",

        # ── model[...] ──
        "model[xaId]":               str(xa_id),
        "model[huyenId]":            "",
        "model[tinhHinhDangKyId]":   "",
        "model[maDon]":              "",
        "model[soThuTu]":            "",
        "model[ngayTiepNhan]":       "",
        "model[thoiDiemDangKy]":     "",
        "model[loaiGiayChungNhanId]":"",
        "model[soPhatHanh]":         "",
        "model[maVach]":             "",
        "model[soVaoSo]":            "",
        "model[soVaoSoCu]":          "",
        "model[ngayVaoSo]":          "",
        "model[soHoSoGoc]":          "",
        "model[soHoSoGocCu]":        "",
        "model[hoTen]":              "",
        "model[soGiayTo]":           "",
        "model[namSinh]":            "",

        "model[soThuTuThua]":        str(so_thua),
        "model[soHieuToBanDo]":      str(so_to),

        "model[soThuTuThuaCu]":      "",
        "model[soHieuToBanDoCu]":    "",
        "model[soNha]":              "",
        "model[diaChiChiTiet]":      "",
        "model[dieuKienCapGiay]":    "",
        "model[phucHoiDuLieu]":      "false",
    }


def api_advanced_search(session, xa_id, so_to, so_thua, logger=None):
    headers = dict(session.headers)
    headers["Content-Type"] = "application/x-www-form-urlencoded; charset=UTF-8"

    res, request_error = post_retry(
        session,
        API_ADVANCED_SEARCH,
        retries=API_SEARCH_RETRIES,
        timeout=API_SEARCH_TIMEOUT,
        data=tao_payload_advanced_search(xa_id, so_to, so_thua),
        headers=headers
    )
    if request_error:
        return {"ok": False, "error": request_error}

    try:
        result = res.json()
    except Exception:
        return {
            "ok": False,
            "error": f"Response AdvancedSearch không phải JSON: {res.text[:500]}",
            "status_code": res.status_code
        }

    # Một số endpoint trả thẳng data dạng list/dict không có 'success'
    if isinstance(result, dict) and ("success" in result) and (not result.get("success")):
        return {"ok": False, "error": str(result)[:800], "raw": result}

    ids = extract_all_tinh_hinh_ids(result)

    if logger:
        logger.log(f"   AdvancedSearch tờ={so_to} thửa={so_thua} -> ids={ids}")

    if not ids:
        return {
            "ok": False,
            "error": "Không tìm thấy tinhHinhDangKyId nào theo tờ/thửa",
            "raw": result
        }

    return {"ok": True, "ids": ids, "raw": result}


def extract_all_tinh_hinh_ids(raw):
    """
    Lấy toàn bộ tinhHinhDangKyId trong response (xử lý nhiều bản ghi).
    Quét cả các key data / Data / value nếu có, fallback quét đệ quy.
    """
    ids = []

    candidates = []
    if isinstance(raw, dict):
        for key in ("data", "Data", "value", "Value", "items", "Items"):
            v = raw.get(key)
            if isinstance(v, list):
                candidates = v
                break
        if not candidates and isinstance(raw, dict):
            candidates = [raw]
    elif isinstance(raw, list):
        candidates = raw

    for item in candidates:
        tid = find_key_recursive(item, "tinhHinhDangKyId")
        if tid:
            ids.append(safe_int(tid))

    # fallback: nếu chưa thấy gì, quét toàn bộ
    if not ids:
        def collect(obj):
            if isinstance(obj, dict):
                for k, v in obj.items():
                    if k == "tinhHinhDangKyId" and v:
                        ids.append(safe_int(v))
                    else:
                        collect(v)
            elif isinstance(obj, list):
                for x in obj:
                    collect(x)
        collect(raw)

    # dedupe giữ thứ tự
    seen = set()
    out = []
    for i in ids:
        if i and i not in seen:
            seen.add(i)
            out.append(i)
    return out


# =========================
# API 2: GET THÔNG TIN ĐĂNG KÝ theo 1 id
# =========================

def lay_thong_tin_dang_ky_by_id(session, tinh_hinh_dang_ky_id):
    """Lấy thông tin đăng ký của 1 tinhHinhDangKyId."""
    tid = safe_int(tinh_hinh_dang_ky_id, 0)
    if not tid:
        return {"ok": False, "error": f"tinhHinhDangKyId không hợp lệ: {tinh_hinh_dang_ky_id}"}

    headers = dict(session.headers)
    headers.update({
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Content-Type": "application/json; charset=UTF-8",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": REFERER_URL,
        "Origin": "https://dla.mplis.gov.vn",
    })

    payload = {"tinhHinhDangKyIds": [tid], "getHoSoQuet": False}

    res, request_error = post_retry(
        session,
        URL_GET_THONG_TIN_DANGKY,
        retries=API_SEARCH_RETRIES,
        timeout=API_SEARCH_TIMEOUT,
        data=json.dumps(payload, ensure_ascii=False),
        headers=headers
    )
    if request_error:
        return {"ok": False, "error": request_error}

    try:
        js = res.json()
    except Exception:
        return {
            "ok": False,
            "error": f"Response GetThongTinDangKy không phải JSON: {res.text[:500]}",
            "status_code": res.status_code
        }

    if isinstance(js, dict) and ("success" in js) and (not js.get("success")):
        return {"ok": False, "error": f"GetThongTinDangKy success=false id={tid}. {str(js)[:500]}", "raw": js}

    value = js.get("value") if isinstance(js, dict) else None
    if not value:
        return {"ok": False, "error": f"Không lấy được value từ id={tid}. {str(js)[:500]}", "raw": js}

    return {"ok": True, "raw": js, "value0": value[0]}


# =========================
# PARSE: GIẤY CHỨNG NHẬN + CHỦ
# =========================

def lay_chu_su_dung(value0):
    """Lấy tên chủ sử dụng từ value[0] (cá nhân hoặc tổ chức)."""
    try:
        chu = find_key_recursive(value0, "ChuSoHuu") or {}
        ca_nhans = chu.get("CaNhans") or []
        if ca_nhans:
            return ca_nhans[0].get("hoTen") or ""
        to_chucs = chu.get("ToChucs") or []
        if to_chucs:
            return to_chucs[0].get("tenToChuc") or to_chucs[0].get("ten") or ""
    except Exception:
        pass
    return ""


def danh_gia_xoa(value0):
    """
    Quyết định có nên xóa hay không dựa trên ListGiayChungNhan.

    Trả về: (nen_xoa, so_gcn, so_phat_hanh_str, ly_do)
      - nen_xoa = True nếu ListGiayChungNhan trống
                  HOẶC không có GCN nào có soPhatHanh hợp lệ.
    """
    list_gcn = find_key_recursive(value0, "ListGiayChungNhan")

    if list_gcn is None:
        list_gcn = []

    if not isinstance(list_gcn, list):
        # cấu trúc lạ -> coi như không có GCN hợp lệ
        return True, 0, "", "ListGiayChungNhan không phải list (coi như trống)"

    so_gcn = len(list_gcn)

    if so_gcn == 0:
        return True, 0, "", "ListGiayChungNhan trống"

    sphs = []
    co_sph_hop_le = False
    for gcn in list_gcn:
        sph = ""
        if isinstance(gcn, dict):
            raw_sph = (
                gcn.get("soPhatHanh")
                or gcn.get("SoPhatHanh")
                or gcn.get("sophathanh")
            )
            sph = str(raw_sph).strip() if raw_sph is not None else ""
        sphs.append(sph)
        if sph:
            co_sph_hop_le = True

    so_phat_hanh_str = " | ".join(s if s else "(trống)" for s in sphs)

    if not co_sph_hop_le:
        return True, so_gcn, so_phat_hanh_str, "soPhatHanh trống ở tất cả GCN"

    return False, so_gcn, so_phat_hanh_str, "Có GCN với soPhatHanh hợp lệ"


# =========================
# API 3: DELETE ĐƠN ĐĂNG KÝ
# =========================

def api_delete_don_dang_ky(session, tinh_hinh_dang_ky_id):
    tid = safe_int(tinh_hinh_dang_ky_id, 0)
    if not tid:
        return {"ok": False, "error": f"tinhHinhDangKyId không hợp lệ: {tinh_hinh_dang_ky_id}"}

    headers = dict(session.headers)
    headers.update({
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Content-Type": "application/json; charset=UTF-8",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": REFERER_URL,
        "Origin": "https://dla.mplis.gov.vn",
    })

    payload = {"tinhHinhDangKyId": tid}

    try:
        res = session.post(
            URL_DELETE_DON_DANGKY,
            data=json.dumps(payload, ensure_ascii=False),
            headers=headers,
            timeout=API_DELETE_TIMEOUT
        )
    except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
        return {"ok": False, "error": f"Timeout/lỗi mạng khi Delete: {e}"}
    except requests.exceptions.RequestException as e:
        return {"ok": False, "error": f"Lỗi request Delete: {e}"}

    print("=" * 80)
    print("DELETE ĐƠN ĐĂNG KÝ")
    print("ID:", tid)
    print("STATUS:", res.status_code)
    print("TEXT:", res.text[:500])

    try:
        js = res.json()
    except Exception:
        js = None

    if not res.ok:
        return {"ok": False, "error": f"HTTP {res.status_code}: {res.text[:500]}"}

    # Coi là OK nếu success=true, hoặc không có cờ lỗi rõ ràng
    if isinstance(js, dict):
        if js.get("success") is True or js.get("Success") is True:
            return {"ok": True, "raw": js}
        if ("success" not in js) and ("error" not in js) and ("Error" not in js):
            return {"ok": True, "raw": js}
        return {"ok": False, "error": f"Delete trả về: {str(js)[:500]}", "raw": js}

    # Không parse được JSON nhưng HTTP 200
    return {"ok": True, "raw": res.text[:500]}


# =========================
# XỬ LÝ 1 DÒNG EXCEL (1 tờ/thửa -> nhiều id)
# =========================

def xu_ly_1_dong(session, item, maxa, logger, progress_cb=None):
    """
    Xử lý 1 dòng Excel (tờ/thửa). Trả về LIST các result_row,
    mỗi tinhHinhDangKyId là 1 result_row.
    """
    row_excel = item["row"]
    soto      = item["soto"]
    sothua    = item["sothua"]

    def progress(percent, text):
        if progress_cb:
            progress_cb(percent, text)

    def make_row(tid="", chu="", so_gcn="", sph="", hanh_dong="", status="Lỗi", note=""):
        return {
            "row_excel":        row_excel,
            "soto":             soto,
            "sothua":           sothua,
            "tinhHinhDangKyId": tid,
            "chu_su_dung":      chu,
            "so_gcn":           so_gcn,
            "so_phat_hanh":     sph,
            "hanh_dong":        hanh_dong,
            "status":           status,
            "note":             note,
        }

    progress(0, "Bắt đầu")

    # 1) Advanced search lấy danh sách id
    progress(15, "Đang search tình hình ĐK")
    res_search = api_advanced_search(session, xa_id=maxa, so_to=soto, so_thua=sothua, logger=logger)
    if not res_search.get("ok"):
        progress(100, "Search lỗi")
        logger.log(f"❌ Dòng {row_excel} (tờ {soto}/thửa {sothua}): {res_search.get('error','')}")
        return [make_row(status="Lỗi", note="Search lỗi: " + res_search.get("error", ""))]

    ids = res_search["ids"]
    logger.log(f"🔎 Dòng {row_excel} | tờ={soto} thửa={sothua} | tìm thấy {len(ids)} id: {ids}")

    results = []
    n = len(ids)

    for idx, tid in enumerate(ids, start=1):
        base_pct = 15 + int((idx - 1) / max(n, 1) * 75)
        progress(base_pct, f"GCN id {tid} ({idx}/{n})")

        # 2) Lấy thông tin đăng ký
        res_get = lay_thong_tin_dang_ky_by_id(session, tid)
        if not res_get.get("ok"):
            logger.log(f"❌ id={tid}: GET lỗi: {res_get.get('error','')}")
            results.append(make_row(
                tid=tid, status="Lỗi",
                note="GET thông tin ĐK lỗi: " + res_get.get("error", "")
            ))
            continue

        value0 = res_get["value0"]
        chu = lay_chu_su_dung(value0)

        nen_xoa, so_gcn, sph_str, ly_do = danh_gia_xoa(value0)

        if not nen_xoa:
            logger.log(f"⏭️ id={tid} | chủ={chu} | GIỮ ({ly_do})")
            results.append(make_row(
                tid=tid, chu=chu, so_gcn=so_gcn, sph=sph_str,
                hanh_dong="Giữ", status="Giữ", note=ly_do
            ))
            continue

        # 3) Cần xóa
        if DRY_RUN:
            logger.log(f"🧪 id={tid} | chủ={chu} | DRY_RUN -> SẼ XÓA ({ly_do})")
            results.append(make_row(
                tid=tid, chu=chu, so_gcn=so_gcn, sph=sph_str,
                hanh_dong="Sẽ xóa", status="DRY_RUN", note="DRY_RUN — " + ly_do
            ))
            continue

        res_del = api_delete_don_dang_ky(session, tid)
        if res_del.get("ok"):
            logger.log(f"✅ id={tid} | chủ={chu} | ĐÃ XÓA ({ly_do})")
            results.append(make_row(
                tid=tid, chu=chu, so_gcn=so_gcn, sph=sph_str,
                hanh_dong="Đã xóa", status="Đã xóa", note=ly_do
            ))
        else:
            logger.log(f"❌ id={tid} | chủ={chu} | XÓA LỖI: {res_del.get('error','')}")
            results.append(make_row(
                tid=tid, chu=chu, so_gcn=so_gcn, sph=sph_str,
                hanh_dong="Xóa lỗi", status="Lỗi",
                note="Xóa lỗi: " + res_del.get("error", "")
            ))

    progress(100, "Hoàn thành")
    return results


# =========================
# WORKER (chạy trong thread riêng)
# =========================

def worker_run(username, password, maxa, excel_path, log_queue):
    logger = ThreadSafeLogger(log_queue)
    driver = None

    try:
        data = doc_excel(excel_path)
        if not data:
            logger.log("❌ Excel không có dữ liệu.")
            return

        tong = len(data)
        logger.log(f"✅ Đọc Excel xong: {tong} dòng tờ/thửa.")
        logger.log(f"DRY_RUN={DRY_RUN} | MAX_WORKERS={MAX_WORKERS} | mã xã={maxa}")
        logger.log("API search:  " + API_ADVANCED_SEARCH)
        logger.log("API get:     " + URL_GET_THONG_TIN_DANGKY)
        logger.log("API delete:  " + URL_DELETE_DON_DANGKY)
        if DRY_RUN:
            logger.log("⚠️ Đang ở chế độ DRY_RUN — KHÔNG xóa thật. Bỏ tick DRY RUN để xóa thật.")

        output_path = os.path.join(
            os.path.dirname(excel_path),
            "ket_qua_xoa_don_dang_ky.xlsx"
        )

        driver  = login_mplis(username, password)
        base_session = tao_session_tu_selenium(driver)
        logger.log("✅ Đã tạo session API.")

        thread_local = threading.local()

        def get_thread_session():
            if not hasattr(thread_local, "session"):
                s = requests.Session()
                s.headers.update(base_session.headers)
                s.cookies.update(base_session.cookies)
                thread_local.session = s
            return thread_local.session

        log_queue.put({"type": "init", "total": tong, "workers": MAX_WORKERS})

        results       = []
        results_lock  = threading.Lock()
        counter_lock  = threading.Lock()
        da_xoa        = 0
        giu_lai       = 0
        loi           = 0
        done_count    = 0

        worker_id_map  = {}
        worker_id_lock = threading.Lock()
        worker_counter = [0]

        def get_worker_idx():
            tid = threading.get_ident()
            with worker_id_lock:
                if tid not in worker_id_map:
                    worker_counter[0] += 1
                    worker_id_map[tid] = worker_counter[0]
                return worker_id_map[tid]

        def process_item(item):
            nonlocal da_xoa, giu_lai, loi, done_count

            w_idx = get_worker_idx()
            row_excel = item["row"]

            def row_progress(percent, text):
                log_queue.put({
                    "type": "worker_row_progress",
                    "worker": w_idx, "row": row_excel,
                    "percent": percent, "text": text
                })

            try:
                kq_list = xu_ly_1_dong(
                    session=get_thread_session(),
                    item=item, maxa=maxa,
                    logger=logger, progress_cb=row_progress
                )
            except Exception as e:
                traceback.print_exc()
                kq_list = [{
                    "row_excel": item.get("row"),
                    "soto": item.get("soto"), "sothua": item.get("sothua"),
                    "tinhHinhDangKyId": "", "chu_su_dung": "",
                    "so_gcn": "", "so_phat_hanh": "", "hanh_dong": "",
                    "status": "Lỗi ngoài", "note": str(e)
                }]

            with results_lock:
                results.extend(kq_list)

            with counter_lock:
                done_count += 1
                for kq in kq_list:
                    st = kq.get("status", "")
                    if st in ("Đã xóa", "DRY_RUN"):
                        da_xoa += 1
                    elif st == "Giữ":
                        giu_lai += 1
                    else:
                        loi += 1

                log_queue.put({
                    "type": "progress",
                    "done": done_count, "total": tong,
                    "da_xoa": da_xoa, "giu_lai": giu_lai, "loi": loi,
                })

                if done_count % WRITE_EVERY_N == 0 or done_count == tong:
                    try:
                        ghi_excel_output(results, output_path, results_lock)
                        logger.log(f"💾 Đã ghi tạm Excel ({done_count}/{tong} dòng)")
                    except Exception as ex:
                        logger.log(f"⚠️ Lỗi ghi Excel: {ex}")

            return kq_list

        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {executor.submit(process_item, item): item for item in data}
            for future in as_completed(futures):
                try:
                    future.result()
                except Exception as e:
                    logger.log(f"❌ Future exception: {e}")

        # Ghi lần cuối chắc chắn
        try:
            ghi_excel_output(results, output_path, results_lock)
        except Exception as ex:
            logger.log(f"⚠️ Lỗi ghi Excel cuối: {ex}")

        logger.log("=" * 90)
        nhan = "Sẽ xóa (DRY_RUN)" if DRY_RUN else "Đã xóa"
        logger.log(f"🎯 XONG. {nhan}: {da_xoa} | Giữ: {giu_lai} | Lỗi: {loi}")
        logger.log("📄 File kết quả: " + output_path)

    except Exception as e:
        logger.log("❌ Lỗi chương trình: " + str(e))
        traceback.print_exc()

    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass
        log_queue.put("__DONE__")


# =========================
# TKINTER APP
# =========================

class App(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Xóa đơn đăng ký không có GCN — MPLIS (parallel)")
        self.geometry("1080x780")

        self.log_queue     = queue.Queue()
        self.worker_thread = None

        self.var_username  = tk.StringVar()
        self.var_password  = tk.StringVar()
        self.var_maxa      = tk.StringVar()
        self.var_excel     = tk.StringVar()
        self.var_workers   = tk.IntVar(value=MAX_WORKERS)
        self.var_dry_run   = tk.BooleanVar(value=DRY_RUN)

        self._total       = 0
        self._pb_total    = None
        self._lbl_total   = None
        self._worker_rows = []

        self.create_widgets()
        self.after(200, self.process_log_queue)

    def create_widgets(self):
        frame_top = ttk.LabelFrame(self, text="Thông tin chạy")
        frame_top.pack(fill="x", padx=10, pady=(10, 4))

        ttk.Label(frame_top, text="Username").grid(row=0, column=0, padx=5, pady=4, sticky="w")
        ttk.Entry(frame_top, textvariable=self.var_username, width=30).grid(row=0, column=1, padx=5, pady=4)
        ttk.Label(frame_top, text="Password").grid(row=0, column=2, padx=5, pady=4, sticky="w")
        ttk.Entry(frame_top, textvariable=self.var_password, width=30, show="*").grid(row=0, column=3, padx=5, pady=4)

        ttk.Label(frame_top, text="Mã xã").grid(row=1, column=0, padx=5, pady=4, sticky="w")
        ttk.Entry(frame_top, textvariable=self.var_maxa, width=20).grid(row=1, column=1, padx=5, pady=4, sticky="w")

        ttk.Label(frame_top, text="Số luồng (workers)").grid(row=1, column=2, padx=5, pady=4, sticky="w")
        ttk.Spinbox(frame_top, textvariable=self.var_workers, from_=1, to=10, width=6).grid(
            row=1, column=3, padx=5, pady=4, sticky="w")

        ttk.Checkbutton(
            frame_top,
            text="DRY RUN (chỉ kiểm tra, KHÔNG xóa thật)",
            variable=self.var_dry_run
        ).grid(row=2, column=1, columnspan=3, padx=5, pady=4, sticky="w")

        ttk.Label(frame_top, text="File Excel").grid(row=3, column=0, padx=5, pady=4, sticky="w")
        ttk.Entry(frame_top, textvariable=self.var_excel, width=90).grid(
            row=3, column=1, columnspan=3, padx=5, pady=4, sticky="we")
        ttk.Button(frame_top, text="Duyệt", command=self.browse_excel).grid(row=3, column=4, padx=5, pady=4)

        self.btn_start = ttk.Button(frame_top, text="▶  BẮT ĐẦU CHẠY", command=self.start_run)
        self.btn_start.grid(row=4, column=1, padx=5, pady=8, sticky="w")
        self.btn_clear = ttk.Button(frame_top, text="Xóa log", command=self.clear_log)
        self.btn_clear.grid(row=4, column=2, padx=5, pady=8, sticky="w")

        ttk.Label(frame_top, text="Excel cần cột: soto | sothua",
                  foreground="blue").grid(row=5, column=0, columnspan=5, padx=5, pady=3, sticky="w")
        ttk.Label(frame_top,
                  text="Logic: ListGiayChungNhan trống HOẶC soPhatHanh trống  ->  XÓA đơn đăng ký",
                  foreground="red").grid(row=6, column=0, columnspan=5, padx=5, pady=(0, 4), sticky="w")
        frame_top.columnconfigure(3, weight=1)

        # Progress
        self.frame_progress = ttk.LabelFrame(self, text="Tiến độ")
        self.frame_progress.pack(fill="x", padx=10, pady=4)

        ttk.Label(self.frame_progress, text="Tổng:", width=12, anchor="w").grid(
            row=0, column=0, padx=6, pady=4, sticky="w")
        self._pb_total = ttk.Progressbar(self.frame_progress, length=600, mode="determinate")
        self._pb_total.grid(row=0, column=1, padx=6, pady=4, sticky="we")
        self._lbl_total = ttk.Label(self.frame_progress, text="0 / 0  |  🗑0  ⏭0  ❌0", width=34)
        self._lbl_total.grid(row=0, column=2, padx=6, pady=4, sticky="w")
        self.frame_progress.columnconfigure(1, weight=1)

        self._worker_rows = []

        # Log
        frame_log = ttk.LabelFrame(self, text="Log xử lý")
        frame_log.pack(fill="both", expand=True, padx=10, pady=(4, 10))

        self.txt_log = tk.Text(frame_log, wrap="word", font=("Consolas", 9))
        self.txt_log.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(frame_log, command=self.txt_log.yview)
        scrollbar.pack(side="right", fill="y")
        self.txt_log.configure(yscrollcommand=scrollbar.set)

    def _init_progress(self, total, n_workers):
        self._total = total
        self._pb_total.config(maximum=total, value=0)
        self._lbl_total.config(text=f"0 / {total}  |  🗑0  ⏭0  ❌0")

        for (lbl, pb, lbl_status) in self._worker_rows:
            lbl.grid_forget(); pb.grid_forget(); lbl_status.grid_forget()
        self._worker_rows.clear()

        for i in range(n_workers):
            row_idx = i + 1
            lbl = ttk.Label(self.frame_progress, text=f"Worker {i+1}:", width=12, anchor="w")
            lbl.grid(row=row_idx, column=0, padx=6, pady=2, sticky="w")
            pb = ttk.Progressbar(self.frame_progress, length=600, mode="determinate", maximum=100)
            pb.grid(row=row_idx, column=1, padx=6, pady=2, sticky="we")
            lbl_status = ttk.Label(self.frame_progress, text="Chờ...", width=34, foreground="gray")
            lbl_status.grid(row=row_idx, column=2, padx=6, pady=2, sticky="w")
            self._worker_rows.append((lbl, pb, lbl_status))

    def browse_excel(self):
        path = filedialog.askopenfilename(
            title="Chọn file Excel",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")]
        )
        if path:
            self.var_excel.set(path)

    def clear_log(self):
        self.txt_log.delete("1.0", tk.END)

    def validate_input(self):
        if not self.var_username.get().strip():
            messagebox.showerror("Thiếu thông tin", "Chưa nhập username."); return False
        if not self.var_password.get().strip():
            messagebox.showerror("Thiếu thông tin", "Chưa nhập password."); return False
        if not self.var_maxa.get().strip():
            messagebox.showerror("Thiếu thông tin", "Chưa nhập mã xã."); return False
        if not os.path.isfile(self.var_excel.get().strip()):
            messagebox.showerror("Sai đường dẫn", "File Excel không tồn tại."); return False
        return True

    def start_run(self):
        if self.worker_thread and self.worker_thread.is_alive():
            messagebox.showwarning("Đang chạy", "Chương trình đang chạy."); return
        if not self.validate_input():
            return

        global MAX_WORKERS, DRY_RUN
        MAX_WORKERS = self.var_workers.get()
        DRY_RUN     = self.var_dry_run.get()

        # Cảnh báo khi xóa thật
        if not DRY_RUN:
            if not messagebox.askyesno(
                "XÁC NHẬN XÓA THẬT",
                "Bạn đang TẮT DRY RUN.\n\n"
                "Chương trình sẽ XÓA THẬT các đơn đăng ký không có GCN / soPhatHanh trống.\n"
                "Thao tác này KHÔNG hồi lại được.\n\nBạn chắc chắn muốn tiếp tục?"
            ):
                return

        self.btn_start.config(state="disabled")

        self.worker_thread = threading.Thread(
            target=worker_run,
            args=(
                self.var_username.get().strip(),
                self.var_password.get().strip(),
                self.var_maxa.get().strip(),
                self.var_excel.get().strip(),
                self.log_queue,
            ),
            daemon=True
        )
        self.worker_thread.start()

    def process_log_queue(self):
        try:
            while True:
                msg = self.log_queue.get_nowait()

                if isinstance(msg, dict):
                    t = msg.get("type")

                    if t == "init":
                        self._init_progress(msg["total"], msg["workers"])

                    elif t == "worker_row_progress":
                        w = msg["worker"] - 1
                        row = msg["row"]
                        percent = max(0, min(100, int(msg.get("percent", 0))))
                        text = msg.get("text", "")
                        if 0 <= w < len(self._worker_rows):
                            _, pb, lbl_s = self._worker_rows[w]
                            pb.config(value=percent)
                            if percent >= 100:
                                lbl_s.config(text=f"✅ Dòng {row} — {percent}% — {text}", foreground="green")
                            else:
                                lbl_s.config(text=f"⏳ Dòng {row} — {percent}% — {text}", foreground="blue")

                    elif t == "progress":
                        done  = msg["done"]; total = msg["total"]
                        dx = msg["da_xoa"]; gl = msg["giu_lai"]; lo = msg["loi"]
                        self._pb_total.config(value=done)
                        pct = int(done / total * 100) if total else 0
                        self._lbl_total.config(
                            text=f"{done} / {total} ({pct}%)  |  🗑{dx}  ⏭{gl}  ❌{lo}"
                        )
                    continue

                if msg == "__DONE__":
                    self.btn_start.config(state="normal")
                    for (_, pb, lbl_s) in self._worker_rows:
                        lbl_s.config(text="Xong", foreground="gray")
                    self.txt_log.insert(tk.END, "\n✅ Tiến trình đã kết thúc.\n")
                    self.txt_log.see(tk.END)
                    continue

                self.txt_log.insert(tk.END, msg + "\n")
                self.txt_log.see(tk.END)

        except queue.Empty:
            pass

        self.after(200, self.process_log_queue)


if __name__ == "__main__":
    app = App()
    app.mainloop()