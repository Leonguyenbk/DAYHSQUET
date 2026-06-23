# -*- coding: utf-8 -*-

import os
import sys
import json
import uuid
import copy
import re
import queue
import threading
import traceback
import time
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime, timezone, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from openpyxl import load_workbook, Workbook

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys


# =========================
# API CONFIG
# =========================

REFERER_URL = "https://dla.mplis.gov.vn/dc/DonDangKy/KeKhaiDangKyV2"

API_SEARCH_HOSOQUET = "https://dla.mplis.gov.vn/dc/QuanLyKhoHoSoQuetAjax/SearchHoSoQuet"
API_UPDATE_HOSOQUET = "https://dla.mplis.gov.vn/dc/HoSoQuetAjax/UpdateHoSoQuetExistFile"
URL_GET_THONG_TIN_DANG_KY = "https://dla.mplis.gov.vn/dc/DangKyAjax/GetThongTinDangKyByTinhHinhDangKyIds"
URL_UPDATE_THONG_TIN_DANG_KY = "https://dla.mplis.gov.vn/dc/DangKyAjax/UpdateThongTinDangKy"

# True = chỉ kiểm tra, không update thật
# False = update thật
DRY_RUN = False

# Số luồng xử lý song song (tăng nếu server chịu được, giảm nếu bị rate-limit)
MAX_WORKERS = 3
API_SEARCH_TIMEOUT = 120
API_UPDATE_TIMEOUT = 180
API_SEARCH_RETRIES = 3
API_RETRY_BACKOFF_SECONDS = 5

# Ghi Excel sau mỗi N dòng (thay vì mỗi dòng)
WRITE_EVERY_N = 10


# =========================
# LOG REDIRECT (thread-safe)
# =========================

class ThreadSafeLogger:
    """Logger thread-safe, dùng chung cho nhiều worker."""
    def __init__(self, log_queue):
        self.log_queue = log_queue
        self._lock = threading.Lock()

    def log(self, msg):
        if msg and str(msg).strip():
            with self._lock:
                self.log_queue.put(str(msg))

    def write(self, text):
        self.log(text)

    def flush(self):
        pass


# =========================
# EXCEL
# =========================

def doc_excel(path_excel):
    wb = load_workbook(path_excel, data_only=True)
    ws = wb.active

    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value:
            headers[str(value).strip().lower()] = col

    required_cols = ["soto", "sothua", "loaidat", "tenfile"]
    missing = [c for c in required_cols if c not in headers]

    if missing:
        raise RuntimeError(f"Thiếu cột trong Excel: {', '.join(missing)}")

    data = []

    for row in range(2, ws.max_row + 1):
        soto    = ws.cell(row=row, column=headers["soto"]).value
        sothua  = ws.cell(row=row, column=headers["sothua"]).value
        loaidat = ws.cell(row=row, column=headers["loaidat"]).value
        tenfile = ws.cell(row=row, column=headers["tenfile"]).value

        if not soto and not sothua and not loaidat and not tenfile:
            continue

        data.append({
            "row":     row,
            "soto":    str(soto).strip()    if soto    is not None else "",
            "sothua":  str(sothua).strip()  if sothua  is not None else "",
            "loaidat": str(loaidat).strip() if loaidat is not None else "",
            "tenfile": str(tenfile).strip() if tenfile is not None else "",
        })

    return data


def ghi_excel_output(rows, output_path, lock):
    """Ghi kết quả ra Excel, dùng lock để tránh ghi đồng thời."""
    with lock:
        wb = Workbook()
        ws = wb.active
        ws.title = "KetQua"

        headers = [
            "STT", "Dòng Excel", "Số tờ", "Số thửa", "Loại đất", "Tên file",
            "Mô tả mới", "tinhHinhDangKyId", "hoSoQuetId", "thongTinHoSoId",
            "Chủ sử dụng", "Trạng thái", "Ghi chú"
        ]
        ws.append(headers)

        # Sắp xếp theo row_excel để output có thứ tự
        sorted_rows = sorted(rows, key=lambda x: x.get("row_excel") or 0)

        for i, r in enumerate(sorted_rows, start=1):
            ws.append([
                i,
                r.get("row_excel"),
                r.get("soto"),
                r.get("sothua"),
                r.get("loaidat"),
                r.get("tenfile"),
                r.get("mo_ta"),
                r.get("tinhHinhDangKyId"),
                r.get("hoSoQuetId"),
                r.get("thongTinHoSoId"),
                r.get("chu_su_dung"),
                r.get("status"),
                r.get("note"),
            ])

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

        wb.save(output_path)


def lay_duong_dan_file(folder_upload, tenfile):
    if os.path.isabs(tenfile):
        return tenfile
    return os.path.join(folder_upload, tenfile)


# =========================
# SELENIUM LOGIN + TOKEN
# =========================

def lay_token_tu_trang(driver):
    js = """
    return (
        document.querySelector('input[name="__RequestVerificationToken"]')?.value ||
        document.querySelector('input[name="__requestverificationtoken"]')?.value ||
        document.querySelector('meta[name="__RequestVerificationToken"]')?.content ||
        document.querySelector('meta[name="__requestverificationtoken"]')?.content ||
        document.querySelector('meta[name="RequestVerificationToken"]')?.content ||
        ''
    );
    """
    return driver.execute_script(js)


def tao_session_tu_selenium(driver):
    session = requests.Session()

    user_agent = driver.execute_script("return navigator.userAgent;")
    token = lay_token_tu_trang(driver)

    if not token:
        raise RuntimeError("Không lấy được __requestverificationtoken.")

    session.headers.update({
        "User-Agent": user_agent,
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "X-Requested-With": "XMLHttpRequest",
        "Origin": "https://dla.mplis.gov.vn",
        "Referer": REFERER_URL,
        "__requestverificationtoken": token,
        "__RequestVerificationToken": token,
        "RequestVerificationToken": token,
    })

    for c in driver.get_cookies():
        session.cookies.set(
            name=c["name"],
            value=c["value"],
            domain=c.get("domain"),
            path=c.get("path", "/")
        )

    return session


def login_mplis(username, password):
    options = Options()
    options.add_argument("--start-maximized")
    options.add_argument("--window-position=100,100")
    options.add_argument("--window-size=1400,900")

    driver = webdriver.Chrome(options=options)
    driver.get(REFERER_URL)

    try:
        import time
        time.sleep(2)

        inputs = driver.find_elements(By.CSS_SELECTOR, "input")
        user_box = None
        pass_box = None

        for inp in inputs:
            typ         = (inp.get_attribute("type")        or "").lower()
            name        = (inp.get_attribute("name")        or "").lower()
            placeholder = (inp.get_attribute("placeholder") or "").lower()
            input_id    = (inp.get_attribute("id")          or "").lower()
            text_all    = f"{name} {placeholder} {input_id}"

            if not user_box and typ in ["text", "email"] and any(
                k in text_all for k in ["user", "username", "login", "account", "ten"]
            ):
                user_box = inp

            if not pass_box and typ == "password":
                pass_box = inp

        if not user_box:
            for inp in inputs:
                typ = (inp.get_attribute("type") or "").lower()
                if typ in ["text", "email"]:
                    user_box = inp
                    break

        if user_box and pass_box:
            user_box.clear()
            user_box.send_keys(username)
            pass_box.clear()
            pass_box.send_keys(password)
            pass_box.send_keys(Keys.ENTER)
        else:
            pass  # người dùng tự đăng nhập tay

    except Exception:
        pass

    messagebox.showinfo(
        "Đăng nhập / Authenticator",
        "Hoàn tất đăng nhập MPLIS và xác thực Authenticator trên Chrome.\n"
        "Sau khi vào được màn hình kê khai đăng ký thì bấm OK để tiếp tục."
    )

    return driver



# =========================
# API: GET + UPDATE THÔNG TIN ĐĂNG KÝ
# =========================

def dotnet_date_to_iso(value):
    """
    Chuyển /Date(1780843978377)/ sang ISO UTC: 2026-06-07T14:52:58.377Z.
    Nếu không phải /Date(...)/ thì giữ nguyên.
    """
    if not isinstance(value, str):
        return value

    m = re.search(r"/Date\((-?\d+)\)/", value)
    if not m:
        return value

    ms = int(m.group(1))
    dt = datetime.fromtimestamp(ms / 1000, tz=timezone.utc)
    return dt.strftime("%Y-%m-%dT%H:%M:%S.") + f"{int(dt.microsecond / 1000):03d}Z"


def convert_dates_recursive(obj):
    """Convert toàn bộ ngày /Date(...)/ trong dict/list sang ISO."""
    if isinstance(obj, dict):
        return {k: convert_dates_recursive(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [convert_dates_recursive(x) for x in obj]
    return dotnet_date_to_iso(obj)


def ddmmyyyy_to_iso_utc_start_of_day_vn(date_str):
    """
    Nhập dd/mm/yyyy theo ngày Việt Nam.
    Ví dụ 23/04/2026 -> 2026-04-22T17:00:00.000Z.
    """
    date_str = str(date_str).strip()
    dt_vn = datetime.strptime(date_str, "%d/%m/%Y")
    tz_vn = timezone(timedelta(hours=7))
    dt_vn = dt_vn.replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=tz_vn)
    dt_utc = dt_vn.astimezone(timezone.utc)
    return dt_utc.strftime("%Y-%m-%dT%H:%M:%S.000Z")


def lay_thong_tin_dang_ky_by_id(session, tinh_hinh_dang_ky_id, token=None):
    """
    Lấy thông tin đăng ký từ tinhHinhDangKyId.

    Theo F12, endpoint này nhận JSON raw body:
        {
            "tinhHinhDangKyIds": [19540778],
            "getHoSoQuet": false
        }

    Không gửi form-urlencoded, vì form có thể bind sai và trả value: [].
    """
    tid = safe_int(tinh_hinh_dang_ky_id, 0)
    if not tid:
        return {"ok": False, "error": f"tinhHinhDangKyId không hợp lệ: {tinh_hinh_dang_ky_id}"}

    headers = dict(session.headers)
    headers.update({
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Content-Type": "application/json; charset=UTF-8",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": REFERER_URL,
        "Origin": "https://dla.mplis.gov.vn",
    })

    if token:
        headers["__RequestVerificationToken"] = token
        headers["RequestVerificationToken"] = token
        headers["__requestverificationtoken"] = token

    payload = {
        "tinhHinhDangKyIds": [tid],
        "getHoSoQuet": False,
    }

    res, request_error = post_retry(
        session,
        URL_GET_THONG_TIN_DANG_KY,
        retries=API_SEARCH_RETRIES,
        timeout=API_SEARCH_TIMEOUT,
        data=json.dumps(payload, ensure_ascii=False),
        headers=headers
    )
    if request_error:
        return {"ok": False, "error": request_error}

    print("=" * 80)
    print("GET THÔNG TIN ĐĂNG KÝ")
    print("ID:", tid)
    print("PAYLOAD:", json.dumps(payload, ensure_ascii=False))
    print("STATUS:", getattr(res, "status_code", ""))
    print("TEXT:", getattr(res, "text", "")[:1000])

    try:
        js = res.json()
    except Exception:
        return {
            "ok": False,
            "error": f"Response GetThongTinDangKy không phải JSON: {res.text[:1000]}",
            "status_code": res.status_code
        }

    if not js.get("success"):
        return {
            "ok": False,
            "error": f"GetThongTinDangKy success=false với tinhHinhDangKyId={tid}. Response={str(js)[:1000]}",
            "raw": js
        }

    if not js.get("value"):
        return {
            "ok": False,
            "error": f"Không lấy được value từ tinhHinhDangKyId={tid}. Response={str(js)[:1000]}",
            "raw": js
        }

    return {"ok": True, "raw": js}


def build_payload_update_quyen_quan_ly(response_json, ngay_dang_ky_lan_dau_ddmmyyyy):
    """
    Build payload lõi thongTinDangKy từ response GetThongTinDangKyByTinhHinhDangKyIds.

    Giữ nguyên toàn bộ object value[0], gồm:
      - TinhHinhDangKy
      - ChuSoHuu
      - TaiSan
      - ListDangKyQuyen
      - ListDangKyTaiSan
      - ListGiayChungNhan
      - ListHanCheQuyen
      - ListHanCheThuaLienKe
      - ListNghiaVuTaiChinh
      - các field khác nếu có

    Chỉ sửa đúng:
      - TinhHinhDangKy.coQuyenQuanLy = True
      - TinhHinhDangKy.thoiDiemDangKyLanDau = ngày nhập UI convert ISO UTC
    """
    if not response_json.get("value"):
        raise Exception("Response không có value để build payload UpdateThongTinDangKy")

    value0 = copy.deepcopy(response_json["value"][0])

    # Trường hợp hiếm nếu response đã bọc sẵn thongTinDangKy thì lấy lõi bên trong.
    if isinstance(value0, dict) and "thongTinDangKy" in value0 and isinstance(value0["thongTinDangKy"], dict):
        thong_tin_dang_ky = value0["thongTinDangKy"]
    else:
        thong_tin_dang_ky = value0

    thong_tin_dang_ky = convert_dates_recursive(thong_tin_dang_ky)

    if "TinhHinhDangKy" not in thong_tin_dang_ky or not isinstance(thong_tin_dang_ky["TinhHinhDangKy"], dict):
        raise Exception("Payload không có TinhHinhDangKy, không thể update.")

    thong_tin_dang_ky["TinhHinhDangKy"]["coQuyenQuanLy"] = True
    thong_tin_dang_ky["TinhHinhDangKy"]["thoiDiemDangKyLanDau"] = ddmmyyyy_to_iso_utc_start_of_day_vn(
        ngay_dang_ky_lan_dau_ddmmyyyy
    )

    return thong_tin_dang_ky


def _response_update_ok(js):
    if not isinstance(js, dict):
        return False
    if js.get("success") is True or js.get("Success") is True or js.get("ok") is True:
        return True
    # Một số API chỉ trả {value: ...} nhưng không có success.
    if "success" not in js and "error" not in js and "Error" not in js:
        return True
    return False


def api_update_thong_tin_dang_ky(session, tinh_hinh_dang_ky_id, ngay_dang_ky_lan_dau_ddmmyyyy):
    """
    1) GET thông tin theo tinhHinhDangKyId bằng JSON raw.
    2) Build payload lõi thongTinDangKy, giữ nguyên toàn bộ dữ liệu.
    3) POST UpdateThongTinDangKy.

    Ưu tiên gửi dạng F12 thường thấy:
        {"thongTinDangKy": {...}}
    Nếu server không nhận thì fallback gửi raw {...}.
    """
    res_get = lay_thong_tin_dang_ky_by_id(session, tinh_hinh_dang_ky_id)
    if not res_get.get("ok"):
        return {"ok": False, "error": "GET thông tin đăng ký lỗi: " + res_get.get("error", "")}

    try:
        thong_tin_dang_ky = build_payload_update_quyen_quan_ly(
            res_get["raw"],
            ngay_dang_ky_lan_dau_ddmmyyyy
        )
    except Exception as e:
        return {"ok": False, "error": "Build payload UpdateThongTinDangKy lỗi: " + str(e)}

    headers_json = dict(session.headers)
    headers_json.update({
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Content-Type": "application/json; charset=UTF-8",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": REFERER_URL,
        "Origin": "https://dla.mplis.gov.vn",
    })

    payloads = [
        ("json_wrapper_thongTinDangKy", {"thongTinDangKy": thong_tin_dang_ky}),
        ("json_raw_thongTinDangKy_core", thong_tin_dang_ky),
    ]

    last_text = ""
    last_js = None

    for mode, payload in payloads:
        try:
            with open(f"debug_UpdateThongTinDangKy_{mode}.json", "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

        try:
            res = session.post(
                URL_UPDATE_THONG_TIN_DANG_KY,
                data=json.dumps(payload, ensure_ascii=False),
                headers=headers_json,
                timeout=API_UPDATE_TIMEOUT
            )
        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
            last_text = f"Timeout/lỗi mạng khi UpdateThongTinDangKy ({mode}): {e}"
            continue
        except requests.exceptions.RequestException as e:
            last_text = f"Lỗi request UpdateThongTinDangKy ({mode}): {e}"
            continue

        print("=" * 80)
        print("UPDATE THÔNG TIN ĐĂNG KÝ")
        print("MODE:", mode)
        print("STATUS:", res.status_code)
        print("TEXT:", res.text[:1000])

        last_text = res.text[:1000]
        try:
            js = res.json()
        except Exception:
            js = None
        last_js = js

        if res.ok and _response_update_ok(js):
            return {"ok": True, "raw": js, "payload": payload, "sent_as": mode}

    return {
        "ok": False,
        "error": "UpdateThongTinDangKy không thành công. "
                 f"Phản hồi cuối: {last_js if last_js is not None else last_text}",
        "payload_core": thong_tin_dang_ky
    }


def normalize_tenfile_key(tenfile):
    """Gom nhóm theo tên file, bỏ khác biệt hoa/thường và folder path."""
    tenfile = str(tenfile or "").strip().replace("/", "\\")
    return os.path.basename(tenfile).strip().upper()

# =========================
# API: SEARCH HỒ SƠ QUÉT
# =========================

def tao_payload_search_hosoquet(xa_id, so_to, so_thua):
    return {
        "draw": "2",

        "columns[0][data]": "",          "columns[0][name]": "",
        "columns[0][searchable]": "true","columns[0][orderable]": "false",
        "columns[0][search][value]": "", "columns[0][search][regex]": "false",

        "columns[1][data]": "ListHoSoQuet",   "columns[1][name]": "ListHoSoQuet",
        "columns[1][searchable]": "true",     "columns[1][orderable]": "false",
        "columns[1][search][value]": "",      "columns[1][search][regex]": "false",

        "columns[2][data]": "ThongTinDangKy", "columns[2][name]": "GiayChungNhan",
        "columns[2][searchable]": "true",     "columns[2][orderable]": "false",
        "columns[2][search][value]": "",      "columns[2][search][regex]": "false",

        "columns[3][data]": "ThongTinDangKy", "columns[3][name]": "ChuSoHuu",
        "columns[3][searchable]": "true",     "columns[3][orderable]": "false",
        "columns[3][search][value]": "",      "columns[3][search][regex]": "false",

        "columns[4][data]": "ThongTinDangKy", "columns[4][name]": "TaiSan",
        "columns[4][searchable]": "true",     "columns[4][orderable]": "false",
        "columns[4][search][value]": "",      "columns[4][search][regex]": "false",

        "start": "0", "length": "10",
        "search[value]": "", "search[regex]": "false",

        "xaId": str(xa_id), "huyenId": "0", "khoId": "0",
        "dayId": "0", "keId": "0", "hopId": "0",

        "maDon": "", "soThuTu": "", "ngayTiepNhan": "",
        "thoiDiemDangKy": "", "tuNgay": "", "denNgay": "",
        "loaiGiayChungNhanId": "", "maVach": "", "soPhatHanh": "",
        "soVaoSo": "", "soVaoSoCu": "", "ngayVaoSo": "",
        "soHoSoGoc": "", "soHoSoGocCu": "",
        "hoTen": "", "soGiayTo": "", "namSinh": "",

        "soThuTuThua":    str(so_thua),
        "soHieuToBanDo":  str(so_to),

        "soThuTuThuaCu": "", "soHieuToBanDoCu": "",
        "soNha": "", "diaChiChiTiet": "", "getHoSoLichSu": ""
    }


def post_retry(session, url, *, retries, timeout, **kwargs):
    last_error = None

    for attempt in range(1, retries + 1):
        try:
            return session.post(url, timeout=timeout, **kwargs), None
        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
            last_error = e
            if attempt < retries:
                time.sleep(API_RETRY_BACKOFF_SECONDS * attempt)
        except requests.exceptions.RequestException as e:
            return None, f"Loi request API: {e}"

    return None, f"Timeout/loi mang sau {retries} lan thu: {last_error}"


def api_search_hosoquet(session, xa_id, so_to, so_thua):
    headers = dict(session.headers)
    headers["Content-Type"] = "application/x-www-form-urlencoded; charset=UTF-8"

    res, request_error = post_retry(
        session,
        API_SEARCH_HOSOQUET,
        retries=API_SEARCH_RETRIES,
        timeout=API_SEARCH_TIMEOUT,
        data=tao_payload_search_hosoquet(xa_id, so_to, so_thua),
        headers=headers
    )
    if request_error:
        return {"ok": False, "error": request_error}

    try:
        result = res.json()
    except Exception:
        return {
            "ok": False,
            "error": f"Response không phải JSON: {res.text[:1000]}",
            "status_code": res.status_code
        }

    if not result.get("success"):
        return {"ok": False, "error": str(result), "raw": result}

    data = result.get("data") or []

    if not data:
        return {
            "ok": False,
            "error": "Không tìm thấy hồ sơ quét theo tờ/thửa",
            "raw": result
        }

    if len(data) > 1:
        return {
            "ok": False,
            "error": f"Tìm thấy {len(data)} bản ghi, không xử lý để tránh nhầm",
            "raw": result
        }

    item = data[0]
    list_hosoquet = item.get("ListHoSoQuet") or []

    if not list_hosoquet:
        return {
            "ok": False,
            "error": "Có bản ghi nhưng không có ListHoSoQuet",
            "raw": result
        }

    return {
        "ok": True,
        "item": item,
        "list_hosoquet": list_hosoquet,
        "raw": result
    }


# =========================
# PARSE
# =========================

def lay_chu_su_dung(item):
    try:
        thong_tin = item.get("ThongTinDangKy") or {}
        chu = thong_tin.get("ChuSoHuu") or {}

        ca_nhans = chu.get("CaNhans") or []
        if ca_nhans:
            return ca_nhans[0].get("hoTen") or ""

        to_chucs = chu.get("ToChucs") or []
        if to_chucs:
            return to_chucs[0].get("tenToChuc") or to_chucs[0].get("ten") or ""
    except Exception:
        pass
    return ""


def lay_dien_tich(item):
    try:
        thong_tin = item.get("ThongTinDangKy") or {}
        tai_san   = thong_tin.get("TaiSan") or {}
        thua_dats = tai_san.get("ThuaDats") or []
        if thua_dats:
            return thua_dats[0].get("dienTich") or thua_dats[0].get("dienTichPhapLy") or ""
    except Exception:
        pass
    return ""


def get_tinh_hinh_id_from_item_hoso(item, hoso):
    """Ưu tiên lấy tinhHinhDangKyId từ ThongTinDangKy.TinhHinhDangKy vì đây là ID đăng ký thật."""
    thong_tin = item.get("ThongTinDangKy") or {}
    tinh_hinh = thong_tin.get("TinhHinhDangKy") or {}

    tid = tinh_hinh.get("tinhHinhDangKyId")
    if not tid:
        tid = hoso.get("tinhHinhDangKyId")
    if not tid:
        tid = hoso.get("Title")
    return tid


def tim_hosoquet_tu_search(raw_search, uu_tien_chuacogiay=True):
    data = raw_search.get("data") or []

    for item in data:
        thong_tin     = item.get("ThongTinDangKy") or {}
        tinh_hinh     = thong_tin.get("TinhHinhDangKy") or {}
        list_hosoquet = item.get("ListHoSoQuet") or []

        if uu_tien_chuacogiay:
            for hoso in list_hosoquet:
                wrapper = hoso.get("ListFileHoSoQuet") or {}
                files   = wrapper.get("ListFileHoSoQuet") or []
                for f in files:
                    mo_ta = (f.get("moTa") or "").upper()
                    if "CHUACOGIAY" in mo_ta:
                        return {
                            "info": {
                                "thongTinHoSoId":   hoso.get("thongTinHoSoId"),
                                "tinhHinhDangKyId": get_tinh_hinh_id_from_item_hoso(item, hoso),
                                "xaId":             hoso.get("xaId") or tinh_hinh.get("xaId"),
                            },
                            "hoso": hoso, "file": f, "item": item
                        }

        if list_hosoquet:
            hoso = list_hosoquet[0]
            return {
                "info": {
                    "thongTinHoSoId":   hoso.get("thongTinHoSoId"),
                    "tinhHinhDangKyId": get_tinh_hinh_id_from_item_hoso(item, hoso),
                    "xaId":             hoso.get("xaId") or tinh_hinh.get("xaId"),
                },
                "hoso": hoso, "file": None, "item": item
            }

    return None


def dem_file_trong_hoso(hoso):
    wrapper = hoso.get("ListFileHoSoQuet") or {}
    files   = wrapper.get("ListFileHoSoQuet") or []
    return len(files)


def co_file_khong_chuacogiay(hoso):
    wrapper = hoso.get("ListFileHoSoQuet") or {}
    files   = wrapper.get("ListFileHoSoQuet") or []

    for f in files:
        mo_ta    = (f.get("moTa")   or "").upper()
        ten_file = (f.get("tenFile") or f.get("Name") or "").upper()
        text     = mo_ta + " " + ten_file

        if text.strip() and "CHUACOGIAY" not in text:
            return True, f.get("moTa") or f.get("tenFile") or "Có file không phải CHUACOGIAY"

    return False, ""


# =========================
# API: UPDATE FILE HỒ SƠ QUÉT
# =========================

def now_iso_z():
    return datetime.now(timezone.utc).isoformat(timespec="milliseconds").replace("+00:00", "Z")


def safe_int(value, default=0):
    try:
        if value is None or value == "":
            return default
        return int(value)
    except Exception:
        return default


def api_update_hosoquet_exist_file(session, file_path, found_hosoquet, mo_ta_moi):
    if not os.path.isfile(file_path):
        return {"ok": False, "error": f"Không tìm thấy file PDF: {file_path}"}

    info = found_hosoquet["info"]
    hoso = found_hosoquet["hoso"]

    ho_so_quet_id       = hoso.get("hoSoQuetId")       or hoso.get("Title")
    thong_tin_ho_so_id  = hoso.get("thongTinHoSoId")   or info.get("thongTinHoSoId")
    tinh_hinh_dang_ky_id = info.get("tinhHinhDangKyId") or hoso.get("tinhHinhDangKyId")
    xa_id               = hoso.get("xaId")             or info.get("xaId")

    ho_so_quet = {
        "hoSoQuetId":        safe_int(ho_so_quet_id),
        "thongTinHoSoId":    safe_int(thong_tin_ho_so_id),
        "tinhHinhDangKyId":  safe_int(tinh_hinh_dang_ky_id),
        "xaId":              safe_int(xa_id),
        "CreatedDate":       now_iso_z(),
        "ModifiedDate":      now_iso_z(),
        "Id":                hoso.get("Id"),
        "Title":             str(hoso.get("Title") or ho_so_quet_id),
        "Name":              hoso.get("Name"),
        "Path":              hoso.get("Path"),
        "ParentPath":        hoso.get("ParentPath"),
        "_id":               1,
        "TuiHoSo":           None,
        "tuiHoSoId":         safe_int(hoso.get("tuiHoSoId"), 0)
    }

    info_ho_so_quet = {
        "loaiHoSoQuet":       2,
        "laGiayToVeNguonGoc": False,
        "giayChungNhanId":    "",
        "moTa":               mo_ta_moi,
        "tenGiayTo":          "",
        "trichYeu":           "",
        "laGiayChungNhan":    False,
        "__id":               str(uuid.uuid4()),
        "files":              None
    }

    data = {
        "hoSoQuet":       json.dumps(ho_so_quet,      ensure_ascii=False),
        "infoHoSoQuet_1": json.dumps(info_ho_so_quet, ensure_ascii=False),
        "count":          "1"
    }

    headers = dict(session.headers)
    headers.pop("Content-Type", None)

    with open(file_path, "rb") as f:
        files = {
            "fileHoSoQuet_1": (
                os.path.basename(file_path),
                f,
                "application/pdf"
            )
        }
        try:
            res = session.post(
                API_UPDATE_HOSOQUET,
                data=data,
                files=files,
                headers=headers,
                timeout=API_UPDATE_TIMEOUT
            )
        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
            return {
                "ok": False,
                "error": f"Timeout/loi mang khi update/upload: {e}"
            }
        except requests.exceptions.RequestException as e:
            return {
                "ok": False,
                "error": f"Loi request API khi update/upload: {e}"
            }

    try:
        result = res.json()
    except Exception:
        return {
            "ok": False,
            "error": f"Response update không phải JSON: {res.text[:1000]}",
            "status_code": res.status_code
        }

    if not result.get("success"):
        return {"ok": False, "error": str(result), "raw": result}

    return {"ok": True, "raw": result}


def kiem_tra_sau_update_bang_search(session, xa_id, so_to, so_thua, mo_ta_moi):
    res = api_search_hosoquet(session, xa_id, so_to, so_thua)

    if not res.get("ok"):
        return False, "Không search lại được hồ sơ quét: " + res.get("error", "")

    target = mo_ta_moi.upper()
    data   = res["raw"].get("data") or []

    for item in data:
        for hoso in item.get("ListHoSoQuet") or []:
            wrapper = hoso.get("ListFileHoSoQuet") or {}
            files   = wrapper.get("ListFileHoSoQuet") or []
            for f in files:
                mo_ta    = (f.get("moTa")   or "").upper()
                ten_file = (f.get("tenFile") or f.get("Name") or "").upper()
                if target in mo_ta or target in ten_file:
                    return True, "Đã thấy mô tả mới trong hồ sơ quét"

    return False, "Chưa thấy mô tả mới sau cập nhật"


# =========================
# XỬ LÝ 1 DÒNG
# =========================

def xu_ly_1_dong(session, item, maxa, folder_upload, ngay_dang_ky_lan_dau, logger, progress_cb=None):
    row_excel = item["row"]
    soto      = item["soto"]
    sothua    = item["sothua"]
    loaidat   = item["loaidat"]
    tenfile   = item["tenfile"]

    file_path  = lay_duong_dan_file(folder_upload, tenfile)
    mo_ta_moi  = f"CHUACOGIAY_{maxa}_{loaidat}-DDK"

    result_row = {
        "row_excel":        row_excel,
        "soto":             soto,
        "sothua":           sothua,
        "loaidat":          loaidat,
        "tenfile":          tenfile,
        "mo_ta":            mo_ta_moi,
        "tinhHinhDangKyId": "",
        "hoSoQuetId":       "",
        "thongTinHoSoId":   "",
        "chu_su_dung":      "",
        "status":           "Lỗi",
        "note":             ""
    }

    def progress(percent, text):
        """Báo tiến độ riêng của dòng hiện tại cho worker đang xử lý."""
        if progress_cb:
            progress_cb(percent, text)

    def done(icon, status, note):
        result_row["status"] = status
        result_row["note"]   = note
        logger.log(f"{icon} Dòng {row_excel}: {note}")
        return result_row

    # 0% → bắt đầu dòng
    progress(0, "Bắt đầu")

    # 10% → kiểm tra file PDF
    progress(10, "Kiểm tra file PDF")
    if not os.path.isfile(file_path):
        progress(100, "Lỗi file")
        return done("❌", "Lỗi", f"Không tìm thấy file: {file_path}")

    # 25% → Search
    progress(25, "Đang search hồ sơ quét")
    res_search = api_search_hosoquet(session, xa_id=maxa, so_to=soto, so_thua=sothua)
    if not res_search.get("ok"):
        progress(100, "Search lỗi")
        return done("❌", "Lỗi", "Search lỗi: " + res_search.get("error", ""))

    item_search               = res_search["item"]
    result_row["chu_su_dung"] = lay_chu_su_dung(item_search)

    # 45% → parse kết quả search
    progress(45, "Đã tìm thấy hồ sơ")
    found = tim_hosoquet_tu_search(res_search["raw"])
    if not found:
        progress(100, "Không tìm thấy HSQ phù hợp")
        return done("⚠️", "Lỗi", "Không tìm thấy ListHoSoQuet phù hợp")

    info = found["info"]
    hoso = found["hoso"]

    result_row["hoSoQuetId"]       = hoso.get("hoSoQuetId")       or hoso.get("Title")
    result_row["thongTinHoSoId"]   = hoso.get("thongTinHoSoId")   or info.get("thongTinHoSoId")
    result_row["tinhHinhDangKyId"] = info.get("tinhHinhDangKyId") or hoso.get("tinhHinhDangKyId")

    logger.log(
        f"🔎 ID lấy được | tinhHinhDangKyId={result_row['tinhHinhDangKyId']} | "
        f"hoSoQuetId={result_row['hoSoQuetId']} | thongTinHoSoId={result_row['thongTinHoSoId']} | "
        f"tờ={soto} | thửa={sothua} | file={tenfile}"
    )

    # 55% → kiểm tra có file thật chưa
    progress(55, "Kiểm tra file đã có")
    da_co_file_that, ghi_chu_file = co_file_khong_chuacogiay(hoso)
    if da_co_file_that:
        progress(100, "Bỏ qua")
        return done("⏭️", "Bỏ qua", "Bỏ qua — đã có file: " + ghi_chu_file)

    if DRY_RUN:
        progress(100, "DRY RUN")
        return done("🧪", "DRY_RUN", "DRY_RUN — chưa update thật")

    # 70% → Upload/update
    progress(70, "Đang upload/update")
    res_update = api_update_hosoquet_exist_file(
        session=session,
        file_path=file_path,
        found_hosoquet=found,
        mo_ta_moi=mo_ta_moi
    )
    if not res_update.get("ok"):
        progress(100, "Update lỗi")
        return done("❌", "Lỗi", "Update hồ sơ quét lỗi: " + res_update.get("error", ""))

    # 82% → Lấy thông tin từ tinhHinhDangKyId và update quyền quản lý + thời điểm đăng ký lần đầu
    progress(82, "Đang cập nhật thông tin đăng ký")
    res_update_ttdk = api_update_thong_tin_dang_ky(
        session=session,
        tinh_hinh_dang_ky_id=result_row["tinhHinhDangKyId"],
        ngay_dang_ky_lan_dau_ddmmyyyy=ngay_dang_ky_lan_dau
    )
    if not res_update_ttdk.get("ok"):
        progress(100, "Update TTĐK lỗi")
        return done(
            "❌",
            "Lỗi",
            "Upload hồ sơ quét OK nhưng UpdateThongTinDangKy lỗi: " + res_update_ttdk.get("error", "")
        )

    # 90% → Verify
    progress(90, "Đang kiểm tra lại")
    ok_check, _ = kiem_tra_sau_update_bang_search(
        session=session,
        xa_id=maxa,
        so_to=soto,
        so_thua=sothua,
        mo_ta_moi=mo_ta_moi
    )

    if ok_check:
        progress(100, "Hoàn thành")
        return done("✅", "Thành công", "Cập nhật thành công")
    else:
        progress(100, "Cần kiểm tra")
        return done("⚠️", "Cần kiểm tra", "Cần kiểm tra lại")

# =========================
# WORKER (chạy trong thread riêng)
# =========================

def worker_run(username, password, maxa, ngay_dang_ky_lan_dau, excel_path, folder_upload, log_queue):
    logger = ThreadSafeLogger(log_queue)
    driver = None

    try:
        data = doc_excel(excel_path)

        if not data:
            logger.log("❌ Excel không có dữ liệu.")
            return

        # Kiểm tra ngày nhập từ UI
        try:
            ngay_iso_test = ddmmyyyy_to_iso_utc_start_of_day_vn(ngay_dang_ky_lan_dau)
        except Exception:
            raise RuntimeError("Ngày đăng ký lần đầu không đúng định dạng dd/mm/yyyy. Ví dụ: 23/04/2026")

        # Gom nhóm theo tên file. Các dòng cùng tenfile chỉ gọi API/search/upload/update 1 lần.
        grouped = {}
        for item in data:
            key = normalize_tenfile_key(item.get("tenfile"))
            if not key:
                key = f"__ROW_{item.get('row')}__"
            grouped.setdefault(key, []).append(item)

        groups = list(grouped.items())
        tong = len(groups)
        logger.log(f"✅ Đọc Excel xong: {len(data)} dòng | Gom còn {tong} nhóm tenfile duy nhất.")
        logger.log(f"Ngày đăng ký lần đầu: {ngay_dang_ky_lan_dau} -> {ngay_iso_test}")
        logger.log(f"DRY_RUN={DRY_RUN} | MAX_WORKERS={MAX_WORKERS}")
        logger.log("API search HSQ:  " + API_SEARCH_HOSOQUET)
        logger.log("API update HSQ:  " + API_UPDATE_HOSOQUET)
        logger.log("API get TTĐK:    " + URL_GET_THONG_TIN_DANG_KY)
        logger.log("API update TTĐK: " + URL_UPDATE_THONG_TIN_DANG_KY)

        output_path = os.path.join(
            os.path.dirname(excel_path),
            "ket_qua_cap_nhat_hosoquet_api_search.xlsx"
        )

        driver  = login_mplis(username, password)
        session = tao_session_tu_selenium(driver)
        thread_local = threading.local()

        def get_thread_session():
            if not hasattr(thread_local, "session"):
                worker_session = requests.Session()
                worker_session.headers.update(session.headers)
                worker_session.cookies.update(session.cookies)
                thread_local.session = worker_session
            return thread_local.session
        logger.log("✅ Đã tạo session API.")

        # Thông báo UI khởi tạo progress bars
        log_queue.put({"type": "init", "total": tong, "workers": MAX_WORKERS})

        results       = []
        results_lock  = threading.Lock()
        counter_lock  = threading.Lock()
        thanh_cong    = 0
        bo_qua        = 0
        that_bai      = 0
        done_count    = 0

        # Map thread → worker index 1..N
        worker_id_map  = {}
        worker_id_lock = threading.Lock()
        worker_counter = [0]

        def get_worker_idx():
            tid = threading.get_ident()
            with worker_id_lock:
                if tid not in worker_id_map:
                    worker_counter[0] += 1
                    worker_id_map[tid] = worker_counter[0]
                return worker_id_map[tid]

        def make_copy_result(master_result, item, master_row, group_key):
            copied = dict(master_result)
            copied["row_excel"] = item.get("row")
            copied["soto"] = item.get("soto")
            copied["sothua"] = item.get("sothua")
            copied["loaidat"] = item.get("loaidat")
            copied["tenfile"] = item.get("tenfile")
            copied["note"] = (
                f"Cùng tên file '{group_key}' với dòng {master_row}; "
                "không gọi API lần nữa. Kết quả xử lý dùng chung từ dòng đại diện. "
                + str(master_result.get("note") or "")
            )
            return copied

        def process_group(group_data):
            nonlocal thanh_cong, bo_qua, that_bai, done_count

            group_key, items_in_group = group_data
            first_item = items_in_group[0]
            w_idx = get_worker_idx()
            row_excel = first_item["row"]

            def row_progress(percent, text):
                log_queue.put({
                    "type": "worker_row_progress",
                    "worker": w_idx,
                    "row": row_excel,
                    "percent": percent,
                    "text": f"{text} | nhóm {len(items_in_group)} dòng"
                })

            try:
                kq_master = xu_ly_1_dong(
                    session=get_thread_session(),
                    item=first_item,
                    maxa=maxa,
                    folder_upload=folder_upload,
                    ngay_dang_ky_lan_dau=ngay_dang_ky_lan_dau,
                    logger=logger,
                    progress_cb=row_progress
                )
            except Exception as e:
                traceback.print_exc()
                kq_master = {
                    "row_excel":        first_item.get("row"),
                    "soto":             first_item.get("soto"),
                    "sothua":           first_item.get("sothua"),
                    "loaidat":          first_item.get("loaidat"),
                    "tenfile":          first_item.get("tenfile"),
                    "mo_ta":            "",
                    "tinhHinhDangKyId": "",
                    "hoSoQuetId":       "",
                    "thongTinHoSoId":   "",
                    "chu_su_dung":      "",
                    "status":           "Lỗi ngoài",
                    "note":             str(e)
                }

            group_results = [kq_master]
            if len(items_in_group) > 1:
                logger.log(
                    f"🔁 Nhóm tenfile '{group_key}' có {len(items_in_group)} dòng; "
                    f"đã xử lý 1 lần ở dòng {row_excel}, các dòng còn lại dùng chung ID/kết quả."
                )
                for item2 in items_in_group[1:]:
                    group_results.append(make_copy_result(kq_master, item2, row_excel, group_key))

            with results_lock:
                results.extend(group_results)

            with counter_lock:
                done_count += 1
                st = kq_master.get("status", "")
                if st == "Thành công":
                    thanh_cong += 1
                elif st == "Bỏ qua":
                    bo_qua += 1
                else:
                    that_bai += 1

                log_queue.put({
                    "type":       "progress",
                    "done":       done_count,
                    "total":      tong,
                    "thanh_cong": thanh_cong,
                    "bo_qua":     bo_qua,
                    "that_bai":   that_bai,
                })
                if done_count % WRITE_EVERY_N == 0 or done_count == tong:
                    try:
                        ghi_excel_output(results, output_path, results_lock)
                        logger.log(f"💾 Đã ghi tạm Excel ({len(results)} dòng output / {done_count} nhóm)")
                    except Exception as ex:
                        logger.log(f"⚠️ Lỗi ghi Excel: {ex}")

            return group_results

        # Chạy parallel theo NHÓM tenfile, không chạy theo từng dòng Excel nữa.
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {executor.submit(process_group, group_data): group_data for group_data in groups}
            for future in as_completed(futures):
                try:
                    future.result()
                except Exception as e:
                    logger.log(f"❌ Future exception: {e}")

        logger.log("=" * 90)
        logger.log(
            f"🎯 XONG THEO NHÓM. Thành công: {thanh_cong} | Bỏ qua: {bo_qua} | "
            f"Lỗi/Cần kiểm tra: {that_bai} | Tổng nhóm: {tong} | Tổng dòng output: {len(results)}"
        )
        logger.log("📄 File kết quả: " + output_path)

    except Exception as e:
        logger.log("❌ Lỗi chương trình: " + str(e))
        traceback.print_exc()

    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass

        log_queue.put("__DONE__")


# =========================
# TKINTER APP
# =========================

class App(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Cập nhật hồ sơ quét — API SearchHoSoQuet (parallel)")
        self.geometry("1080x800")

        self.log_queue     = queue.Queue()
        self.worker_thread = None

        self.var_username  = tk.StringVar()
        self.var_password  = tk.StringVar()
        self.var_maxa      = tk.StringVar()
        self.var_ngay_dk   = tk.StringVar()
        self.var_excel     = tk.StringVar()
        self.var_folder    = tk.StringVar()
        self.var_workers   = tk.IntVar(value=MAX_WORKERS)
        self.var_dry_run   = tk.BooleanVar(value=DRY_RUN)

        # Progress state
        self._total         = 0
        self._pb_total      = None   # ttk.Progressbar tổng
        self._lbl_total     = None   # label tổng
        self._worker_frames = []     # list of (frame, pb, lbl) cho từng worker

        self.create_widgets()
        self.after(200, self.process_log_queue)

    def create_widgets(self):
        # ── Form ──────────────────────────────────────────────────────────
        frame_top = ttk.LabelFrame(self, text="Thông tin chạy")
        frame_top.pack(fill="x", padx=10, pady=(10, 4))

        ttk.Label(frame_top, text="Username").grid(row=0, column=0, padx=5, pady=4, sticky="w")
        ttk.Entry(frame_top, textvariable=self.var_username, width=30).grid(row=0, column=1, padx=5, pady=4)
        ttk.Label(frame_top, text="Password").grid(row=0, column=2, padx=5, pady=4, sticky="w")
        ttk.Entry(frame_top, textvariable=self.var_password, width=30, show="*").grid(row=0, column=3, padx=5, pady=4)

        ttk.Label(frame_top, text="Mã xã").grid(row=1, column=0, padx=5, pady=4, sticky="w")
        ttk.Entry(frame_top, textvariable=self.var_maxa, width=20).grid(row=1, column=1, padx=5, pady=4, sticky="w")
        ttk.Label(frame_top, text="Ngày ĐK lần đầu").grid(row=1, column=2, padx=5, pady=4, sticky="w")
        ttk.Entry(frame_top, textvariable=self.var_ngay_dk, width=16).grid(row=1, column=3, padx=5, pady=4, sticky="w")
        ttk.Label(frame_top, text="dd/mm/yyyy", foreground="gray").grid(row=1, column=4, padx=0, pady=4, sticky="w")

        ttk.Label(frame_top, text="Số luồng (workers)").grid(row=2, column=0, padx=5, pady=4, sticky="w")
        ttk.Spinbox(frame_top, textvariable=self.var_workers, from_=1, to=10, width=6).grid(row=2, column=1, padx=5, pady=4, sticky="w")
        ttk.Checkbutton(frame_top, text="DRY RUN (chỉ kiểm tra, không update)", variable=self.var_dry_run).grid(
            row=2, column=2, columnspan=3, padx=10, pady=4, sticky="w"
        )

        ttk.Label(frame_top, text="File Excel").grid(row=3, column=0, padx=5, pady=4, sticky="w")
        ttk.Entry(frame_top, textvariable=self.var_excel, width=90).grid(row=3, column=1, columnspan=3, padx=5, pady=4, sticky="we")
        ttk.Button(frame_top, text="Duyệt", command=self.browse_excel).grid(row=3, column=4, padx=5, pady=4)

        ttk.Label(frame_top, text="Folder PDF").grid(row=4, column=0, padx=5, pady=4, sticky="w")
        ttk.Entry(frame_top, textvariable=self.var_folder, width=90).grid(row=4, column=1, columnspan=3, padx=5, pady=4, sticky="we")
        ttk.Button(frame_top, text="Duyệt", command=self.browse_folder).grid(row=4, column=4, padx=5, pady=4)

        self.btn_start = ttk.Button(frame_top, text="▶  BẮT ĐẦU CHẠY", command=self.start_run)
        self.btn_start.grid(row=5, column=1, padx=5, pady=8, sticky="w")
        self.btn_clear = ttk.Button(frame_top, text="Xóa log", command=self.clear_log)
        self.btn_clear.grid(row=5, column=2, padx=5, pady=8, sticky="w")

        ttk.Label(frame_top, text="Excel cần cột: soto | sothua | loaidat | tenfile",
                  foreground="blue").grid(row=6, column=0, columnspan=5, padx=5, pady=3, sticky="w")
        frame_top.columnconfigure(3, weight=1)

        # ── Progress section ───────────────────────────────────────────────
        self.frame_progress = ttk.LabelFrame(self, text="Tiến độ")
        self.frame_progress.pack(fill="x", padx=10, pady=4)

        # Progress bar tổng
        ttk.Label(self.frame_progress, text="Tổng nhóm:", width=12, anchor="w").grid(
            row=0, column=0, padx=6, pady=4, sticky="w")
        self._pb_total = ttk.Progressbar(self.frame_progress, length=600, mode="determinate")
        self._pb_total.grid(row=0, column=1, padx=6, pady=4, sticky="we")
        self._lbl_total = ttk.Label(self.frame_progress, text="0 / 0  |  ✅0  ⏭0  ❌0", width=32)
        self._lbl_total.grid(row=0, column=2, padx=6, pady=4, sticky="w")
        self.frame_progress.columnconfigure(1, weight=1)

        # Worker rows sẽ được tạo động khi nhận message "init"
        self._worker_rows = []

        # ── Log area ───────────────────────────────────────────────────────
        frame_log = ttk.LabelFrame(self, text="Log xử lý")
        frame_log.pack(fill="both", expand=True, padx=10, pady=(4, 10))

        self.txt_log = tk.Text(frame_log, wrap="word", font=("Consolas", 9))
        self.txt_log.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(frame_log, command=self.txt_log.yview)
        scrollbar.pack(side="right", fill="y")
        self.txt_log.configure(yscrollcommand=scrollbar.set)

    def _init_progress(self, total, n_workers):
        """Tạo progress bars khi biết tổng dòng và số worker."""
        self._total = total

        # Reset tổng
        self._pb_total.config(maximum=total, value=0)
        self._lbl_total.config(text=f"0 / {total}  |  ✅0  ⏭0  ❌0")

        # Xóa worker rows cũ nếu có
        for (lbl, pb, lbl_status) in self._worker_rows:
            lbl.grid_forget()
            pb.grid_forget()
            lbl_status.grid_forget()
        self._worker_rows.clear()

        # Tạo rows mới
        for i in range(n_workers):
            row_idx = i + 1
            lbl = ttk.Label(self.frame_progress, text=f"Worker {i+1}:", width=12, anchor="w")
            lbl.grid(row=row_idx, column=0, padx=6, pady=2, sticky="w")

            pb = ttk.Progressbar(self.frame_progress, length=600, mode="determinate", maximum=100)
            pb.grid(row=row_idx, column=1, padx=6, pady=2, sticky="we")

            lbl_status = ttk.Label(self.frame_progress, text="Chờ...", width=32, foreground="gray")
            lbl_status.grid(row=row_idx, column=2, padx=6, pady=2, sticky="w")

            self._worker_rows.append((lbl, pb, lbl_status))

        # Worker progress chỉ là % của dòng đang xử lý, không phải % toàn trình.

    def browse_excel(self):
        path = filedialog.askopenfilename(
            title="Chọn file Excel",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")]
        )
        if path:
            self.var_excel.set(path)

    def browse_folder(self):
        folder = filedialog.askdirectory(title="Chọn folder chứa PDF")
        if folder:
            self.var_folder.set(folder)

    def clear_log(self):
        self.txt_log.delete("1.0", tk.END)

    def validate_input(self):
        if not self.var_username.get().strip():
            messagebox.showerror("Thiếu thông tin", "Chưa nhập username.")
            return False
        if not self.var_password.get().strip():
            messagebox.showerror("Thiếu thông tin", "Chưa nhập password.")
            return False
        if not self.var_maxa.get().strip():
            messagebox.showerror("Thiếu thông tin", "Chưa nhập mã xã.")
            return False
        if not self.var_ngay_dk.get().strip():
            messagebox.showerror("Thiếu thông tin", "Chưa nhập ngày đăng ký lần đầu.")
            return False
        try:
            ddmmyyyy_to_iso_utc_start_of_day_vn(self.var_ngay_dk.get().strip())
        except Exception:
            messagebox.showerror("Sai định dạng", "Ngày đăng ký lần đầu phải là dd/mm/yyyy. Ví dụ: 23/04/2026")
            return False
        if not os.path.isfile(self.var_excel.get().strip()):
            messagebox.showerror("Sai đường dẫn", "File Excel không tồn tại.")
            return False
        if not os.path.isdir(self.var_folder.get().strip()):
            messagebox.showerror("Sai đường dẫn", "Folder PDF không tồn tại.")
            return False
        return True

    def start_run(self):
        if self.worker_thread and self.worker_thread.is_alive():
            messagebox.showwarning("Đang chạy", "Chương trình đang chạy.")
            return
        if not self.validate_input():
            return

        global MAX_WORKERS, DRY_RUN
        MAX_WORKERS = self.var_workers.get()
        DRY_RUN     = self.var_dry_run.get()

        self.btn_start.config(state="disabled")

        self.worker_thread = threading.Thread(
            target=worker_run,
            args=(
                self.var_username.get().strip(),
                self.var_password.get().strip(),
                self.var_maxa.get().strip(),
                self.var_ngay_dk.get().strip(),
                self.var_excel.get().strip(),
                self.var_folder.get().strip(),
                self.log_queue,
            ),
            daemon=True
        )
        self.worker_thread.start()

    def process_log_queue(self):
        try:
            while True:
                msg = self.log_queue.get_nowait()

                # ── Structured messages từ worker ──
                if isinstance(msg, dict):
                    t = msg.get("type")

                    if t == "init":
                        self._init_progress(msg["total"], msg["workers"])

                    elif t == "worker_row_progress":
                        w = msg["worker"] - 1   # 0-based index
                        row = msg["row"]
                        percent = int(msg.get("percent", 0))
                        text = msg.get("text", "")

                        # Chặn giá trị ngoài 0..100
                        percent = max(0, min(100, percent))

                        if w < len(self._worker_rows):
                            _, pb, lbl_s = self._worker_rows[w]
                            pb.config(value=percent)

                            if percent >= 100:
                                lbl_s.config(
                                    text=f"✅ Dòng {row} — {percent}% — {text}",
                                    foreground="green"
                                )
                            else:
                                lbl_s.config(
                                    text=f"⏳ Dòng {row} — {percent}% — {text}",
                                    foreground="blue"
                                )

                    elif t == "progress":
                        done  = msg["done"]
                        total = msg["total"]
                        tc    = msg["thanh_cong"]
                        bq    = msg["bo_qua"]
                        tb    = msg["that_bai"]
                        self._pb_total.config(value=done)
                        pct = int(done / total * 100) if total else 0
                        self._lbl_total.config(
                            text=f"{done} / {total} ({pct}%)  |  ✅{tc}  ⏭{bq}  ❌{tb}"
                        )
                    continue

                # ── Plain string: log text ──
                if msg == "__DONE__":
                    self.btn_start.config(state="normal")
                    # Reset tất cả worker label về trạng thái chờ
                    for (_, pb, lbl_s) in self._worker_rows:
                        lbl_s.config(text="Xong", foreground="gray")
                    self.txt_log.insert(tk.END, "\n✅ Tiến trình đã kết thúc.\n")
                    self.txt_log.see(tk.END)
                    continue

                self.txt_log.insert(tk.END, msg + "\n")
                self.txt_log.see(tk.END)

        except queue.Empty:
            pass

        self.after(200, self.process_log_queue)


if __name__ == "__main__":
    app = App()
    app.mainloop()
