# -*- coding: utf-8 -*-
"""
Tool cập nhật Thông tin đăng ký (VBDLIS / MPLIS) — KHÔNG upload file.

Quy tắc:
  (C) Mã định danh: CCCD 12 số -> ghi maSoDinhDanh (cấp cá nhân) + maDinhDanhCaNhan (giấy tờ CCCD).
  (D) Nguồn gốc thửa đất: ONT/ODT -> loaiNguonGocSuDungDatId = 2 (có thu tiền), còn lại -> 3.
  (E) Giấy chứng nhận: ngayVaoSo trống -> lấy của GCN khác / ngày tạo đơn; soVaoSo trống -> "--\\--".
      Số/ngày vào sổ gửi qua endpoint delta (CapNhatThongKePhanLoaiThuaDatChiTiet).
  (F) Hồ sơ quét nguồn gốc: tích laGiayChungNhan + laGiayToVeNguonGoc cho 1 file phù hợp.

CHẠY THẬT (DRY_RUN=False).
"""

import os
import re
import json
import copy
import queue
import threading
import traceback
import time
import uuid
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime, timezone, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from openpyxl import load_workbook, Workbook

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys


# =========================
# API CONFIG
# =========================

REFERER_URL = "https://dla.mplis.gov.vn/dc/DonDangKy/KeKhaiDangKyV2"

API_SEARCH_HOSOQUET          = "https://dla.mplis.gov.vn/dc/QuanLyKhoHoSoQuetAjax/SearchHoSoQuet"
URL_GET_THONG_TIN_DANG_KY    = "https://dla.mplis.gov.vn/dc/DangKyAjax/GetThongTinDangKyByTinhHinhDangKyIds"
URL_UPDATE_THONG_TIN_DANG_KY = "https://dla.mplis.gov.vn/dc/DangKyAjax/UpdateThongTinDangKy"
URL_UPDATE_CA_NHAN           = "https://dla.mplis.gov.vn/dc/ChuSoHuuAjax/UpdateCaNhan"
URL_UPDATE_THUA_DAT          = "https://dla.mplis.gov.vn/dc/TaiSanAjax/UpdateThuaDat"
API_UPDATE_HOSOQUET          = "https://dla.mplis.gov.vn/dc/HoSoQuetAjax/UpdateHoSoQuetExistFile"
URL_UPDATE_DELTA_GCN         = "https://dla.mplis.gov.vn/dc/LamSachDuLieuAjax/CapNhatThongKePhanLoaiThuaDatChiTiet"
URL_PHAN_LOAI_LAI            = "https://dla.mplis.gov.vn/dc/LamSachDuLieuAjax/GuiYeuCauPhanLoaiLai"

DRY_RUN     = False         # False = gửi cập nhật thật
MAX_WORKERS = 3
API_SEARCH_TIMEOUT = 120
API_UPDATE_TIMEOUT = 180
API_SEARCH_RETRIES = 3
API_RETRY_BACKOFF_SECONDS = 5
WRITE_EVERY_N = 10

# Bật/tắt từng quy tắc (UI override). Đây chỉ là default.
RULE_MA_DINH_DANH    = True   # (C)
RULE_NGUON_GOC_THUA  = True   # (D)
RULE_GIAY_CHUNG_NHAN = True   # (E)
RULE_HSQ_NGUON_GOC   = True   # (F) Cập nhật HSQ luôn, không cần tick chọn


# =========================
# LOG REDIRECT (thread-safe)
# =========================

class ThreadSafeLogger:
    def __init__(self, log_queue):
        self.log_queue = log_queue
        self._lock = threading.Lock()

    def log(self, msg):
        if msg and str(msg).strip():
            with self._lock:
                self.log_queue.put(str(msg))

    def write(self, text):
        self.log(text)

    def flush(self):
        pass


# =========================
# EXCEL
# =========================

def doc_excel(path_excel):
    wb = load_workbook(path_excel, data_only=True)
    ws = wb.active

    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value:
            headers[str(value).strip().lower()] = col

    required_cols = ["soto", "sothua"]
    missing = [c for c in required_cols if c not in headers]
    if missing:
        raise RuntimeError(f"Thiếu cột trong Excel: {', '.join(missing)}")

    data = []
    for row in range(2, ws.max_row + 1):
        soto    = ws.cell(row=row, column=headers["soto"]).value
        sothua  = ws.cell(row=row, column=headers["sothua"]).value
        loaidat = ws.cell(row=row, column=headers["loaidat"]).value if "loaidat" in headers else ""

        if not soto and not sothua:
            continue

        data.append({
            "row":     row,
            "soto":    str(soto).strip()    if soto    is not None else "",
            "sothua":  str(sothua).strip()  if sothua  is not None else "",
            "loaidat": str(loaidat).strip() if loaidat is not None else "",
        })
    return data


def ghi_excel_output(rows, output_path, lock):
    with lock:
        wb = Workbook()
        ws = wb.active
        ws.title = "KetQua"

        headers = [
            "STT", "Dòng Excel", "Số tờ", "Số thửa", "Loại đất",
            "tinhHinhDangKyId", "hoSoQuetId", "thongTinHoSoId",
            "Chủ sử dụng", "Trạng thái", "Ghi chú"
        ]
        ws.append(headers)

        sorted_rows = sorted(rows, key=lambda x: x.get("row_excel") or 0)
        for i, r in enumerate(sorted_rows, start=1):
            ws.append([
                i,
                r.get("row_excel"),
                r.get("soto"),
                r.get("sothua"),
                r.get("loaidat"),
                r.get("tinhHinhDangKyId"),
                r.get("hoSoQuetId"),
                r.get("thongTinHoSoId"),
                r.get("chu_su_dung"),
                r.get("status"),
                r.get("note"),
            ])

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

        wb.save(output_path)


# =========================
# SELENIUM LOGIN + TOKEN
# =========================

def lay_token_tu_trang(driver):
    js = """
    return (
        document.querySelector('input[name="__RequestVerificationToken"]')?.value ||
        document.querySelector('input[name="__requestverificationtoken"]')?.value ||
        document.querySelector('meta[name="__RequestVerificationToken"]')?.content ||
        document.querySelector('meta[name="__requestverificationtoken"]')?.content ||
        document.querySelector('meta[name="RequestVerificationToken"]')?.content ||
        ''
    );
    """
    return driver.execute_script(js)


def tao_session_tu_selenium(driver):
    session = requests.Session()
    user_agent = driver.execute_script("return navigator.userAgent;")
    token = lay_token_tu_trang(driver)
    if not token:
        raise RuntimeError("Không lấy được __requestverificationtoken.")

    session.headers.update({
        "User-Agent": user_agent,
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "X-Requested-With": "XMLHttpRequest",
        "Origin": "https://dla.mplis.gov.vn",
        "Referer": REFERER_URL,
        "__requestverificationtoken": token,
        "__RequestVerificationToken": token,
        "RequestVerificationToken": token,
    })
    for c in driver.get_cookies():
        session.cookies.set(name=c["name"], value=c["value"],
                            domain=c.get("domain"), path=c.get("path", "/"))
    return session


def login_mplis(username, password):
    options = Options()
    options.add_argument("--start-maximized")
    options.add_argument("--window-position=100,100")
    options.add_argument("--window-size=1400,900")
    driver_path = r"D:\Soft\CHROMEDRIVER\chromedriver-win64\chromedriver-win64\chromedriver.exe"

    service = Service(executable_path=driver_path)
    driver = webdriver.Chrome(options=options, service=service)
    driver.get(REFERER_URL)
    try:
        time.sleep(2)
        inputs = driver.find_elements(By.CSS_SELECTOR, "input")
        user_box = None
        pass_box = None
        for inp in inputs:
            typ         = (inp.get_attribute("type")        or "").lower()
            name        = (inp.get_attribute("name")        or "").lower()
            placeholder = (inp.get_attribute("placeholder") or "").lower()
            input_id    = (inp.get_attribute("id")          or "").lower()
            text_all    = f"{name} {placeholder} {input_id}"
            if not user_box and typ in ["text", "email"] and any(
                k in text_all for k in ["user", "username", "login", "account", "ten"]
            ):
                user_box = inp
            if not pass_box and typ == "password":
                pass_box = inp
        if not user_box:
            for inp in inputs:
                if (inp.get_attribute("type") or "").lower() in ["text", "email"]:
                    user_box = inp
                    break
        if user_box and pass_box:
            user_box.clear(); user_box.send_keys(username)
            pass_box.clear(); pass_box.send_keys(password); pass_box.send_keys(Keys.ENTER)
    except Exception:
        pass

    messagebox.showinfo(
        "Đăng nhập / Authenticator",
        "Hoàn tất đăng nhập MPLIS và xác thực Authenticator trên Chrome.\n"
        "Sau khi vào được màn hình kê khai đăng ký thì bấm OK để tiếp tục."
    )
    return driver


# =========================
# DATE HELPERS
# =========================

def dotnet_date_to_iso(value):
    if not isinstance(value, str):
        return value
    m = re.search(r"/Date\((-?\d+)\)/", value)
    if not m:
        return value
    ms = int(m.group(1))
    dt = datetime.fromtimestamp(ms / 1000, tz=timezone.utc)
    return dt.strftime("%Y-%m-%dT%H:%M:%S.") + f"{int(dt.microsecond / 1000):03d}Z"


def convert_dates_recursive(obj):
    if isinstance(obj, dict):
        return {k: convert_dates_recursive(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [convert_dates_recursive(x) for x in obj]
    return dotnet_date_to_iso(obj)


def now_iso_z():
    return datetime.now(timezone.utc).isoformat(timespec="milliseconds").replace("+00:00", "Z")


def safe_int(value, default=0):
    try:
        if value is None or value == "":
            return default
        return int(value)
    except Exception:
        return default


# =========================
# (C) CẬP NHẬT MÃ ĐỊNH DANH — đệ quy, mọi loại chủ
# =========================

def _is_person(node):
    return isinstance(node, dict) and "caNhanId" in node and "hoTen" in node


def iter_persons(obj):
    if isinstance(obj, dict):
        if _is_person(obj):
            yield obj
        for v in obj.values():
            yield from iter_persons(v)
    elif isinstance(obj, list):
        for x in obj:
            yield from iter_persons(x)


def cap_nhat_ma_dinh_danh(person, notes, ghi_note=True):
    """
    CCCD (loai 5) 12 số -> ghi:
      - CaNhan.maSoDinhDanh
      - ListGiayToTuyThan CCCD.maDinhDanhCaNhan
    CMND 9 số / không có CCCD -> bỏ qua + ghi chú.
    """
    ds = person.get("ListGiayToTuyThan") or []

    cccd = None
    for g in ds:
        so_tmp = re.sub(r"\D", "", str(g.get("soGiayTo") or ""))
        if str(g.get("loaiGiayToTuyThanId")) == "5" and len(so_tmp) == 12:
            cccd = g
            break

    ten = person.get("hoTen") or person.get("caNhanId")

    if not cccd:
        so_person = re.sub(r"\D", "", str(person.get("maSoDinhDanh") or ""))
        if len(so_person) == 12:
            person["maSoDinhDanh"] = so_person
            person["isChange"] = True
            if ghi_note:
                notes.append(f"Giữ mã định danh đã có - {ten}: {so_person}; không có CCCD loại 5 để ghi maDinhDanhCaNhan")
        else:
            if ghi_note:
                notes.append(f"Không cập nhật mã định danh - {ten}: không có CCCD loại 5 đủ 12 số")
        return

    so = re.sub(r"\D", "", str(cccd.get("soGiayTo") or ""))

    person["maSoDinhDanh"] = so
    person["isChange"] = True
    person["isNew"] = person.get("isNew", True)

    cccd["maDinhDanhCaNhan"] = so
    cccd["daDinhDanh"] = True
    cccd["isChange"] = True

    if ghi_note:
        notes.append(f"Cập nhật mã định danh CCCD - {ten}: {so}")


def cap_nhat_thong_tin_ca_nhan(payload, notes=None, do_ma_dinh_danh=True):
    notes = notes if notes is not None else []
    seen_note = set()
    for person in iter_persons(payload):
        cid = person.get("caNhanId")
        if do_ma_dinh_danh:
            cap_nhat_ma_dinh_danh(person, notes, ghi_note=(cid not in seen_note))
            seen_note.add(cid)
    return notes


# =========================
# (D) NGUỒN GỐC THỬA ĐẤT
# =========================

NGUON_GOC_TEXT = {
    2: "Công nhận QSDĐ như giao đất có thu tiền sử dụng đất",
    3: "Công nhận QSDĐ như giao đất không thu tiền sử dụng đất",
}
NGUON_GOC_MA = {2: "CNQ-CTT", 3: "CNQ-KTT"}
ONT_ODT = {"ONT", "ODT"}


def _nguon_goc_id_theo_mucdich(loai_mdsd):
    return 2 if (loai_mdsd or "").upper() in ONT_ODT else 3


def _set_nguon_goc_cho_mucdich(md):
    ngid = _nguon_goc_id_theo_mucdich(md.get("loaiMucDichSuDungId"))
    for ng in md.get("ListNguonGocSuDungDat") or []:
        ng["loaiNguonGocSuDungDatId"] = ngid
        sub = ng.get("LoaiNguonGocSuDungDat")
        if isinstance(sub, dict):
            sub["loaiNguonGocSuDungDatId"]   = ngid
            sub["maLoaiNguonGocSuDungDat"]   = NGUON_GOC_MA[ngid]
            sub["tenLoaiNguonGocInGiay"]     = NGUON_GOC_TEXT[ngid]
            sub["tenLoaiNguonGocSoDiaChinh"] = NGUON_GOC_TEXT[ngid]
        ng["isChange"] = True


def _iter_mucdich(obj):
    if isinstance(obj, dict):
        if "loaiMucDichSuDungId" in obj and "ListNguonGocSuDungDat" in obj:
            yield obj
        for v in obj.values():
            yield from _iter_mucdich(v)
    elif isinstance(obj, list):
        for x in obj:
            yield from _iter_mucdich(x)


def cap_nhat_nguon_goc_thua_dat(payload):
    n = 0
    for md in _iter_mucdich(payload):
        _set_nguon_goc_cho_mucdich(md)
        n += 1
    return n


# =========================
# (E) GIẤY CHỨNG NHẬN: số/ngày vào sổ
# =========================

SO_VAO_SO_DEFAULT = "--\\--"   # in ra: --\--


def cap_nhat_giay_chung_nhan(ttdk, ngay_tao_don_iso, notes):
    """
    - ngayVaoSo trống: lấy ngày vào sổ của GCN khác trong cùng list; không có thì lấy ngày tạo đơn.
    - soVaoSo trống: ghi --\\--.
    Ghi marker _delta_gcn_changed_fields để build_delta_gcn_payload đẩy đúng trường đã đổi.
    """
    ds_gcn = ttdk.get("ListGiayChungNhan") or []
    if not ds_gcn:
        return

    ngay_co_san = None
    for gcn in ds_gcn:
        if gcn.get("ngayVaoSo"):
            ngay_co_san = gcn.get("ngayVaoSo")
            break
    ngay_fallback = ngay_co_san or ngay_tao_don_iso

    for gcn in ds_gcn:
        changed_fields = []

        if not gcn.get("ngayVaoSo"):
            gcn["ngayVaoSo"] = ngay_fallback
            changed_fields.append("ngayVaoSo")

        if not gcn.get("soVaoSo"):
            gcn["soVaoSo"] = SO_VAO_SO_DEFAULT
            changed_fields.append("soVaoSo")

        if changed_fields:
            gcn["isChange"] = True
            gcn["_delta_gcn_changed_fields"] = changed_fields
            notes.append(
                "Cập nhật lại thông tin GCN "
                f"(id={gcn.get('giayChungNhanId')}, "
                f"soPhatHanh={gcn.get('soPhatHanh')}, "
                f"trường={','.join(changed_fields)})"
            )


# =========================
# (F) HỒ SƠ QUÉT: đánh dấu node nguồn gốc / giấy chứng nhận
# =========================

# Số phát hành GCN: 1-2 chữ cái IN HOA + (cách/underscore tùy chọn) + 6-8 số.
_RE_SO_GCN = re.compile(r"(?<![A-Z])([A-Z]{1,2})[\s_]?(\d{6,8})(?!\d)")


def _diem_uu_tien_node_gcn(moTa):
    s = (moTa or "").upper()
    if _RE_SO_GCN.search(s):
        return 2
    if "GIẤY CHỨNG NHẬN" in s or re.search(r"\bGCN\b", s):
        return 1
    return 0


def cap_nhat_hosoquet_nguon_goc(list_file_hsq, notes):
    """
    Nếu đã có CÙNG MỘT file vừa là GCN vừa là giấy tờ nguồn gốc thì bỏ qua.
    Nếu chưa: ưu tiên file đang là GCN; chưa có thì tìm theo mô tả/mã phát hành GCN.
    Set 2 cờ trên chính file đó.
    """
    if not list_file_hsq:
        notes.append("HSQ không có file")
        return False

    for f in list_file_hsq:
        if f.get("laGiayChungNhan") is True and f.get("laGiayToVeNguonGoc") is True:
            notes.append("HSQ đã có file vừa là GCN vừa là giấy tờ nguồn gốc, bỏ qua")
            return False

    candidate = next((f for f in list_file_hsq if f.get("laGiayChungNhan") is True), None)

    if candidate is None:
        best_score = -1
        for f in list_file_hsq:
            sc = _diem_uu_tien_node_gcn(f.get("moTa"))
            if sc > 0 and sc >= best_score:
                best_score = sc
                candidate = f

    if candidate is None:
        notes.append("Không tìm thấy file có GCN hoặc mã GCN trong HSQ")
        return False

    candidate["laGiayChungNhan"] = True
    candidate["laGiayToVeNguonGoc"] = True
    candidate["loaiHoSoQuet"] = 1
    candidate["files"] = None

    notes.append(f"Cập nhật lại thông tin HSQ: tích GCN + giấy tờ nguồn gốc cho '{candidate.get('moTa')}'")
    return True


def api_update_hosoquet(session, hoso, debug_dir, row_excel, notes):
    """
    Cập nhật metadata HSQ qua UpdateHoSoQuetExistFile (form-urlencoded):
      hoSoQuet=<json core>, infoHoSoQuet_1..N=<json node>, count=<N>
    """
    list_file = (hoso.get("ListFileHoSoQuet") or {}).get("ListFileHoSoQuet") or []
    if not list_file:
        return {"ok": False, "error": "HSQ không có ListFileHoSoQuet", "notes": notes}

    hoso_core = copy.deepcopy(hoso)
    hoso_core.pop("ListFileHoSoQuet", None)

    data = {
        "hoSoQuet": json.dumps(hoso_core, ensure_ascii=False),
        "count": str(len(list_file)),
    }

    for i, node in enumerate(list_file, start=1):
        n = copy.deepcopy(node)
        n["files"] = None
        data[f"infoHoSoQuet_{i}"] = json.dumps(n, ensure_ascii=False)

    try:
        os.makedirs(debug_dir, exist_ok=True)
        hid = hoso.get("hoSoQuetId") or hoso.get("Title") or "unknown"
        with open(os.path.join(debug_dir, f"HSQ_UPDATE_FORM_{hid}_row{row_excel}.json"),
                  "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

    if DRY_RUN:
        return {"ok": True, "dry_run": True, "notes": notes}

    headers = dict(session.headers)
    headers.update({
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": REFERER_URL,
        "Origin": "https://dla.mplis.gov.vn",
    })
    headers.pop("Content-Type", None)

    try:
        res = session.post(API_UPDATE_HOSOQUET, data=data, headers=headers, timeout=API_UPDATE_TIMEOUT)
    except requests.exceptions.RequestException as e:
        return {"ok": False, "error": f"Lỗi request HSQ: {e}", "notes": notes}

    text = res.text[:1000]
    try:
        js = res.json()
    except Exception:
        return {"ok": False, "error": text, "notes": notes}

    if res.ok and _response_update_ok(js):
        notes.append("Update HSQ OK")
        return {"ok": True, "raw": js, "notes": notes}

    return {"ok": False, "error": str(js)[:1000], "notes": notes}


# =========================
# BUILD PAYLOAD UPDATE THÔNG TIN ĐĂNG KÝ
# =========================

def _lay_ngay_tao_don(ttdk):
    """Ưu tiên CreatedDate của TinhHinhDangKy, fallback thoiDiemDangKy / ngayTiepNhan."""
    thd = ttdk.get("TinhHinhDangKy") or {}
    for k in ("CreatedDate", "thoiDiemDangKy", "ngayTiepNhan", "ModifiedDate"):
        v = thd.get(k)
        if v:
            return v
    return now_iso_z()


def build_payload_update(response_json, flags=None):
    """Trả về (thong_tin_dang_ky, notes)."""
    flags = flags or {}
    if not response_json.get("value"):
        raise Exception("Response không có value để build payload")

    value0 = copy.deepcopy(response_json["value"][0])
    if isinstance(value0, dict) and isinstance(value0.get("thongTinDangKy"), dict):
        ttdk = value0["thongTinDangKy"]
    else:
        ttdk = value0

    ttdk = convert_dates_recursive(ttdk)

    if "TinhHinhDangKy" not in ttdk or not isinstance(ttdk["TinhHinhDangKy"], dict):
        raise Exception("Payload không có TinhHinhDangKy.")

    notes = []
    ngay_tao_don_iso = _lay_ngay_tao_don(ttdk)

    if flags.get("RULE_MA_DINH_DANH", True):
        cap_nhat_thong_tin_ca_nhan(ttdk, notes=notes, do_ma_dinh_danh=True)

    if flags.get("RULE_NGUON_GOC_THUA", True):
        cap_nhat_nguon_goc_thua_dat(ttdk)

    if flags.get("RULE_GIAY_CHUNG_NHAN", True):
        cap_nhat_giay_chung_nhan(ttdk, ngay_tao_don_iso, notes)

    return ttdk, notes


# =========================
# API: GET + UPDATE THÔNG TIN ĐĂNG KÝ
# =========================

def post_retry(session, url, *, retries, timeout, **kwargs):
    last_error = None
    for attempt in range(1, retries + 1):
        try:
            return session.post(url, timeout=timeout, **kwargs), None
        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
            last_error = e
            if attempt < retries:
                time.sleep(API_RETRY_BACKOFF_SECONDS * attempt)
        except requests.exceptions.RequestException as e:
            return None, f"Loi request API: {e}"
    return None, f"Timeout/loi mang sau {retries} lan thu: {last_error}"


def lay_thong_tin_dang_ky_by_id(session, tinh_hinh_dang_ky_id):
    tid = safe_int(tinh_hinh_dang_ky_id, 0)
    if not tid:
        return {"ok": False, "error": f"tinhHinhDangKyId không hợp lệ: {tinh_hinh_dang_ky_id}"}

    headers = dict(session.headers)
    headers.update({
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Content-Type": "application/json; charset=UTF-8",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": REFERER_URL,
        "Origin": "https://dla.mplis.gov.vn",
    })

    payload = {"tinhHinhDangKyIds": [tid], "getHoSoQuet": False}

    res, request_error = post_retry(
        session, URL_GET_THONG_TIN_DANG_KY,
        retries=API_SEARCH_RETRIES, timeout=API_SEARCH_TIMEOUT,
        data=json.dumps(payload, ensure_ascii=False), headers=headers
    )
    if request_error:
        return {"ok": False, "error": request_error}

    try:
        js = res.json()
    except Exception:
        return {"ok": False, "error": f"GET không phải JSON: {res.text[:500]}",
                "status_code": res.status_code}

    if not js.get("success"):
        return {"ok": False, "error": f"GET success=false (id={tid}): {str(js)[:500]}"}
    if not js.get("value"):
        return {"ok": False, "error": f"GET không có value (id={tid})"}

    return {"ok": True, "raw": js}


def _response_update_ok(js):
    if not isinstance(js, dict):
        return False
    if js.get("success") is True or js.get("Success") is True or js.get("ok") is True:
        return True
    if "success" not in js and "error" not in js and "Error" not in js:
        return True
    return False


def _post_json_update(session, url, payload, ten_api):
    headers = dict(session.headers)
    headers.update({
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Content-Type": "application/json; charset=UTF-8",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": REFERER_URL,
        "Origin": "https://dla.mplis.gov.vn",
    })
    try:
        res = session.post(url, data=json.dumps(payload, ensure_ascii=False), headers=headers, timeout=API_UPDATE_TIMEOUT)
    except requests.exceptions.RequestException as e:
        return {"ok": False, "error": f"Lỗi request {ten_api}: {e}"}
    text = res.text[:1000]
    try:
        js = res.json()
    except Exception:
        return {"ok": False, "error": text, "status_code": res.status_code}
    if res.ok and _response_update_ok(js):
        return {"ok": True, "raw": js}
    return {"ok": False, "error": str(js)[:1000], "raw": js}


def api_update_ca_nhan(session, person):
    return _post_json_update(session, URL_UPDATE_CA_NHAN, person, "UpdateCaNhan")


def api_update_thua_dat(session, thua_dat):
    return _post_json_update(session, URL_UPDATE_THUA_DAT, thua_dat, "UpdateThuaDat")


def build_delta_gcn_payload(ttdk):
    """
    Build delta cho số/ngày vào sổ GCN, ĐỌC marker _delta_gcn_changed_fields
    (do cap_nhat_giay_chung_nhan ghi) thay vì check lại giá trị rỗng —
    vì giá trị đã bị set trước đó nên check rỗng sẽ luôn sai.
    """
    thdk = ttdk.get("TinhHinhDangKy") or {}
    tid = thdk.get("tinhHinhDangKyId")

    data = {}
    ds_gcn = ttdk.get("ListGiayChungNhan") or []

    ngay_co_san = next((g.get("ngayVaoSo") for g in ds_gcn if g.get("ngayVaoSo")), None)
    ngay_fallback = ngay_co_san or _lay_ngay_tao_don(ttdk)

    if not thdk.get("thoiDiemDangKy"):
        thdk["thoiDiemDangKy"] = ngay_fallback
        data[f"TINHHINHDANGKY.{tid}|thoiDiemDangKy"] = ngay_fallback

    for gcn in ds_gcn:
        gid = gcn.get("giayChungNhanId")
        ver = gcn.get("version")
        if not gid or not ver:
            continue

        changed = gcn.get("_delta_gcn_changed_fields") or []
        if not changed:
            continue

        prefix = f"TINHHINHDANGKY.{tid}|GIAYCHUNGNHAN.{gid}_{ver}"
        for field in changed:                 # 'soVaoSo' và/hoặc 'ngayVaoSo'
            data[f"{prefix}|{field}"] = gcn.get(field)

    return {
        "id": uuid.uuid4().hex[:24],
        "tinhHinhDangKyId": tid,
        "data": json.dumps(data, ensure_ascii=False)
    }


def clean_delta_gcn_markers(obj):
    if isinstance(obj, dict):
        obj.pop("_delta_gcn_changed_fields", None)
        for v in obj.values():
            clean_delta_gcn_markers(v)
    elif isinstance(obj, list):
        for x in obj:
            clean_delta_gcn_markers(x)


def api_update_delta_gcn(session, delta_payload):
    return _post_json_update(session, URL_UPDATE_DELTA_GCN, delta_payload, "CapNhatThongKePhanLoaiThuaDatChiTiet")


def iter_unique_persons(ttdk):
    seen = set()
    for person in iter_persons(ttdk):
        key = (person.get("caNhanId"), person.get("version"), person.get("InId"))
        if key in seen:
            continue
        seen.add(key)
        yield person


def iter_unique_thua_dat(obj):
    seen = set()
    def walk(x):
        if isinstance(x, dict):
            if "thuaDatId" in x and "ListMucDichSuDung" in x:
                key = (x.get("thuaDatId"), x.get("version"), x.get("InId"))
                if key not in seen:
                    seen.add(key)
                    yield x
            for v in x.values():
                yield from walk(v)
        elif isinstance(x, list):
            for it in x:
                yield from walk(it)
    yield from walk(obj)


def get_thua_dat_ids_from_ttdk(ttdk):
    ids = []
    for thua in iter_unique_thua_dat(ttdk):
        tid = thua.get("thuaDatId")
        if tid and tid not in ids:
            ids.append(tid)
    return ids


def api_gui_yeu_cau_phan_loai_lai(session, thua_dat_ids):
    headers = dict(session.headers)
    headers.update({
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Content-Type": "application/json; charset=UTF-8",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": REFERER_URL,
        "Origin": "https://dla.mplis.gov.vn",
    })
    payload = {"thuaDatIds": [int(x) for x in thua_dat_ids if x]}
    if not payload["thuaDatIds"]:
        return {"ok": False, "error": "Không có thuaDatId để phân loại lại"}
    try:
        res = session.post(URL_PHAN_LOAI_LAI, data=json.dumps(payload, ensure_ascii=False),
                           headers=headers, timeout=API_UPDATE_TIMEOUT)
    except requests.exceptions.RequestException as e:
        return {"ok": False, "error": f"Lỗi request phân loại lại: {e}"}
    try:
        js = res.json()
    except Exception:
        return {"ok": False, "error": res.text[:1000]}
    return {"ok": res.ok and _response_update_ok(js), "raw": js}


def api_update_thong_tin_dang_ky(session, tinh_hinh_dang_ky_id, flags, debug_dir, row_excel):
    res_get = lay_thong_tin_dang_ky_by_id(session, tinh_hinh_dang_ky_id)
    if not res_get.get("ok"):
        return {"ok": False, "error": "GET TTĐK lỗi: " + res_get.get("error", ""), "notes": []}

    try:
        ttdk, notes = build_payload_update(res_get["raw"], flags=flags)
    except Exception as e:
        return {"ok": False, "error": "Build payload lỗi: " + str(e), "notes": []}

    os.makedirs(debug_dir, exist_ok=True)
    try:
        with open(os.path.join(debug_dir, f"TTDK_{tinh_hinh_dang_ky_id}_row{row_excel}.json"),
                  "w", encoding="utf-8") as f:
            json.dump(ttdk, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

    if DRY_RUN:
        return {"ok": True, "notes": notes, "dry_run": True}

    # 1) Số/ngày vào sổ GCN qua endpoint delta.
    delta_gcn_payload = build_delta_gcn_payload(ttdk)
    try:
        delta_gcn_data = json.loads(delta_gcn_payload.get("data") or "{}")
    except Exception:
        delta_gcn_data = {}

    if delta_gcn_data:
        try:
            with open(os.path.join(debug_dir, f"DELTA_GCN_{tinh_hinh_dang_ky_id}_row{row_excel}.json"),
                      "w", encoding="utf-8") as f:
                json.dump(delta_gcn_payload, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

        res_delta = api_update_delta_gcn(session, delta_gcn_payload)
        if not res_delta.get("ok"):
            return {"ok": False,
                    "error": "Update delta GCN lỗi: " + str(res_delta.get("raw") or res_delta.get("error")),
                    "notes": notes}
        notes.append("Đã update delta GCN: số/ngày vào sổ")

    clean_delta_gcn_markers(ttdk)

    # 2) Update từng cá nhân.
    for person in iter_unique_persons(ttdk):
        res_cn = api_update_ca_nhan(session, person)
        ten = person.get("hoTen") or person.get("caNhanId")
        if not res_cn.get("ok"):
            return {"ok": False, "error": f"Update cá nhân lỗi ({ten}): " + str(res_cn.get("raw") or res_cn.get("error")), "notes": notes}
        notes.append(f"Đã update cá nhân: {ten}")

    # 3) Update từng thửa đất.
    for thua in iter_unique_thua_dat(ttdk):
        res_td = api_update_thua_dat(session, thua)
        ma_thua = thua.get("maThua") or f"{thua.get('soHieuToBanDo')}/{thua.get('soThuTuThua')}"
        if not res_td.get("ok"):
            return {"ok": False, "error": f"Update thửa đất lỗi ({ma_thua}): " + str(res_td.get("raw") or res_td.get("error")), "notes": notes}
        notes.append(f"Đã update thửa đất: {ma_thua}")

    # 4) Update thông tin đăng ký tổng.
    headers_json = dict(session.headers)
    headers_json.update({
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Content-Type": "application/json; charset=UTF-8",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": REFERER_URL,
        "Origin": "https://dla.mplis.gov.vn",
    })

    payloads = [
        ("json_wrapper_thongTinDangKy", {"thongTinDangKy": ttdk}),
        ("json_raw_thongTinDangKy_core", ttdk),
    ]
    last_js, last_text = None, ""
    for mode, payload in payloads:
        try:
            res = session.post(URL_UPDATE_THONG_TIN_DANG_KY,
                               data=json.dumps(payload, ensure_ascii=False),
                               headers=headers_json, timeout=API_UPDATE_TIMEOUT)
        except requests.exceptions.RequestException as e:
            last_text = f"Lỗi request ({mode}): {e}"
            continue
        last_text = res.text[:1000]
        try:
            js = res.json()
        except Exception:
            js = None
        last_js = js
        if res.ok and _response_update_ok(js):
            notes.append(f"Update TTĐK OK ({mode})")

            # 5) Gửi yêu cầu phân loại lại thửa.
            thua_ids = get_thua_dat_ids_from_ttdk(ttdk)
            res_pl = api_gui_yeu_cau_phan_loai_lai(session, thua_ids)
            if not res_pl.get("ok"):
                notes.append("Gửi yêu cầu phân loại lại lỗi: " + str(res_pl.get("raw") or res_pl.get("error")))
            else:
                notes.append(f"Đã gửi yêu cầu phân loại lại thửa: {thua_ids}")

            return {"ok": True, "raw": js, "sent_as": mode, "notes": notes}

    return {"ok": False,
            "error": "UpdateTTĐK không thành công. Phản hồi cuối: "
                     f"{last_js if last_js is not None else last_text}",
            "notes": notes}


# =========================
# API: SEARCH HỒ SƠ QUÉT
# =========================

def tao_payload_search_hosoquet(xa_id, so_to, so_thua):
    return {
        "draw": "2",
        "columns[0][data]": "", "columns[0][name]": "", "columns[0][searchable]": "true",
        "columns[0][orderable]": "false", "columns[0][search][value]": "", "columns[0][search][regex]": "false",
        "columns[1][data]": "ListHoSoQuet", "columns[1][name]": "ListHoSoQuet", "columns[1][searchable]": "true",
        "columns[1][orderable]": "false", "columns[1][search][value]": "", "columns[1][search][regex]": "false",
        "columns[2][data]": "ThongTinDangKy", "columns[2][name]": "GiayChungNhan", "columns[2][searchable]": "true",
        "columns[2][orderable]": "false", "columns[2][search][value]": "", "columns[2][search][regex]": "false",
        "columns[3][data]": "ThongTinDangKy", "columns[3][name]": "ChuSoHuu", "columns[3][searchable]": "true",
        "columns[3][orderable]": "false", "columns[3][search][value]": "", "columns[3][search][regex]": "false",
        "columns[4][data]": "ThongTinDangKy", "columns[4][name]": "TaiSan", "columns[4][searchable]": "true",
        "columns[4][orderable]": "false", "columns[4][search][value]": "", "columns[4][search][regex]": "false",
        "start": "0", "length": "10", "search[value]": "", "search[regex]": "false",
        "xaId": str(xa_id), "huyenId": "0", "khoId": "0", "dayId": "0", "keId": "0", "hopId": "0",
        "maDon": "", "soThuTu": "", "ngayTiepNhan": "", "thoiDiemDangKy": "", "tuNgay": "", "denNgay": "",
        "loaiGiayChungNhanId": "", "maVach": "", "soPhatHanh": "", "soVaoSo": "", "soVaoSoCu": "",
        "ngayVaoSo": "", "soHoSoGoc": "", "soHoSoGocCu": "", "hoTen": "", "soGiayTo": "", "namSinh": "",
        "soThuTuThua": str(so_thua), "soHieuToBanDo": str(so_to),
        "soThuTuThuaCu": "", "soHieuToBanDoCu": "", "soNha": "", "diaChiChiTiet": "", "getHoSoLichSu": ""
    }


def api_search_hosoquet(session, xa_id, so_to, so_thua):
    headers = dict(session.headers)
    headers["Content-Type"] = "application/x-www-form-urlencoded; charset=UTF-8"
    res, request_error = post_retry(
        session, API_SEARCH_HOSOQUET,
        retries=API_SEARCH_RETRIES, timeout=API_SEARCH_TIMEOUT,
        data=tao_payload_search_hosoquet(xa_id, so_to, so_thua), headers=headers
    )
    if request_error:
        return {"ok": False, "error": request_error}
    try:
        result = res.json()
    except Exception:
        return {"ok": False, "error": f"Response không phải JSON: {res.text[:500]}"}
    if not result.get("success"):
        return {"ok": False, "error": str(result)[:500], "raw": result}

    data = result.get("data") or []
    if not data:
        return {"ok": False, "error": "Không tìm thấy hồ sơ quét theo tờ/thửa", "raw": result}

    def _get_tid(it):
        tinh_hinh = (it.get("ThongTinDangKy") or {}).get("TinhHinhDangKy") or {}
        try:
            return int(tinh_hinh.get("tinhHinhDangKyId"))
        except Exception:
            return None

    if len(data) > 1:
        data_co_id = [it for it in data if _get_tid(it) is not None]
        if not data_co_id:
            return {"ok": False,
                    "error": f"Tìm thấy {len(data)} bản ghi nhưng không bản ghi nào có tinhHinhDangKyId",
                    "raw": result}
        data = sorted(data_co_id, key=_get_tid)
        result = dict(result)
        result["data"] = [data[0]]

    item = data[0]
    list_hosoquet = item.get("ListHoSoQuet") or []
    if not list_hosoquet:
        return {"ok": False, "error": "Có bản ghi nhưng không có ListHoSoQuet", "raw": result}

    return {"ok": True, "item": item, "list_hosoquet": list_hosoquet, "raw": result}


# =========================
# PARSE
# =========================

def lay_chu_su_dung(item):
    try:
        chu = (item.get("ThongTinDangKy") or {}).get("ChuSoHuu") or {}
        ca = chu.get("CaNhans") or []
        if ca and ca[0].get("hoTen"):
            return ca[0]["hoTen"]
        tc = chu.get("ToChucs") or []
        if tc:
            return tc[0].get("tenToChuc") or tc[0].get("ten") or ""
        vc = chu.get("VoChongs") or []
        if vc:
            chong = (vc[0].get("Chong") or {}).get("hoTen") or ""
            vo = (vc[0].get("Vo") or {}).get("hoTen") or ""
            return (chong + (" & " + vo if vo else "")).strip(" &")
        hgd = chu.get("HoGiaDinhs") or []
        if hgd:
            return (hgd[0].get("ChuHo") or {}).get("hoTen") or ""
    except Exception:
        pass
    return ""


def get_tinh_hinh_id_from_item_hoso(item, hoso):
    thong_tin = item.get("ThongTinDangKy") or {}
    tinh_hinh = thong_tin.get("TinhHinhDangKy") or {}
    return tinh_hinh.get("tinhHinhDangKyId") or hoso.get("tinhHinhDangKyId") or hoso.get("Title")


def tim_hosoquet_tu_search(raw_search):
    data = raw_search.get("data") or []
    for item in data:
        tinh_hinh = (item.get("ThongTinDangKy") or {}).get("TinhHinhDangKy") or {}
        list_hosoquet = item.get("ListHoSoQuet") or []
        if list_hosoquet:
            hoso = list_hosoquet[0]
            return {"info": {"thongTinHoSoId": hoso.get("thongTinHoSoId"),
                             "tinhHinhDangKyId": get_tinh_hinh_id_from_item_hoso(item, hoso),
                             "xaId": hoso.get("xaId") or tinh_hinh.get("xaId")},
                    "hoso": hoso, "item": item}
    return None


# =========================
# XỬ LÝ 1 DÒNG
# =========================

def xu_ly_1_dong(session, item, maxa, logger, flags, debug_dir, progress_cb=None):
    row_excel = item["row"]
    soto, sothua, loaidat = item["soto"], item["sothua"], item["loaidat"]

    result_row = {
        "row_excel": row_excel, "soto": soto, "sothua": sothua, "loaidat": loaidat,
        "tinhHinhDangKyId": "", "hoSoQuetId": "", "thongTinHoSoId": "",
        "chu_su_dung": "", "status": "Lỗi", "note": "",
    }
    extra_notes = []

    def progress(p, t):
        if progress_cb:
            progress_cb(p, t)

    def done(icon, status, note):
        result_row["status"] = status
        all_notes = ([note] if note else []) + extra_notes
        result_row["note"] = " | ".join(n for n in all_notes if n)
        logger.log(f"{icon} Dòng {row_excel}: {result_row['note']}")
        return result_row

    progress(0, "Bắt đầu")

    progress(25, "Đang search hồ sơ quét")
    res_search = api_search_hosoquet(session, xa_id=maxa, so_to=soto, so_thua=sothua)
    if not res_search.get("ok"):
        progress(100, "Search lỗi")
        return done("❌", "Lỗi", "Search lỗi: " + res_search.get("error", ""))

    result_row["chu_su_dung"] = lay_chu_su_dung(res_search["item"])

    found = tim_hosoquet_tu_search(res_search["raw"])
    if not found:
        progress(100, "Không tìm thấy HSQ")
        return done("⚠️", "Lỗi", "Không tìm thấy ListHoSoQuet phù hợp")

    info, hoso = found["info"], found["hoso"]
    result_row["hoSoQuetId"]     = hoso.get("hoSoQuetId") or hoso.get("Title")
    result_row["thongTinHoSoId"] = hoso.get("thongTinHoSoId") or info.get("thongTinHoSoId")
    result_row["tinhHinhDangKyId"] = info.get("tinhHinhDangKyId") or hoso.get("tinhHinhDangKyId")

    logger.log(f"🔎 tinhHinhDangKyId={result_row['tinhHinhDangKyId']} "
               f"hoSoQuetId={result_row['hoSoQuetId']} tờ={soto} thửa={sothua}")

    if any(flags.get(k) for k in ("RULE_MA_DINH_DANH", "RULE_NGUON_GOC_THUA", "RULE_GIAY_CHUNG_NHAN")):
        progress(70, "Cập nhật thông tin đăng ký")
        res_ttdk = api_update_thong_tin_dang_ky(
            session, result_row["tinhHinhDangKyId"], flags=flags, debug_dir=debug_dir, row_excel=row_excel
        )
        if res_ttdk.get("notes"):
            extra_notes.extend(res_ttdk["notes"])
        if not res_ttdk.get("ok"):
            progress(100, "Update TTĐK lỗi")
            return done("❌", "Lỗi", "Update TTĐK lỗi: " + res_ttdk.get("error", ""))
        if res_ttdk.get("dry_run"):
            extra_notes.append("DRY_RUN — chưa gửi TTĐK")

    if flags.get("RULE_HSQ_NGUON_GOC", True):
        files_node = (hoso.get("ListFileHoSoQuet") or {}).get("ListFileHoSoQuet") or []
        hsq_notes = []
        if cap_nhat_hosoquet_nguon_goc(files_node, hsq_notes):
            progress(90, "Cập nhật HSQ nguồn gốc")
            res_hsq = api_update_hosoquet(session, hoso, debug_dir, row_excel, hsq_notes)
            extra_notes.extend(res_hsq.get("notes") or [])
            if not res_hsq.get("ok"):
                progress(100, "Update HSQ lỗi")
                return done("❌", "Lỗi", "Update HSQ lỗi: " + res_hsq.get("error", ""))
            if res_hsq.get("dry_run"):
                extra_notes.append("DRY_RUN — chưa gửi HSQ")
        else:
            extra_notes.extend(hsq_notes)

    progress(100, "Hoàn thành")
    if DRY_RUN:
        return done("🧪", "DRY_RUN", "Đã build payload (xem debug JSON)")
    return done("✅", "Thành công", "Cập nhật thành công")


# =========================
# WORKER
# =========================

def worker_run(username, password, maxa, excel_path, log_queue, flags):
    logger = ThreadSafeLogger(log_queue)
    driver = None
    try:
        data = doc_excel(excel_path)
        if not data:
            logger.log("❌ Excel không có dữ liệu.")
            return

        tong = len(data)
        logger.log(f"✅ Đọc Excel: {tong} dòng. DRY_RUN={DRY_RUN} | MAX_WORKERS={MAX_WORKERS}")
        logger.log("Quy tắc bật: " + ", ".join(k for k, v in flags.items() if v))

        output_path = os.path.join(os.path.dirname(excel_path), "ket_qua_cap_nhat_dangky.xlsx")
        debug_dir = os.path.join(os.path.dirname(excel_path), "debug_payload")

        driver = login_mplis(username, password)
        base_session = tao_session_tu_selenium(driver)
        thread_local = threading.local()

        def get_thread_session():
            if not hasattr(thread_local, "session"):
                ws = requests.Session()
                ws.headers.update(base_session.headers)
                ws.cookies.update(base_session.cookies)
                thread_local.session = ws
            return thread_local.session

        logger.log("✅ Đã tạo session API.")
        log_queue.put({"type": "init", "total": tong, "workers": MAX_WORKERS})

        results, results_lock = [], threading.Lock()
        counter_lock = threading.Lock()
        thanh_cong = bo_qua = that_bai = done_count = 0

        worker_id_map, worker_id_lock, worker_counter = {}, threading.Lock(), [0]

        def get_worker_idx():
            tid = threading.get_ident()
            with worker_id_lock:
                if tid not in worker_id_map:
                    worker_counter[0] += 1
                    worker_id_map[tid] = worker_counter[0]
                return worker_id_map[tid]

        def process_item(item):
            nonlocal thanh_cong, bo_qua, that_bai, done_count
            w_idx = get_worker_idx()
            row_excel = item["row"]

            def row_progress(percent, text):
                log_queue.put({"type": "worker_row_progress", "worker": w_idx,
                               "row": row_excel, "percent": percent, "text": text})

            try:
                kq = xu_ly_1_dong(get_thread_session(), item, maxa, logger, flags, debug_dir, row_progress)
            except Exception as e:
                traceback.print_exc()
                kq = {"row_excel": item.get("row"), "soto": item.get("soto"),
                      "sothua": item.get("sothua"), "loaidat": item.get("loaidat"),
                      "tinhHinhDangKyId": "", "hoSoQuetId": "", "thongTinHoSoId": "",
                      "chu_su_dung": "", "status": "Lỗi ngoài", "note": str(e)}

            with results_lock:
                results.append(kq)
            with counter_lock:
                done_count += 1
                st = kq.get("status", "")
                if st in ("Thành công", "DRY_RUN"):
                    thanh_cong += 1
                elif st == "Bỏ qua":
                    bo_qua += 1
                else:
                    that_bai += 1
                log_queue.put({"type": "progress", "done": done_count, "total": tong,
                               "thanh_cong": thanh_cong, "bo_qua": bo_qua, "that_bai": that_bai})
                if done_count % WRITE_EVERY_N == 0 or done_count == tong:
                    try:
                        ghi_excel_output(results, output_path, results_lock)
                        logger.log(f"💾 Ghi tạm Excel ({done_count} dòng)")
                    except Exception as ex:
                        logger.log(f"⚠️ Lỗi ghi Excel: {ex}")
            return kq

        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {executor.submit(process_item, it): it for it in data}
            for fut in as_completed(futures):
                try:
                    fut.result()
                except Exception as e:
                    logger.log(f"❌ Future exception: {e}")

        ghi_excel_output(results, output_path, results_lock)
        logger.log("=" * 70)
        logger.log(f"🎯 XONG. OK/DryRun: {thanh_cong} | Bỏ qua: {bo_qua} | Lỗi: {that_bai} | Tổng: {tong}")
        logger.log("📄 File kết quả: " + output_path)
        logger.log("📁 Debug payload: " + debug_dir)

    except Exception as e:
        logger.log("❌ Lỗi chương trình: " + str(e))
        traceback.print_exc()
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass
        log_queue.put("__DONE__")


# =========================
# TKINTER APP
# =========================

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Cập nhật Thông tin đăng ký (VBDLIS) — CHẠY THẬT")
        self.geometry("1120x820")

        self.log_queue = queue.Queue()
        self.worker_thread = None

        self.var_username = tk.StringVar()
        self.var_password = tk.StringVar()
        self.var_maxa     = tk.StringVar()
        self.var_excel    = tk.StringVar()
        self.var_workers  = tk.IntVar(value=MAX_WORKERS)

        self._worker_rows = []
        self.create_widgets()
        self.after(200, self.process_log_queue)

    def create_widgets(self):
        frame_top = ttk.LabelFrame(self, text="Thông tin chạy")
        frame_top.pack(fill="x", padx=10, pady=(10, 4))

        ttk.Label(frame_top, text="Username").grid(row=0, column=0, padx=5, pady=4, sticky="w")
        ttk.Entry(frame_top, textvariable=self.var_username, width=28).grid(row=0, column=1, padx=5, pady=4)
        ttk.Label(frame_top, text="Password").grid(row=0, column=2, padx=5, pady=4, sticky="w")
        ttk.Entry(frame_top, textvariable=self.var_password, width=28, show="*").grid(row=0, column=3, padx=5, pady=4)

        ttk.Label(frame_top, text="Mã xã").grid(row=1, column=0, padx=5, pady=4, sticky="w")
        ttk.Entry(frame_top, textvariable=self.var_maxa, width=18).grid(row=1, column=1, padx=5, pady=4, sticky="w")

        ttk.Label(frame_top, text="Số luồng").grid(row=2, column=0, padx=5, pady=4, sticky="w")
        ttk.Spinbox(frame_top, textvariable=self.var_workers, from_=1, to=10, width=6).grid(row=2, column=1, padx=5, pady=4, sticky="w")
        ttk.Label(frame_top, text="CHẠY THẬT: cập nhật TTĐK + HSQ, không DRY RUN",
                  foreground="red").grid(row=2, column=2, columnspan=3, padx=10, pady=4, sticky="w")

        ttk.Label(frame_top, text="File Excel").grid(row=3, column=0, padx=5, pady=4, sticky="w")
        ttk.Entry(frame_top, textvariable=self.var_excel, width=80).grid(row=3, column=1, columnspan=3, padx=5, pady=4, sticky="we")
        ttk.Button(frame_top, text="Duyệt", command=self.browse_excel).grid(row=3, column=4, padx=5, pady=4)
        frame_top.columnconfigure(3, weight=1)

        frame_btn = ttk.Frame(self)
        frame_btn.pack(fill="x", padx=10, pady=2)
        self.btn_start = ttk.Button(frame_btn, text="▶  BẮT ĐẦU", command=self.start_run)
        self.btn_start.pack(side="left", padx=5, pady=6)
        ttk.Button(frame_btn, text="Xóa log", command=self.clear_log).pack(side="left", padx=5)
        ttk.Label(frame_btn, text="Excel cần cột: soto | sothua (loaidat tùy chọn)",
                  foreground="blue").pack(side="left", padx=15)

        self.frame_progress = ttk.LabelFrame(self, text="Tiến độ")
        self.frame_progress.pack(fill="x", padx=10, pady=4)
        ttk.Label(self.frame_progress, text="Tổng:", width=12, anchor="w").grid(row=0, column=0, padx=6, pady=4, sticky="w")
        self._pb_total = ttk.Progressbar(self.frame_progress, length=600, mode="determinate")
        self._pb_total.grid(row=0, column=1, padx=6, pady=4, sticky="we")
        self._lbl_total = ttk.Label(self.frame_progress, text="0 / 0  |  ✅0  ⏭0  ❌0", width=34)
        self._lbl_total.grid(row=0, column=2, padx=6, pady=4, sticky="w")
        self.frame_progress.columnconfigure(1, weight=1)

        frame_log = ttk.LabelFrame(self, text="Log xử lý")
        frame_log.pack(fill="both", expand=True, padx=10, pady=(4, 10))
        self.txt_log = tk.Text(frame_log, wrap="word", font=("Consolas", 9))
        self.txt_log.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(frame_log, command=self.txt_log.yview)
        sb.pack(side="right", fill="y")
        self.txt_log.configure(yscrollcommand=sb.set)

    def _init_progress(self, total, n_workers):
        self._pb_total.config(maximum=total, value=0)
        self._lbl_total.config(text=f"0 / {total}  |  ✅0  ⏭0  ❌0")
        for (lbl, pb, lbl_s) in self._worker_rows:
            lbl.grid_forget(); pb.grid_forget(); lbl_s.grid_forget()
        self._worker_rows.clear()
        for i in range(n_workers):
            r = i + 1
            lbl = ttk.Label(self.frame_progress, text=f"Worker {i+1}:", width=12, anchor="w")
            lbl.grid(row=r, column=0, padx=6, pady=2, sticky="w")
            pb = ttk.Progressbar(self.frame_progress, length=600, mode="determinate", maximum=100)
            pb.grid(row=r, column=1, padx=6, pady=2, sticky="we")
            lbl_s = ttk.Label(self.frame_progress, text="Chờ...", width=40, foreground="gray")
            lbl_s.grid(row=r, column=2, padx=6, pady=2, sticky="w")
            self._worker_rows.append((lbl, pb, lbl_s))

    def browse_excel(self):
        p = filedialog.askopenfilename(title="Chọn file Excel",
                                       filetypes=[("Excel", "*.xlsx *.xlsm"), ("All", "*.*")])
        if p:
            self.var_excel.set(p)

    def clear_log(self):
        self.txt_log.delete("1.0", tk.END)

    def validate_input(self):
        if not self.var_username.get().strip():
            messagebox.showerror("Thiếu", "Chưa nhập username."); return False
        if not self.var_password.get().strip():
            messagebox.showerror("Thiếu", "Chưa nhập password."); return False
        if not self.var_maxa.get().strip():
            messagebox.showerror("Thiếu", "Chưa nhập mã xã."); return False
        if not os.path.isfile(self.var_excel.get().strip()):
            messagebox.showerror("Sai", "File Excel không tồn tại."); return False
        return True

    def start_run(self):
        if self.worker_thread and self.worker_thread.is_alive():
            messagebox.showwarning("Đang chạy", "Chương trình đang chạy."); return
        if not self.validate_input():
            return

        global MAX_WORKERS, DRY_RUN
        MAX_WORKERS = self.var_workers.get()
        DRY_RUN = False

        flags = {
            "RULE_MA_DINH_DANH":    True,
            "RULE_NGUON_GOC_THUA":  True,
            "RULE_GIAY_CHUNG_NHAN": True,
            "RULE_HSQ_NGUON_GOC":   True,
        }

        self.btn_start.config(state="disabled")
        self.worker_thread = threading.Thread(
            target=worker_run,
            args=(self.var_username.get().strip(), self.var_password.get().strip(),
                  self.var_maxa.get().strip(), self.var_excel.get().strip(),
                  self.log_queue, flags),
            daemon=True
        )
        self.worker_thread.start()

    def process_log_queue(self):
        try:
            while True:
                msg = self.log_queue.get_nowait()
                if isinstance(msg, dict):
                    t = msg.get("type")
                    if t == "init":
                        self._init_progress(msg["total"], msg["workers"])
                    elif t == "worker_row_progress":
                        w = msg["worker"] - 1
                        pct = max(0, min(100, int(msg.get("percent", 0))))
                        if w < len(self._worker_rows):
                            _, pb, lbl_s = self._worker_rows[w]
                            pb.config(value=pct)
                            color = "green" if pct >= 100 else "blue"
                            lbl_s.config(text=f"Dòng {msg['row']} — {pct}% — {msg.get('text','')}",
                                         foreground=color)
                    elif t == "progress":
                        done, total = msg["done"], msg["total"]
                        self._pb_total.config(value=done)
                        pct = int(done / total * 100) if total else 0
                        self._lbl_total.config(
                            text=f"{done}/{total} ({pct}%)  |  ✅{msg['thanh_cong']}  ⏭{msg['bo_qua']}  ❌{msg['that_bai']}")
                    continue

                if msg == "__DONE__":
                    self.btn_start.config(state="normal")
                    for (_, pb, lbl_s) in self._worker_rows:
                        lbl_s.config(text="Xong", foreground="gray")
                    self.txt_log.insert(tk.END, "\n✅ Tiến trình đã kết thúc.\n")
                    self.txt_log.see(tk.END)
                    continue

                self.txt_log.insert(tk.END, msg + "\n")
                self.txt_log.see(tk.END)
        except queue.Empty:
            pass
        self.after(200, self.process_log_queue)


if __name__ == "__main__":
    App().mainloop()