import os
import sys
import queue
import threading
import traceback
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from helpers import wait_query_done
from ultils import *


URL = "https://dla.mplis.gov.vn/dc/DonDangKy/KeKhaiDangKyV2"


class TextRedirector:
    def __init__(self, log_queue):
        self.log_queue = log_queue

    def write(self, text):
        if text.strip():
            self.log_queue.put(text)

    def flush(self):
        pass


def doc_excel(path_excel):
    """
    Đọc Excel có cột:
    soto, sothua, loaidat, tenfile

    Return:
        list dict
    """

    wb = load_workbook(path_excel, data_only=True)
    ws = wb.active

    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value:
            headers[str(value).strip().lower()] = col

    required_cols = ["soto", "sothua", "loaidat", "tenfile"]
    missing = [c for c in required_cols if c not in headers]

    if missing:
        raise RuntimeError(f"Thiếu cột trong Excel: {', '.join(missing)}")

    data = []

    for row in range(2, ws.max_row + 1):
        soto = ws.cell(row=row, column=headers["soto"]).value
        sothua = ws.cell(row=row, column=headers["sothua"]).value
        loaidat = ws.cell(row=row, column=headers["loaidat"]).value
        tenfile = ws.cell(row=row, column=headers["tenfile"]).value

        if not soto and not sothua and not loaidat and not tenfile:
            continue

        data.append({
            "row": row,
            "soto": str(soto).strip() if soto is not None else "",
            "sothua": str(sothua).strip() if sothua is not None else "",
            "loaidat": str(loaidat).strip() if loaidat is not None else "",
            "tenfile": str(tenfile).strip() if tenfile is not None else "",
        })

    return data


def lay_duong_dan_file(folder_upload, tenfile):
    """
    Nếu tenfile là đường dẫn tuyệt đối thì dùng luôn.
    Nếu không thì ghép với folder_upload.
    """

    if os.path.isabs(tenfile):
        return tenfile

    return os.path.join(folder_upload, tenfile)


def kiem_tra_hosoquet_co_chuacogiay(driver, modal_ho_so_quet, timeout=10):
    """
    Kiểm tra trong:
    #vModuleHoSoQuetChon #lstHoSoQuet

    Nếu có span.value chứa CHUACOGIAY thì True.
    Nếu không có thì False.
    """

    try:
        wait = WebDriverWait(driver, timeout)

        wait.until(
            lambda d: modal_ho_so_quet.find_element(
                By.CSS_SELECTOR,
                "#vModuleHoSoQuetChon #lstHoSoQuet"
            )
        )

        spans = modal_ho_so_quet.find_elements(
            By.CSS_SELECTOR,
            "#vModuleHoSoQuetChon #lstHoSoQuet span.value"
        )

        texts = []

        for sp in spans:
            text = sp.text.strip()
            if text:
                texts.append(text)

        print(f"🔎 Nội dung hồ sơ quét tìm thấy: {texts}")

        for text in texts:
            if "CHUACOGIAY" in text.upper():
                print("✅ Hồ sơ quét có chứa CHUACOGIAY.")
                return True

        print("⚠️ Hồ sơ quét không chứa CHUACOGIAY.")
        return False

    except Exception as e:
        print(f"❌ Lỗi khi kiểm tra CHUACOGIAY trong hồ sơ quét: {e}")
        return False


def dong_modal_ho_so_quet(driver, timeout=10):
    """
    Đóng modal Hồ sơ quét bằng #closeModal trong div[id^='mdlHoSoQuet-']
    """

    try:
        wait = WebDriverWait(driver, timeout)

        modal = wait.until(
            EC.visibility_of_element_located((
                By.CSS_SELECTOR,
                "div[id^='mdlHoSoQuet-'].in, "
                "div[id^='mdlHoSoQuet-'].show, "
                "div[id^='mdlHoSoQuet-'].modal.in, "
                "div[id^='mdlHoSoQuet-'].modal.show"
            ))
        )

        btn_close = modal.find_element(By.CSS_SELECTOR, "#closeModal")

        try:
            btn_close.click()
        except Exception:
            driver.execute_script("arguments[0].click();", btn_close)

        wait.until(
            EC.invisibility_of_element_located((
                By.CSS_SELECTOR,
                "div[id^='mdlHoSoQuet-']"
            ))
        )

        print("✅ Đã đóng modal Hồ sơ quét.")
        return True

    except Exception as e:
        print(f"❌ Lỗi khi đóng modal Hồ sơ quét: {e}")
        return False


def xu_ly_1_dong(driver, wait, item, maxa, folder_upload):
    """
    Xử lý 1 dòng Excel.
    """

    row_excel = item["row"]
    soto = item["soto"]
    sothua = item["sothua"]
    loaidat = item["loaidat"]
    tenfile = item["tenfile"]

    file_path = lay_duong_dan_file(folder_upload, tenfile)

    print("=" * 80)
    print(f"▶️ Dòng Excel {row_excel}: Số tờ={soto}, Số thửa={sothua}, Loại đất={loaidat}, File={tenfile}")

    if not os.path.isfile(file_path):
        print(f"❌ Không tìm thấy file upload: {file_path}")
        return False
    
    mo_tra_cuu_don_dang_ky(
        driver=driver,
        wait=wait
        )

    # 1. Tìm số tờ số thửa
    so_ban_ghi = nhap_to_thua_va_tim_kiem(
        driver=driver,
        wait=wait,
        so_to=soto,
        so_thua=sothua,
        timeout=60
    )

    if not so_ban_ghi:
        print("⚠️ Không tìm thấy thửa đất. Bỏ qua dòng này.")
        return False

    # 2. Chọn bản ghi đầu tiên
    chon_ban_ghi_dau_tien(driver)
    print("✅ Đã chọn bản ghi đầu tiên.")
    wait_query_done(driver)

    # 3. Mở modal Hồ sơ quét
    modal_ho_so_quet = mo_ho_so_quet(driver, timeout=60)

    if not modal_ho_so_quet:
        print("❌ Không mở được modal Hồ sơ quét.")
        return False

    # 4. Kiểm tra hồ sơ quét có CHUACOGIAY không
    co_chuacogiay = kiem_tra_hosoquet_co_chuacogiay(
        driver=driver,
        modal_ho_so_quet=modal_ho_so_quet,
        timeout=10
    )

    if not co_chuacogiay:
        print("⚠️ Hồ sơ quét không có CHUACOGIAY. Đóng modal, bỏ đơn và qua dòng tiếp theo.")

        dong_modal_ho_so_quet(driver, timeout=10)

        bo_don_dang_ky(
            driver=driver,
            timeout=30
        )

        return False

    # 5. Chọn hồ sơ quét đầu tiên
    ok_chon_hsq = chon_ho_so_quet_dau_tien(
        driver=driver,
        modal_ho_so_quet=modal_ho_so_quet,
        timeout=10
    )

    if not ok_chon_hsq:
        print("❌ Không chọn được hồ sơ quét đầu tiên.")
        dong_modal_ho_so_quet(driver, timeout=10)
        return False

    # 6. Mở modal cập nhật hồ sơ quét
    modal_add_hsq = cap_nhat_ho_so_quet_dau_tien(
        driver=driver,
        modal_ho_so_quet=modal_ho_so_quet,
        timeout=30
    )

    if not modal_add_hsq:
        print("❌ Không mở được modal cập nhật hồ sơ quét.")
        dong_modal_ho_so_quet(driver, timeout=10)
        return False

    # 7. Xóa file đầu tiên
    ok_xoa = xoa_file_dau_tien_trong_add_hosoquet(
        driver=driver,
        modal_add_hsq=modal_add_hsq,
        timeout=30
    )

    if not ok_xoa:
        print("⚠️ Không xóa được file đầu tiên. Vẫn tiếp tục thêm file mới.")

    # 8. Thêm dòng Đơn đăng ký
    ok_them = them_file_don_dang_ky_trong_add_hosoquet(
        driver=driver,
        modal_add_hsq=modal_add_hsq,
        maxa=maxa,
        loaidat=loaidat,
        timeout=30
    )

    if not ok_them:
        print("❌ Không thêm được dòng Đơn đăng ký.")
        return False

    # 9. Upload file theo mô tả
    ok_upload = upload_file_theo_mo_ta_trong_add_hosoquet(
        driver=driver,
        modal_add_hsq=modal_add_hsq,
        maxa=maxa,
        loaidat=loaidat,
        file_path=file_path,
        timeout=60
    )

    # 10. Cập nhật và đóng modal hồ sơ quét
    ok_cap_nhat = cap_nhat_va_dong_modal_hosoquet(
        driver=driver,
        modal_add_hsq=modal_add_hsq,
        timeout=60
    )

    if not ok_cap_nhat:
        print("❌ Không cập nhật hoặc đóng được modal Hồ sơ quét.")
        return False

    # 11. Bỏ đơn
    ok_bo_don = bo_don_dang_ky(
        driver=driver,
        timeout=30
    )

    print(f"✅ Hoàn thành dòng Excel {row_excel}.")
    return True


def worker_run(username, password, maxa, excel_path, folder_upload, log_queue):
    old_stdout = sys.stdout
    sys.stdout = TextRedirector(log_queue)

    driver = None

    try:
        data = doc_excel(excel_path)

        if not data:
            print("❌ Excel không có dữ liệu.")
            return

        print(f"✅ Đã đọc Excel: {len(data)} dòng dữ liệu.")

        options = Options()
        options.add_argument("--start-maximized")
        options.add_argument("--window-position=100,100")
        options.add_argument("--window-size=1400,900")

        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)
        wait = WebDriverWait(driver, 20)

        driver.get(URL)
        print(f"🌐 Mở trang: {URL}")

        username_box, password_box = get_login_fields(wait)
        username_box.send_keys(username)
        password_box.send_keys(password)
        password_box.send_keys(Keys.ENTER)

        messagebox.showinfo(
            "Xác minh Authenticator",
            "Vui lòng hoàn tất xác minh bằng Authenticator trên trình duyệt, sau đó bấm OK để tiếp tục.",
        )  

        wait_query_done(driver)

        chon_xa(driver, wait, maxa)

        tong = len(data)
        thanh_cong = 0
        that_bai = 0

        for idx, item in enumerate(data, start=1):
            print(f"\n🚀 Đang xử lý {idx}/{tong}")

            try:
                ok = xu_ly_1_dong(
                    driver=driver,
                    wait=wait,
                    item=item,
                    maxa=maxa,
                    folder_upload=folder_upload
                )

                if ok:
                    thanh_cong += 1
                else:
                    that_bai += 1

            except Exception as e:
                that_bai += 1
                print(f"❌ Lỗi ngoài khi xử lý dòng Excel {item.get('row')}: {e}")
                traceback.print_exc()

        print("=" * 80)
        print(f"🎯 XONG. Thành công: {thanh_cong} | Bỏ qua/Lỗi: {that_bai} | Tổng: {tong}")

    except Exception as e:
        print(f"❌ Lỗi chương trình: {e}")
        traceback.print_exc()

    finally:
        sys.stdout = old_stdout

        if driver:
            try:
                driver.quit()
            except Exception:
                pass

        log_queue.put("__DONE__")


class App(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Gắn hồ sơ quét Đơn đăng ký")
        self.geometry("950x650")

        self.log_queue = queue.Queue()
        self.worker_thread = None

        self.var_username = tk.StringVar()
        self.var_password = tk.StringVar()
        self.var_maxa = tk.StringVar(value="24121")
        self.var_excel = tk.StringVar()
        self.var_folder = tk.StringVar()

        self.create_widgets()
        self.after(200, self.process_log_queue)

    def create_widgets(self):
        frame_top = ttk.LabelFrame(self, text="Thông tin chạy")
        frame_top.pack(fill="x", padx=10, pady=10)

        ttk.Label(frame_top, text="Username").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(frame_top, textvariable=self.var_username, width=30).grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(frame_top, text="Password").grid(row=0, column=2, padx=5, pady=5, sticky="w")
        ttk.Entry(frame_top, textvariable=self.var_password, width=30, show="*").grid(row=0, column=3, padx=5, pady=5)

        ttk.Label(frame_top, text="Mã xã").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(frame_top, textvariable=self.var_maxa, width=30).grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(frame_top, text="File Excel").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(frame_top, textvariable=self.var_excel, width=80).grid(row=2, column=1, columnspan=3, padx=5, pady=5, sticky="we")
        ttk.Button(frame_top, text="Duyệt Excel", command=self.browse_excel).grid(row=2, column=4, padx=5, pady=5)

        ttk.Label(frame_top, text="Folder file PDF").grid(row=3, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(frame_top, textvariable=self.var_folder, width=80).grid(row=3, column=1, columnspan=3, padx=5, pady=5, sticky="we")
        ttk.Button(frame_top, text="Duyệt Folder", command=self.browse_folder).grid(row=3, column=4, padx=5, pady=5)

        self.btn_start = ttk.Button(frame_top, text="BẮT ĐẦU CHẠY", command=self.start_run)
        self.btn_start.grid(row=4, column=1, padx=5, pady=10, sticky="w")

        self.btn_clear = ttk.Button(frame_top, text="Xóa log", command=self.clear_log)
        self.btn_clear.grid(row=4, column=2, padx=5, pady=10, sticky="w")

        frame_top.columnconfigure(3, weight=1)

        frame_log = ttk.LabelFrame(self, text="Log xử lý")
        frame_log.pack(fill="both", expand=True, padx=10, pady=10)

        self.txt_log = tk.Text(frame_log, wrap="word")
        self.txt_log.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(frame_log, command=self.txt_log.yview)
        scrollbar.pack(side="right", fill="y")

        self.txt_log.configure(yscrollcommand=scrollbar.set)

    def browse_excel(self):
        path = filedialog.askopenfilename(
            title="Chọn file Excel",
            filetypes=[
                ("Excel files", "*.xlsx *.xlsm"),
                ("All files", "*.*")
            ]
        )

        if path:
            self.var_excel.set(path)

    def browse_folder(self):
        folder = filedialog.askdirectory(title="Chọn folder chứa file upload")

        if folder:
            self.var_folder.set(folder)

    def clear_log(self):
        self.txt_log.delete("1.0", tk.END)

    def validate_input(self):
        username = self.var_username.get().strip()
        password = self.var_password.get().strip()
        maxa = self.var_maxa.get().strip()
        excel = self.var_excel.get().strip()
        folder = self.var_folder.get().strip()

        if not username:
            messagebox.showerror("Thiếu thông tin", "Chưa nhập username.")
            return False

        if not password:
            messagebox.showerror("Thiếu thông tin", "Chưa nhập password.")
            return False

        if not maxa:
            messagebox.showerror("Thiếu thông tin", "Chưa nhập mã xã.")
            return False

        if not os.path.isfile(excel):
            messagebox.showerror("Sai đường dẫn", "File Excel không tồn tại.")
            return False

        if not os.path.isdir(folder):
            messagebox.showerror("Sai đường dẫn", "Folder file upload không tồn tại.")
            return False

        return True

    def start_run(self):
        if self.worker_thread and self.worker_thread.is_alive():
            messagebox.showwarning("Đang chạy", "Chương trình đang chạy.")
            return

        if not self.validate_input():
            return

        self.btn_start.config(state="disabled")

        username = self.var_username.get().strip()
        password = self.var_password.get().strip()
        maxa = self.var_maxa.get().strip()
        excel = self.var_excel.get().strip()
        folder = self.var_folder.get().strip()

        self.worker_thread = threading.Thread(
            target=worker_run,
            args=(username, password, maxa, excel, folder, self.log_queue),
            daemon=True
        )

        self.worker_thread.start()

    def process_log_queue(self):
        try:
            while True:
                msg = self.log_queue.get_nowait()

                if msg == "__DONE__":
                    self.btn_start.config(state="normal")
                    self.txt_log.insert(tk.END, "\n✅ Tiến trình đã kết thúc.\n")
                    self.txt_log.see(tk.END)
                    continue

                self.txt_log.insert(tk.END, msg + "\n")
                self.txt_log.see(tk.END)

        except queue.Empty:
            pass

        self.after(200, self.process_log_queue)


if __name__ == "__main__":
    app = App()
    app.mainloop()