import os
import sys
import json
import uuid
import queue
import threading
import traceback
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime, timezone

import requests
from openpyxl import load_workbook, Workbook

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys


# =========================
# API CONFIG
# =========================

REFERER_URL = "https://dla.mplis.gov.vn/dc/DonDangKy/KeKhaiDangKyV2"

API_SEARCH_THDK = "https://dla.mplis.gov.vn/dc/DangKyAjax/AdvancedSearchTinhHinhDangKy"
API_GET_HOSOQUET = "https://dla.mplis.gov.vn/dc/HoSoQuetAjax/GetHoSoQuetKeKhaiByTinhHinhDangKyId"
API_UPDATE_HOSOQUET = "https://dla.mplis.gov.vn/dc/HoSoQuetAjax/UpdateHoSoQuetExistFile"

# True = chỉ kiểm tra, không update thật
# False = update thật
DRY_RUN = False


# =========================
# LOG REDIRECT
# =========================

class TextRedirector:
    def __init__(self, log_queue):
        self.log_queue = log_queue

    def write(self, text):
        if text and text.strip():
            self.log_queue.put(text)

    def flush(self):
        pass


# =========================
# EXCEL
# =========================

def doc_excel(path_excel):
    wb = load_workbook(path_excel, data_only=True)
    ws = wb.active

    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value:
            headers[str(value).strip().lower()] = col

    required_cols = ["soto", "sothua", "loaidat", "tenfile"]
    missing = [c for c in required_cols if c not in headers]

    if missing:
        raise RuntimeError(f"Thiếu cột trong Excel: {', '.join(missing)}")

    data = []

    for row in range(2, ws.max_row + 1):
        soto = ws.cell(row=row, column=headers["soto"]).value
        sothua = ws.cell(row=row, column=headers["sothua"]).value
        loaidat = ws.cell(row=row, column=headers["loaidat"]).value
        tenfile = ws.cell(row=row, column=headers["tenfile"]).value

        if not soto and not sothua and not loaidat and not tenfile:
            continue

        data.append({
            "row": row,
            "soto": str(soto).strip() if soto is not None else "",
            "sothua": str(sothua).strip() if sothua is not None else "",
            "loaidat": str(loaidat).strip() if loaidat is not None else "",
            "tenfile": str(tenfile).strip() if tenfile is not None else "",
        })

    return data


def ghi_excel_output(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "KetQua"

    headers = [
        "STT",
        "Dòng Excel",
        "Số tờ",
        "Số thửa",
        "Loại đất",
        "Tên file",
        "Mô tả mới",
        "tinhHinhDangKyId",
        "hoSoQuetId",
        "thongTinHoSoId",
        "Trạng thái",
        "Ghi chú"
    ]

    ws.append(headers)

    for i, r in enumerate(rows, start=1):
        ws.append([
            i,
            r.get("row_excel"),
            r.get("soto"),
            r.get("sothua"),
            r.get("loaidat"),
            r.get("tenfile"),
            r.get("mo_ta"),
            r.get("tinhHinhDangKyId"),
            r.get("hoSoQuetId"),
            r.get("thongTinHoSoId"),
            r.get("status"),
            r.get("note"),
        ])

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max_len + 2, 70)

    wb.save(output_path)


def lay_duong_dan_file(folder_upload, tenfile):
    if os.path.isabs(tenfile):
        return tenfile
    return os.path.join(folder_upload, tenfile)


# =========================
# SELENIUM LOGIN + TOKEN
# =========================

def lay_token_tu_trang(driver):
    js = """
    return (
        document.querySelector('input[name="__RequestVerificationToken"]')?.value ||
        document.querySelector('input[name="__requestverificationtoken"]')?.value ||
        document.querySelector('meta[name="__RequestVerificationToken"]')?.content ||
        document.querySelector('meta[name="__requestverificationtoken"]')?.content ||
        document.querySelector('meta[name="RequestVerificationToken"]')?.content ||
        ''
    );
    """
    return driver.execute_script(js)


def tao_session_tu_selenium(driver):
    session = requests.Session()

    user_agent = driver.execute_script("return navigator.userAgent;")
    token = lay_token_tu_trang(driver)

    if not token:
        raise RuntimeError("Không lấy được __requestverificationtoken.")

    session.headers.update({
        "User-Agent": user_agent,
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "X-Requested-With": "XMLHttpRequest",
        "Origin": "https://dla.mplis.gov.vn",
        "Referer": REFERER_URL,
        "__requestverificationtoken": token,
    })

    for c in driver.get_cookies():
        session.cookies.set(
            name=c["name"],
            value=c["value"],
            domain=c.get("domain"),
            path=c.get("path", "/")
        )

    print("✅ Đã tạo session API.")
    print("🔐 Token:", token[:30] + "...")

    return session


def login_mplis(username, password):
    options = Options()
    options.add_argument("--start-maximized")
    options.add_argument("--window-position=100,100")
    options.add_argument("--window-size=1400,900")

    driver = webdriver.Chrome(options=options)
    driver.get(REFERER_URL)

    # Cố gắng tự nhập username/password nếu thấy form
    try:
        import time
        time.sleep(2)

        inputs = driver.find_elements(By.CSS_SELECTOR, "input")

        user_box = None
        pass_box = None

        for inp in inputs:
            typ = (inp.get_attribute("type") or "").lower()
            name = (inp.get_attribute("name") or "").lower()
            placeholder = (inp.get_attribute("placeholder") or "").lower()
            input_id = (inp.get_attribute("id") or "").lower()

            text_all = f"{name} {placeholder} {input_id}"

            if not user_box and typ in ["text", "email"] and any(k in text_all for k in ["user", "username", "login", "account", "ten"]):
                user_box = inp

            if not pass_box and typ == "password":
                pass_box = inp

        if not user_box:
            for inp in inputs:
                typ = (inp.get_attribute("type") or "").lower()
                if typ in ["text", "email"]:
                    user_box = inp
                    break

        if user_box and pass_box:
            user_box.clear()
            user_box.send_keys(username)
            pass_box.clear()
            pass_box.send_keys(password)
            pass_box.send_keys(Keys.ENTER)
            print("✅ Đã nhập username/password, chờ xác thực nếu có.")
        else:
            print("⚠️ Không tự tìm thấy ô đăng nhập. Mày đăng nhập tay trên Chrome.")

    except Exception as e:
        print("⚠️ Không tự login được, mày đăng nhập tay:", e)

    messagebox.showinfo(
        "Đăng nhập / Authenticator",
        "Hoàn tất đăng nhập MPLIS và xác thực Authenticator trên Chrome.\n"
        "Sau khi vào được màn hình kê khai đăng ký thì bấm OK để tiếp tục."
    )

    return driver


# =========================
# API 1: SEARCH TỜ THỬA
# =========================

def tao_payload_tim_to_thua(xa_id, so_to, so_thua):
    return {
        "draw": "2",

        "columns[0][data]": "",
        "columns[0][name]": "",
        "columns[0][searchable]": "true",
        "columns[0][orderable]": "false",
        "columns[0][search][value]": "",
        "columns[0][search][regex]": "false",

        "columns[1][data]": "tinhHinhDangKyId",
        "columns[1][name]": "tinhHinhDangKyId",
        "columns[1][searchable]": "true",
        "columns[1][orderable]": "true",
        "columns[1][search][value]": "",
        "columns[1][search][regex]": "false",

        "columns[2][data]": "maDon",
        "columns[2][name]": "maDon",
        "columns[2][searchable]": "true",
        "columns[2][orderable]": "true",
        "columns[2][search][value]": "",
        "columns[2][search][regex]": "false",

        "columns[3][data]": "soThuTu",
        "columns[3][name]": "soThuTu",
        "columns[3][searchable]": "true",
        "columns[3][orderable]": "true",
        "columns[3][search][value]": "",
        "columns[3][search][regex]": "false",

        "columns[4][data]": "DaiDienKhaiTrinh",
        "columns[4][name]": "DaiDienKhaiTrinh",
        "columns[4][searchable]": "true",
        "columns[4][orderable]": "false",
        "columns[4][search][value]": "",
        "columns[4][search][regex]": "false",

        "columns[5][data]": "ngayTiepNhan",
        "columns[5][name]": "ngayTiepNhan",
        "columns[5][searchable]": "true",
        "columns[5][orderable]": "true",
        "columns[5][search][value]": "",
        "columns[5][search][regex]": "false",

        "columns[6][data]": "thoiDiemDangKy",
        "columns[6][name]": "thoiDiemDangKy",
        "columns[6][searchable]": "true",
        "columns[6][orderable]": "true",
        "columns[6][search][value]": "",
        "columns[6][search][regex]": "false",

        "order[0][column]": "5",
        "order[0][dir]": "desc",
        "start": "0",
        "length": "10",
        "search[value]": "",
        "search[regex]": "false",

        "model[xaId]": str(xa_id),
        "model[huyenId]": "",
        "model[tinhHinhDangKyId]": "",
        "model[maDon]": "",
        "model[soThuTu]": "",
        "model[ngayTiepNhan]": "",
        "model[thoiDiemDangKy]": "",
        "model[loaiGiayChungNhanId]": "",
        "model[soPhatHanh]": "",
        "model[maVach]": "",
        "model[soVaoSo]": "",
        "model[soVaoSoCu]": "",
        "model[ngayVaoSo]": "",
        "model[soHoSoGoc]": "",
        "model[soHoSoGocCu]": "",
        "model[hoTen]": "",
        "model[soGiayTo]": "",
        "model[namSinh]": "",

        "model[soThuTuThua]": str(so_thua),
        "model[soHieuToBanDo]": str(so_to),

        "model[soThuTuThuaCu]": "",
        "model[soHieuToBanDoCu]": "",
        "model[soNha]": "",
        "model[diaChiChiTiet]": "",
        "model[dieuKienCapGiay]": "",
        "model[phucHoiDuLieu]": "false",
    }


def api_tim_to_thua(session, xa_id, so_to, so_thua):
    headers = dict(session.headers)
    headers["Content-Type"] = "application/x-www-form-urlencoded; charset=UTF-8"

    res = session.post(
        API_SEARCH_THDK,
        data=tao_payload_tim_to_thua(xa_id, so_to, so_thua),
        headers=headers,
        timeout=30
    )

    result = res.json()

    if result.get("success") is False:
        return {
            "ok": False,
            "error": result.get("Error", "API search lỗi"),
            "raw": result
        }

    records_total = result.get("recordsTotal", 0)
    data = result.get("data", [])

    if records_total == 0 or not data:
        return {
            "ok": False,
            "error": "Không tìm thấy bản ghi",
            "recordsTotal": records_total,
            "raw": result
        }

    if records_total > 1:
        return {
            "ok": False,
            "error": f"Tìm thấy {records_total} bản ghi, không xử lý để tránh nhầm",
            "recordsTotal": records_total,
            "raw": result
        }

    item = data[0]

    return {
        "ok": True,
        "tinhHinhDangKyId": item.get("tinhHinhDangKyId"),
        "item": item,
        "raw": result
    }


# =========================
# API 2: LOAD HỒ SƠ QUÉT
# =========================

def api_get_hosoquet(session, tinh_hinh_dang_ky_id):
    headers = dict(session.headers)
    headers["Content-Type"] = "application/json; charset=UTF-8"

    payload = {
        "tinhHinhDangKyId": int(tinh_hinh_dang_ky_id),
        "giaoDichBaoDamId": 0
    }

    res = session.post(
        API_GET_HOSOQUET,
        json=payload,
        headers=headers,
        timeout=60
    )

    result = res.json()

    if not result.get("success"):
        return {
            "ok": False,
            "error": str(result),
            "raw": result
        }

    return {
        "ok": True,
        "raw": result
    }


def tim_hosoquet_chuacogiay(raw_hosoquet):
    value = raw_hosoquet.get("Value") or raw_hosoquet.get("value") or {}
    list_info = value.get("ListHoSoQuetInfo") or []

    for info in list_info:
        for hoso in info.get("ListHoSoQuet") or []:
            wrapper = hoso.get("ListFileHoSoQuet") or {}
            files = wrapper.get("ListFileHoSoQuet") or []

            for f in files:
                mo_ta = (f.get("moTa") or "").upper()
                if "CHUACOGIAY" in mo_ta:
                    return {
                        "info": info,
                        "hoso": hoso,
                        "file": f
                    }

    return None


# =========================
# API 3: UPDATE FILE HỒ SƠ QUÉT
# =========================

def now_iso_z():
    return datetime.now(timezone.utc).isoformat(timespec="milliseconds").replace("+00:00", "Z")


def api_update_hosoquet_exist_file(session, file_path, found_hosoquet, mo_ta_moi):
    if not os.path.isfile(file_path):
        return {
            "ok": False,
            "error": f"Không tìm thấy file PDF: {file_path}"
        }

    info = found_hosoquet["info"]
    hoso = found_hosoquet["hoso"]

    ho_so_quet_id = hoso.get("hoSoQuetId") or hoso.get("Title")
    thong_tin_ho_so_id = hoso.get("thongTinHoSoId") or info.get("thongTinHoSoId")
    tinh_hinh_dang_ky_id = hoso.get("tinhHinhDangKyId") or info.get("tinhHinhDangKyId")
    xa_id = hoso.get("xaId") or info.get("xaId")

    ho_so_quet = {
        "hoSoQuetId": int(ho_so_quet_id),
        "thongTinHoSoId": int(thong_tin_ho_so_id),
        "tinhHinhDangKyId": int(tinh_hinh_dang_ky_id),
        "xaId": int(xa_id),
        "CreatedDate": now_iso_z(),
        "ModifiedDate": now_iso_z(),
        "Id": hoso.get("Id"),
        "Title": str(hoso.get("Title") or ho_so_quet_id),
        "Name": hoso.get("Name"),
        "Path": hoso.get("Path"),
        "ParentPath": hoso.get("ParentPath"),
        "_id": 1,
        "TuiHoSo": None,
        "tuiHoSoId": int(hoso.get("tuiHoSoId") or 0)
    }

    info_ho_so_quet = {
        "loaiHoSoQuet": 2,
        "laGiayToVeNguonGoc": False,
        "giayChungNhanId": "",
        "moTa": mo_ta_moi,
        "tenGiayTo": "",
        "trichYeu": "",
        "laGiayChungNhan": False,
        "__id": str(uuid.uuid4()),
        "files": None
    }

    data = {
        "hoSoQuet": json.dumps(ho_so_quet, ensure_ascii=False),
        "infoHoSoQuet_1": json.dumps(info_ho_so_quet, ensure_ascii=False),
        "count": "1"
    }

    headers = dict(session.headers)

    # Multipart upload: không set Content-Type thủ công
    headers.pop("Content-Type", None)

    with open(file_path, "rb") as f:
        files = {
            "fileHoSoQuet_1": (
                os.path.basename(file_path),
                f,
                "application/pdf"
            )
        }

        res = session.post(
            API_UPDATE_HOSOQUET,
            data=data,
            files=files,
            headers=headers,
            timeout=180
        )

    try:
        result = res.json()
    except Exception:
        return {
            "ok": False,
            "error": f"Response update không phải JSON: {res.text[:1000]}"
        }

    if not result.get("success"):
        return {
            "ok": False,
            "error": str(result),
            "raw": result
        }

    return {
        "ok": True,
        "raw": result
    }


def kiem_tra_sau_update(session, tinh_hinh_dang_ky_id, mo_ta_moi):
    res_hoso = api_get_hosoquet(session, tinh_hinh_dang_ky_id)

    if not res_hoso.get("ok"):
        return False, "Không load lại được hồ sơ quét: " + res_hoso.get("error", "")

    raw = res_hoso["raw"]
    value = raw.get("Value") or raw.get("value") or {}
    list_info = value.get("ListHoSoQuetInfo") or []
    target = mo_ta_moi.upper()

    for info in list_info:
        for hoso in info.get("ListHoSoQuet") or []:
            wrapper = hoso.get("ListFileHoSoQuet") or {}
            files = wrapper.get("ListFileHoSoQuet") or []

            for f in files:
                mo_ta = (f.get("moTa") or "").upper()
                if target in mo_ta:
                    return True, "Đã thấy mô tả mới trong hồ sơ quét"

    return False, "Chưa thấy mô tả mới sau cập nhật"


# =========================
# XỬ LÝ 1 DÒNG
# =========================

def xu_ly_1_dong(session, item, maxa, folder_upload):
    row_excel = item["row"]
    soto = item["soto"]
    sothua = item["sothua"]
    loaidat = item["loaidat"]
    tenfile = item["tenfile"]

    file_path = lay_duong_dan_file(folder_upload, tenfile)

    # Theo yêu cầu của mày:
    mo_ta_moi = f"CHUACOGIAY_{maxa}_{loaidat}-DDK"

    result_row = {
        "row_excel": row_excel,
        "soto": soto,
        "sothua": sothua,
        "loaidat": loaidat,
        "tenfile": tenfile,
        "mo_ta": mo_ta_moi,
        "tinhHinhDangKyId": "",
        "hoSoQuetId": "",
        "thongTinHoSoId": "",
        "status": "Lỗi",
        "note": ""
    }

    print("=" * 90)
    print(f"▶️ Dòng Excel {row_excel}: tờ={soto}, thửa={sothua}, loại đất={loaidat}, file={tenfile}")
    print(f"📌 Mô tả mới: {mo_ta_moi}")

    if not os.path.isfile(file_path):
        result_row["note"] = f"Không tìm thấy file: {file_path}"
        print("❌", result_row["note"])
        return result_row

    # 1. Tìm tình hình đăng ký
    res_search = api_tim_to_thua(
        session=session,
        xa_id=maxa,
        so_to=soto,
        so_thua=sothua
    )

    if not res_search.get("ok"):
        result_row["note"] = "Tra cứu lỗi: " + res_search.get("error", "")
        print("❌", result_row["note"])
        return result_row

    tinh_hinh_id = res_search["tinhHinhDangKyId"]
    result_row["tinhHinhDangKyId"] = tinh_hinh_id

    print("✅ tinhHinhDangKyId:", tinh_hinh_id)

    # 2. Load hồ sơ quét
    res_hoso = api_get_hosoquet(session, tinh_hinh_id)

    if not res_hoso.get("ok"):
        result_row["note"] = "Load hồ sơ quét lỗi: " + res_hoso.get("error", "")
        print("❌", result_row["note"])
        return result_row

    found = tim_hosoquet_chuacogiay(res_hoso["raw"])

    if not found:
        result_row["note"] = "Không tìm thấy hồ sơ quét có CHUACOGIAY"
        print("⚠️", result_row["note"])
        return result_row

    info = found["info"]
    hoso = found["hoso"]

    ho_so_quet_id = hoso.get("hoSoQuetId") or hoso.get("Title")
    thong_tin_ho_so_id = hoso.get("thongTinHoSoId") or info.get("thongTinHoSoId")

    result_row["hoSoQuetId"] = ho_so_quet_id
    result_row["thongTinHoSoId"] = thong_tin_ho_so_id

    print("✅ hoSoQuetId:", ho_so_quet_id)
    print("✅ thongTinHoSoId:", thong_tin_ho_so_id)

    if DRY_RUN:
        result_row["status"] = "DRY_RUN"
        result_row["note"] = "Chỉ kiểm tra, chưa cập nhật thật"
        print("🧪 DRY_RUN=True, bỏ qua update thật.")
        return result_row

    # 3. Update file + mô tả
    res_update = api_update_hosoquet_exist_file(
        session=session,
        file_path=file_path,
        found_hosoquet=found,
        mo_ta_moi=mo_ta_moi
    )

    if not res_update.get("ok"):
        result_row["note"] = "Update lỗi: " + res_update.get("error", "")
        print("❌", result_row["note"])
        return result_row

    print("✅ UpdateHoSoQuetExistFile thành công.")

    # 4. Kiểm tra lại
    ok_check, note_check = kiem_tra_sau_update(
        session=session,
        tinh_hinh_dang_ky_id=tinh_hinh_id,
        mo_ta_moi=mo_ta_moi
    )

    if ok_check:
        result_row["status"] = "Thành công"
        result_row["note"] = note_check
        print("✅", note_check)
    else:
        result_row["status"] = "Cần kiểm tra"
        result_row["note"] = note_check
        print("⚠️", note_check)

    return result_row


# =========================
# WORKER
# =========================

def worker_run(username, password, maxa, excel_path, folder_upload, log_queue):
    old_stdout = sys.stdout
    sys.stdout = TextRedirector(log_queue)

    driver = None

    try:
        data = doc_excel(excel_path)

        if not data:
            print("❌ Excel không có dữ liệu.")
            return

        print(f"✅ Đã đọc Excel: {len(data)} dòng.")
        print("DRY_RUN =", DRY_RUN)

        output_path = os.path.join(
            os.path.dirname(excel_path),
            "ket_qua_cap_nhat_hosoquet_api.xlsx"
        )

        driver = login_mplis(username, password)
        session = tao_session_tu_selenium(driver)

        results = []

        tong = len(data)
        thanh_cong = 0
        that_bai = 0

        for idx, item in enumerate(data, start=1):
            print(f"\n🚀 Đang xử lý {idx}/{tong}")

            try:
                kq = xu_ly_1_dong(
                    session=session,
                    item=item,
                    maxa=maxa,
                    folder_upload=folder_upload
                )
            except Exception as e:
                traceback.print_exc()
                kq = {
                    "row_excel": item.get("row"),
                    "soto": item.get("soto"),
                    "sothua": item.get("sothua"),
                    "loaidat": item.get("loaidat"),
                    "tenfile": item.get("tenfile"),
                    "mo_ta": "",
                    "tinhHinhDangKyId": "",
                    "hoSoQuetId": "",
                    "thongTinHoSoId": "",
                    "status": "Lỗi ngoài",
                    "note": str(e)
                }

            results.append(kq)

            if kq.get("status") == "Thành công":
                thanh_cong += 1
            else:
                that_bai += 1

            # Ghi sau mỗi dòng để không mất kết quả nếu lỗi giữa chừng
            ghi_excel_output(results, output_path)

        print("=" * 90)
        print(f"🎯 XONG. Thành công: {thanh_cong} | Lỗi/Bỏ qua: {that_bai} | Tổng: {tong}")
        print("📄 File kết quả:", output_path)

    except Exception as e:
        print("❌ Lỗi chương trình:", e)
        traceback.print_exc()

    finally:
        sys.stdout = old_stdout

        if driver:
            try:
                driver.quit()
            except Exception:
                pass

        log_queue.put("__DONE__")


# =========================
# TKINTER APP
# =========================

class App(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Cập nhật hồ sơ quét bằng API")
        self.geometry("980x680")

        self.log_queue = queue.Queue()
        self.worker_thread = None

        self.var_username = tk.StringVar()
        self.var_password = tk.StringVar()
        self.var_maxa = tk.StringVar(value="24121")
        self.var_excel = tk.StringVar()
        self.var_folder = tk.StringVar()

        self.create_widgets()
        self.after(200, self.process_log_queue)

    def create_widgets(self):
        frame_top = ttk.LabelFrame(self, text="Thông tin chạy")
        frame_top.pack(fill="x", padx=10, pady=10)

        ttk.Label(frame_top, text="Username").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(frame_top, textvariable=self.var_username, width=30).grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(frame_top, text="Password").grid(row=0, column=2, padx=5, pady=5, sticky="w")
        ttk.Entry(frame_top, textvariable=self.var_password, width=30, show="*").grid(row=0, column=3, padx=5, pady=5)

        ttk.Label(frame_top, text="Mã xã").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(frame_top, textvariable=self.var_maxa, width=30).grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(frame_top, text="File Excel").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(frame_top, textvariable=self.var_excel, width=85).grid(row=2, column=1, columnspan=3, padx=5, pady=5, sticky="we")
        ttk.Button(frame_top, text="Duyệt Excel", command=self.browse_excel).grid(row=2, column=4, padx=5, pady=5)

        ttk.Label(frame_top, text="Folder PDF").grid(row=3, column=0, padx=5, pady=5, sticky="w")
        ttk.Entry(frame_top, textvariable=self.var_folder, width=85).grid(row=3, column=1, columnspan=3, padx=5, pady=5, sticky="we")
        ttk.Button(frame_top, text="Duyệt Folder", command=self.browse_folder).grid(row=3, column=4, padx=5, pady=5)

        self.btn_start = ttk.Button(frame_top, text="BẮT ĐẦU CHẠY", command=self.start_run)
        self.btn_start.grid(row=4, column=1, padx=5, pady=10, sticky="w")

        self.btn_clear = ttk.Button(frame_top, text="Xóa log", command=self.clear_log)
        self.btn_clear.grid(row=4, column=2, padx=5, pady=10, sticky="w")

        note = (
            "Excel cần cột: soto, sothua, loaidat, tenfile | "
            "Mô tả mới = CHUACOGIAY_{mã xã}_{loaidat}-DDK"
        )
        ttk.Label(frame_top, text=note, foreground="blue").grid(
            row=5, column=0, columnspan=5, padx=5, pady=5, sticky="w"
        )

        frame_top.columnconfigure(3, weight=1)

        frame_log = ttk.LabelFrame(self, text="Log xử lý")
        frame_log.pack(fill="both", expand=True, padx=10, pady=10)

        self.txt_log = tk.Text(frame_log, wrap="word")
        self.txt_log.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(frame_log, command=self.txt_log.yview)
        scrollbar.pack(side="right", fill="y")

        self.txt_log.configure(yscrollcommand=scrollbar.set)

    def browse_excel(self):
        path = filedialog.askopenfilename(
            title="Chọn file Excel",
            filetypes=[
                ("Excel files", "*.xlsx *.xlsm"),
                ("All files", "*.*")
            ]
        )

        if path:
            self.var_excel.set(path)

    def browse_folder(self):
        folder = filedialog.askdirectory(title="Chọn folder chứa PDF")
        if folder:
            self.var_folder.set(folder)

    def clear_log(self):
        self.txt_log.delete("1.0", tk.END)

    def validate_input(self):
        username = self.var_username.get().strip()
        password = self.var_password.get().strip()
        maxa = self.var_maxa.get().strip()
        excel = self.var_excel.get().strip()
        folder = self.var_folder.get().strip()

        if not username:
            messagebox.showerror("Thiếu thông tin", "Chưa nhập username.")
            return False

        if not password:
            messagebox.showerror("Thiếu thông tin", "Chưa nhập password.")
            return False

        if not maxa:
            messagebox.showerror("Thiếu thông tin", "Chưa nhập mã xã.")
            return False

        if not os.path.isfile(excel):
            messagebox.showerror("Sai đường dẫn", "File Excel không tồn tại.")
            return False

        if not os.path.isdir(folder):
            messagebox.showerror("Sai đường dẫn", "Folder PDF không tồn tại.")
            return False

        return True

    def start_run(self):
        if self.worker_thread and self.worker_thread.is_alive():
            messagebox.showwarning("Đang chạy", "Chương trình đang chạy.")
            return

        if not self.validate_input():
            return

        self.btn_start.config(state="disabled")

        username = self.var_username.get().strip()
        password = self.var_password.get().strip()
        maxa = self.var_maxa.get().strip()
        excel = self.var_excel.get().strip()
        folder = self.var_folder.get().strip()

        self.worker_thread = threading.Thread(
            target=worker_run,
            args=(username, password, maxa, excel, folder, self.log_queue),
            daemon=True
        )

        self.worker_thread.start()

    def process_log_queue(self):
        try:
            while True:
                msg = self.log_queue.get_nowait()

                if msg == "__DONE__":
                    self.btn_start.config(state="normal")
                    self.txt_log.insert(tk.END, "\n✅ Tiến trình đã kết thúc.\n")
                    self.txt_log.see(tk.END)
                    continue

                self.txt_log.insert(tk.END, msg + "\n")
                self.txt_log.see(tk.END)

        except queue.Empty:
            pass

        self.after(200, self.process_log_queue)


if __name__ == "__main__":
    app = App()
    app.mainloop()