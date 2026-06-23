import json
import threading
import traceback

import requests
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.chrome.options import Options


REFERER_URL_DEFAULT = "https://dla.mplis.gov.vn/dc/DonDangKy/KeKhaiDangKyV2"
BASE_URL = "https://dla.mplis.gov.vn"
TIMEOUT = 120


def norm(v):
    return str(v).strip().lower() if v is not None else ""


def val(row, headers, names, default=""):
    for name in names:
        if name.lower() in headers:
            v = row[headers[name.lower()] - 1]
            return default if v is None else str(v).strip()
    return default


def to_float(v):
    return float(str(v).strip().replace(",", "."))


def to_int(v):
    return int(float(str(v).strip()))


def read_excel(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h:
            headers[norm(h)] = c

    required = ["soto", "sothua", "dientich", "loaidat", "nguongocid", "diachi", "maxa"]
    missing = [x for x in required if x not in headers]
    if missing:
        raise RuntimeError("Thiếu cột Excel: " + ", ".join(missing))

    rows = []
    for r in range(2, ws.max_row + 1):
        raw = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if all(x is None or str(x).strip() == "" for x in raw):
            continue

        loaidat = val(raw, headers, ["loaidat"]).upper()
        tenloaidat = val(raw, headers, ["tenloaidat"], loaidat)

        rows.append({
            "row_excel": r,
            "soto": to_int(val(raw, headers, ["soto"])),
            "sothua": to_int(val(raw, headers, ["sothua"])),
            "dientich": to_float(val(raw, headers, ["dientich"])),
            "loaidat": loaidat,
            "tenloaidat": tenloaidat,
            "motaloaidat": val(raw, headers, ["motaloaidat"], f"{loaidat} ({tenloaidat})"),
            "nguongocid": val(raw, headers, ["nguongocid"], "6"),
            "diachi": val(raw, headers, ["diachi"]),
            "maxa": val(raw, headers, ["maxa"]),
            "huyenid": val(raw, headers, ["huyenid"], "0"),
            "tinhid": val(raw, headers, ["tinhid"], "66"),
            "thoihan": val(raw, headers, ["thoihan"], "Lâu dài"),
        })

    return rows


def get_token(driver):
    js = """
    return (
        document.querySelector('input[name="__RequestVerificationToken"]')?.value ||
        document.querySelector('input[name="__requestverificationtoken"]')?.value ||
        document.querySelector('meta[name="__RequestVerificationToken"]')?.content ||
        document.querySelector('meta[name="__requestverificationtoken"]')?.content ||
        document.querySelector('meta[name="RequestVerificationToken"]')?.content ||
        ''
    );
    """
    return driver.execute_script(js)


def get_cookie_raw(driver):
    return "; ".join([f"{c['name']}={c['value']}" for c in driver.get_cookies()])


def create_session(driver, referer, log):
    token = get_token(driver)
    cookie_raw = get_cookie_raw(driver)
    ua = driver.execute_script("return navigator.userAgent;")

    log("=" * 100)
    log("TOKEN:")
    log(token or "<RỖNG>")
    log("-" * 100)
    log("COOKIE:")
    log(cookie_raw or "<RỖNG>")
    log("=" * 100)

    if not token:
        raise RuntimeError("Không lấy được token.")
    if not cookie_raw:
        raise RuntimeError("Không lấy được cookie.")

    s = requests.Session()
    s.headers.update({
        "User-Agent": ua,
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "X-Requested-With": "XMLHttpRequest",
        "Origin": BASE_URL,
        "Referer": referer,
        "__requestverificationtoken": token,
        "__RequestVerificationToken": token,
        "RequestVerificationToken": token,
    })

    for c in driver.get_cookies():
        s.cookies.set(c["name"], c["value"], domain=c.get("domain"), path=c.get("path", "/"))

    return s


def build_payload(x):
    dt = x["dientich"]

    return {
        "soThuTuThua": x["sothua"],
        "soHieuToBanDo": x["soto"],
        "dienTich": dt,
        "dienTichPhapLy": str(dt),
        "soThuTuThuaCu": "",
        "soHieuToBanDoCu": "",
        "loaiThuaDat": "",
        "quaTrinhSuDung": "",
        "inSoLieuCu": False,
        "lichSuHinhThanh": "",
        "noiDungQuyHoach": "",
        "ghiChuDienTich": "",
        "ListMucDichSuDung": [
            {
                "loaiMucDichSuDungId": x["loaidat"],
                "dienTich": dt,
                "soThuTu": 1,
                "ngayHinhThanh": None,
                "ngaySuDung": None,
                "loaiMucDichSuDungQuyHoachId": "0",
                "mucDichSuDungChiTiet": "",
                "thoiHanSuDung": x["thoihan"],
                "loaiMucDichSuDungPhuId": "0",
                "ghiChu": "",
                "ListNguonGocSuDungDat": [
                    {
                        "loaiNguonGocSuDungDatId": str(x["nguongocid"]),
                        "loaiNguonGocChuyenQuyenId": "0",
                        "dienTich": dt,
                        "chiTiet": ""
                    }
                ],
                "LoaiMucDichSuDung": {
                    "loaiMucDichSuDungId": x["loaidat"],
                    "kyHieuLoaiMucDichSuDung": None,
                    "tenLoaiMucDichSuDung": x["tenloaidat"],
                    "moTaLoaiMucDichSuDung": x["motaloaidat"],
                    "trangThai": True
                },
                "LoaiMucDichSuDungPhu": None
            }
        ],
        "ListDiaChi": [
            {
                "tinhId": str(x["tinhid"]),
                "huyenId": str(x["huyenid"]),
                "xaId": str(x["maxa"]),
                "duongId": "",
                "ngoPho": "",
                "soNha": "",
                "toDanPhoId": "",
                "laDiaChiChinh": False,
                "laDiaChiCu": False,
                "diaChiChiTiet": x["diachi"]
            }
        ],
        "TaiLieuDoDac": None,
        "diaChi": x["diachi"],
        "tinhId": int(float(x["tinhid"])),
        "huyenId": int(float(x["huyenid"])),
        "xaId": str(x["maxa"]),
        "duongDanSoDo": None,
        "tenFileSoDo": None
    }


def ok_response(js):
    if not isinstance(js, dict):
        return False
    return (
        js.get("success") is True
        or js.get("Success") is True
        or js.get("ok") is True
        or ("Value" in js and js.get("success") is not False)
        or ("value" in js and js.get("success") is not False)
    )


def post_auto(session, url, payload):
    """
    Cố ý đơn giản:
    1) Gửi JSON raw trước.
    2) Nếu không success thì thử form-urlencoded với key 'thuaDat'.
    3) Nếu vẫn không success thì thử form-urlencoded với key 'model'.

    User không cần chọn cách gửi nữa.
    """
    attempts = []

    headers_json = dict(session.headers)
    headers_json["Content-Type"] = "application/json; charset=UTF-8"
    attempts.append(("JSON raw", lambda: session.post(
        url,
        data=json.dumps(payload, ensure_ascii=False),
        headers=headers_json,
        timeout=TIMEOUT
    )))

    headers_form = dict(session.headers)
    headers_form["Content-Type"] = "application/x-www-form-urlencoded; charset=UTF-8"

    attempts.append(("FORM thuaDat", lambda: session.post(
        url,
        data={"thuaDat": json.dumps(payload, ensure_ascii=False)},
        headers=headers_form,
        timeout=TIMEOUT
    )))

    attempts.append(("FORM model", lambda: session.post(
        url,
        data={"model": json.dumps(payload, ensure_ascii=False)},
        headers=headers_form,
        timeout=TIMEOUT
    )))

    last = None
    for name, fn in attempts:
        res = fn()
        last = (name, res)
        try:
            js = res.json()
        except Exception:
            js = None

        if res.ok and ok_response(js):
            return name, res, js

    name, res = last
    try:
        js = res.json()
    except Exception:
        js = None
    return name, res, js


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("ADD THỬA ĐẤT EXCEL - BẢN ĐƠN GIẢN")
        self.geometry("1080x720")

        self.driver = None
        self.session = None

        self.var_login_url = tk.StringVar(value=REFERER_URL_DEFAULT)
        self.var_add_url = tk.StringVar()
        self.var_excel = tk.StringVar()
        self.var_dry = tk.BooleanVar(value=True)

        self.ui()

    def ui(self):
        f = ttk.LabelFrame(self, text="Cấu hình")
        f.pack(fill="x", padx=10, pady=10)

        ttk.Label(f, text="URL đăng nhập").grid(row=0, column=0, padx=5, pady=4, sticky="w")
        ttk.Entry(f, textvariable=self.var_login_url, width=110).grid(row=0, column=1, columnspan=3, padx=5, pady=4, sticky="we")

        ttk.Label(f, text="URL ADD thửa").grid(row=1, column=0, padx=5, pady=4, sticky="w")
        ttk.Entry(f, textvariable=self.var_add_url, width=110).grid(row=1, column=1, columnspan=3, padx=5, pady=4, sticky="we")

        ttk.Label(f, text="Excel").grid(row=2, column=0, padx=5, pady=4, sticky="w")
        ttk.Entry(f, textvariable=self.var_excel, width=90).grid(row=2, column=1, padx=5, pady=4, sticky="we")
        ttk.Button(f, text="Duyệt", command=self.browse).grid(row=2, column=2, padx=5, pady=4)

        ttk.Checkbutton(f, text="DRY RUN - chỉ in payload, chưa gửi", variable=self.var_dry).grid(row=3, column=1, padx=5, pady=4, sticky="w")

        ttk.Button(f, text="1. Mở Chrome đăng nhập", command=self.login).grid(row=4, column=0, padx=5, pady=8)
        ttk.Button(f, text="2. Lấy token + cookie", command=self.get_session).grid(row=4, column=1, padx=5, pady=8, sticky="w")
        ttk.Button(f, text="3. ADD từ Excel", command=self.run_thread).grid(row=4, column=2, padx=5, pady=8, sticky="w")
        ttk.Button(f, text="Xóa log", command=lambda: self.txt.delete("1.0", tk.END)).grid(row=4, column=3, padx=5, pady=8)

        ttk.Label(
            f,
            text="Excel tối thiểu: soto | sothua | dientich | loaidat | nguongocid | diachi | maxa",
            foreground="blue"
        ).grid(row=5, column=0, columnspan=4, padx=5, pady=4, sticky="w")

        f.columnconfigure(1, weight=1)

        lf = ttk.LabelFrame(self, text="Log")
        lf.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        self.txt = tk.Text(lf, wrap="word", font=("Consolas", 9))
        self.txt.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(lf, command=self.txt.yview)
        sb.pack(side="right", fill="y")
        self.txt.config(yscrollcommand=sb.set)

    def log(self, s):
        self.txt.insert(tk.END, str(s) + "\n")
        self.txt.see(tk.END)
        self.update_idletasks()

    def browse(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm"), ("All files", "*.*")])
        if p:
            self.var_excel.set(p)

    def login(self):
        try:
            if self.driver:
                try:
                    self.driver.quit()
                except Exception:
                    pass
            options = Options()
            options.add_argument("--start-maximized")
            self.driver = webdriver.Chrome(options=options)
            self.driver.get(self.var_login_url.get().strip())
            self.log("Đã mở Chrome. Đăng nhập xong thì bấm nút 2.")
        except Exception as e:
            messagebox.showerror("Lỗi", str(e))
            self.log(traceback.format_exc())

    def get_session(self):
        try:
            if not self.driver:
                raise RuntimeError("Chưa mở Chrome.")
            self.session = create_session(self.driver, self.var_login_url.get().strip(), self.log)
            messagebox.showinfo("OK", "Đã lấy token + cookie, xem trong log.")
        except Exception as e:
            messagebox.showerror("Lỗi", str(e))
            self.log(traceback.format_exc())

    def run_thread(self):
        threading.Thread(target=self.run, daemon=True).start()

    def run(self):
        try:
            if not self.session:
                raise RuntimeError("Chưa lấy token + cookie.")
            if not self.var_add_url.get().strip().startswith("http"):
                raise RuntimeError("Chưa nhập URL ADD thửa.")
            if not self.var_excel.get().strip():
                raise RuntimeError("Chưa chọn Excel.")

            rows = read_excel(self.var_excel.get().strip())
            self.log("=" * 100)
            self.log(f"Đọc Excel: {len(rows)} dòng")
            self.log(f"DRY_RUN={self.var_dry.get()}")
            self.log("=" * 100)

            ok = fail = 0

            for i, row in enumerate(rows, 1):
                payload = build_payload(row)
                self.log(f"\nDòng {row['row_excel']} - {i}/{len(rows)}: tờ {row['soto']} thửa {row['sothua']}")
                self.log(json.dumps(payload, ensure_ascii=False, indent=2))

                with open(f"debug_payload_add_thua_{row['row_excel']}.json", "w", encoding="utf-8") as f:
                    json.dump(payload, f, ensure_ascii=False, indent=2)

                if self.var_dry.get():
                    self.log("DRY RUN: chưa gửi.")
                    continue

                mode, res, js = post_auto(self.session, self.var_add_url.get().strip(), payload)

                self.log(f"Cách gửi đã dùng cuối: {mode}")
                self.log(f"HTTP {res.status_code}")
                self.log(res.text[:3000])

                with open(f"debug_response_add_thua_{row['row_excel']}.txt", "w", encoding="utf-8") as f:
                    f.write(res.text)

                if ok_response(js):
                    ok += 1
                    self.log("✅ Thành công.")
                else:
                    fail += 1
                    self.log("❌ Chưa thành công.")

            self.log("=" * 100)
            self.log(f"XONG: Thành công {ok} | Lỗi {fail} | Tổng {len(rows)}")
        except Exception as e:
            messagebox.showerror("Lỗi", str(e))
            self.log(traceback.format_exc())


if __name__ == "__main__":
    App().mainloop()
